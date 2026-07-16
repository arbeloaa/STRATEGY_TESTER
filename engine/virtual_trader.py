"""
virtual_trader.py  --  Virtual Portfolio Simulator V5
====================================================
Reads gate_report_latest.json from tester.py, opens positions on all
BUY signals, tracks each position daily, and exits via:

  EXIT TRIGGER 1 -- Hard Stop Loss (capital protection)
    HIGH confidence : -25% below entry  (V5: was -20%)
    MED  confidence : -20% below entry  (V5: was -16%)
    LOW  confidence : -16% below entry  (V5: was -14%)

  EXIT TRIGGER 2 -- Trailing Stop (lock in gains)
    Once a position is up +8%, trailing stop activates at -15% from peak (V5: was +5%/-12%)
    Once a position is up +20%, trailing stop tightens to -11% from peak (V5: was +15%/-9%)
    Once a position is up +50%, trailing stop tightens to -7% from peak (V5: was +40%/-6%)

  EXIT TRIGGER 3 -- Gate Deterioration (fundamental exit)
    Re-score stock every GATE_RECHECK_DAYS days using latest price momentum.
    If score drops below threshold: SELL (thesis broke).
    NOTE: full re-fetch is expensive; we use price-based proxy (MA crossdown).

  EXIT TRIGGER 4 -- Time Stop (avoid dead money)
    If position is between -2% and +2% after TIME_STOP_DAYS days: SELL.

  EXIT TRIGGER 5 -- Drawdown Duration (V3)
    If position has been below -8% for 60+ consecutive trading days: SELL.

POSITION SIZING:
  Base = portfolio_value / n_positions  (equal weight, max 10% per stock)
  HIGH confidence + score > thr+2.0 : 1.5x base
  HIGH confidence                   : 1.25x base
  MED  confidence                   : 1.0x base
  LOW  confidence                   : 0.6x base
  All sizes normalized to sum = 100% of portfolio.

OUTPUT:
  - Console report with per-position P&L and portfolio summary
  - virtual_trader_report_<timestamp>.json
  - virtual_trader_report_<timestamp>.xlsx  (if openpyxl available)

Usage:
  python virtual_trader.py                          # use gate_report_latest.json
  python virtual_trader.py --json path/to/file.json
  python virtual_trader.py --capital 100000         # custom starting capital ($)
  python virtual_trader.py --no-excel               # skip Excel output
"""

import os, sys, re, json, time, argparse, math
from datetime import datetime, timedelta
from pathlib import Path

import yfinance as yf
import pandas as pd

# -- Try optional imports ---------------------------------------------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ---------------------------------------------
#  SETTINGS -- edit these
# ---------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[1]
BASE_DIR     = PROJECT_ROOT / "scripts"
DEFAULT_JSON = PROJECT_ROOT / "reports" / "gate_report_latest.json"
OUTPUT_DIR   = PROJECT_ROOT / "reports"

STARTING_CAPITAL   = 100_000   # USD

# Stop-loss per confidence tier (negative = below entry)
# V5: widened all stops significantly -- V4 HARD_STOP_MED killed NVDA(-30%), ON(-30%),
# WEX(-28.5%), AVGO(-27.5%), GFS(-27%) within 68-96 days.
# These are all MED conf semis that recovered after drawdown in BULL_WEAK.
# Wider stops let cyclical positions ride through vol without premature exit.
STOP_LOSS = {
    "HIGH": -0.25,   # V5: was -0.20 -- NVDA/AVGO type need room
    "MED":  -0.20,   # V5: was -0.16 -- ON(-30%), WEX(-28.5%) all hit old -16% stop
    "LOW":  -0.16,   # V5: was -0.14
    "VETO": -0.05,   # fallback
}

# Trailing stop: V5 reworked -- V4 trail@+5%/-12% activated too early on small bounces,
# then locked in tiny gains while stocks kept running. 23 TRAIL_STOP exits with many
# leaving gains on the table. Raise activation to +8% and widen initial trail to -15%.
TRAIL_ACTIVATE_PCT   = 0.08    # V5: was 0.05 -- don't activate on noise bounces
TRAIL_STOP_PCT       = -0.15   # V5: was -0.12 -- much wider trail so runners aren't cut
TRAIL_TIGHTEN_PCT    = 0.20    # V5: was 0.15 -- tighten later, only after solid gain
TRAIL_TIGHT_STOP_PCT = -0.11   # V5: was -0.09 -- wider than V4 to avoid whipsaw
TRAIL_ULTRA_PCT      = 0.50    # V5: was 0.40 -- only tighten for truly big winners
TRAIL_ULTRA_STOP_PCT = -0.07   # V5: was -0.06 -- lock in bulk of gains on huge winners

# Time stop: sell if flat after this many calendar days
TIME_STOP_DAYS     = 210   # V5: was 180 -- extend further; BULL_WEAK recoveries take time
TIME_STOP_BAND     = 0.02  # same as V4

# Drawdown duration stop
# V5: wider floor and longer duration -- V4's -5%/50d was causing exits on normal
# semi cyclical drawdowns (NVDA, ON, GFS all had extended drawdowns before recovery)
DRAWDOWN_FLOOR_PCT      = -0.08   # V5: was -0.05 -- don't count shallow dips as "drawdown"
DRAWDOWN_DURATION_DAYS  = 60      # V5: was 50 -- give more recovery time

# Gate deterioration proxy: price crosses below N-day MA
# V5: MA100 cross caused 3 exits in V4; require 5-day confirm to reduce false signals
MA_DETERIORATION_DAYS = 100  # same as V4
MA_CONFIRM_DAYS       = 5    # V5: was 3 -- require 5 consecutive days below MA100

# BEAR_CONFIRMED: V5 adjustments -- V4 still had premature bear exits
BEAR_CONFIRMED_DROP   = -0.20  # V5: was -0.18 -- require even deeper NDX drop
BEAR_ALERT_DROP       = -0.14  # V5: was -0.12 -- widen alert threshold
BEAR_MA200_THRESHOLD  = 0.84   # V5: was 0.86 -- NDX must be >16% below MA200

# V5: extend grace periods -- V4 had BEAR_CONFIRMED exits on stocks that recovered
BEAR_GRACE_DAYS_HIGH  = 120    # V5: was 90 -- HIGH conf gets full quarter of grace
BEAR_GRACE_DAYS_MED   = 75     # V5: was 45 -- MED conf also gets substantial grace

# Position size multipliers by confidence
SIZE_MULT = {
    "HIGH_PLUS": 1.5,   # HIGH confidence AND score > threshold + 2.0
    "HIGH":      1.25,
    "MED":       1.0,
    "LOW":       0.6,
}
MAX_POSITION_PCT = 0.10    # cap any single position at 10% of portfolio

# Fetch price throttle (seconds between yfinance calls)
FETCH_DELAY = 0.3

# After BEAR_ALERT fires (any position), pause entering new positions for this many trading days.
# Set to 0 to disable. Positions skipped show exit_reason = "SKIPPED_BEAR_ALERT".
PAUSE_ENTRIES_ON_BEAR_ALERT = 20

# V5: BEAR_ALERT stop tightening less aggressive
BEAR_ALERT_STOP_MULT = 0.75  # V5: was 0.70 -- tighten stop to 75% of original (not 70%)

# V5: Regime-aware stop widening for BULL_WEAK
# Evidence: BULL_WEAK regime had outsized losses from premature exits
# NVDA, ON, WEX, AVGO, GFS all exited via HARD_STOP_MED in BULL_WEAK
BULL_WEAK_STOP_MULT = 1.20   # V5: was 1.15 -- stops are 20% wider in BULL_WEAK

# V5: NEW -- Universe-aware stop adjustment
# Semi stocks are more volatile (cyclical), need wider stops
# Evidence: 5 worst positions (NVDA -30%, ON -30%, WEX -28.5%, AVGO -27.5%, GFS -27%)
# are ALL semi universe stocks that hit HARD_STOP_MED
SEMI_STOP_MULT = 1.15  # V5: NEW -- semi stocks get 15% wider stops on top of regime mult

# V5: NEW -- Score-margin stop adjustment
# Stocks with score >> threshold have stronger fundamentals, deserve wider stops
# Evidence: AVGO (score well above threshold) hit hard stop despite strong gates
SCORE_MARGIN_STOP_BONUS = 0.02  # V5: NEW -- +2% wider stop per 1.0 score above threshold, capped

# ---------------------------------------------
#  COLORS
# ---------------------------------------------
GREEN   = "\033[92m"
RED     = "\033[91m"
YELLOW  = "\033[93m"
CYAN    = "\033[96m"
MAGENTA = "\033[95m"
BOLD    = "\033[1m"
RESET   = "\033[0m"

def log(msg, color=""):
    codes = {"green":GREEN,"red":RED,"yellow":YELLOW,
             "cyan":CYAN,"bold":BOLD,"reset":RESET}
    c  = codes.get(color, "")
    rs = RESET if color else ""
    print(f"{c}{msg}{rs}", flush=True)

# ---------------------------------------------
#  PRICE FETCHING
# ---------------------------------------------
def fetch_daily_prices(ticker: str, start: str, end: str) -> pd.DataFrame:
    """Fetch daily OHLCV for ticker between start and end dates."""
    try:
        t = yf.Ticker(ticker)
        df = t.history(start=start, end=end, interval="1d", auto_adjust=True)
        if df.empty:
            return pd.DataFrame()
        df.index = pd.to_datetime(df.index).tz_localize(None)
        return df[["Close", "High", "Low", "Volume"]].copy()
    except Exception as e:
        log(f"  [WARN] {ticker}: price fetch error: {e}", "yellow")
        return pd.DataFrame()

def fetch_ndx_prices(start: str, end: str) -> pd.DataFrame:
    """Fetch daily NDX (QQQ) prices for regime monitoring."""
    try:
        t  = yf.Ticker("QQQ")
        df = t.history(start=start, end=end, interval="1d", auto_adjust=True)
        if df.empty:
            return pd.DataFrame()
        df.index = pd.to_datetime(df.index).tz_localize(None)
        return df[["Close"]].copy()
    except Exception as e:
        log(f"  [WARN] NDX fetch error: {e}", "yellow")
        return pd.DataFrame()

def check_market_regime_shift(entry_ndx: float, ndx_df: pd.DataFrame,
                               current_dt: pd.Timestamp) -> str:
    """
    Check if NDX has shifted to bear regime since position entry.
    Returns 'BEAR_ALERT', 'BEAR_CONFIRMED', or 'OK'.
    V5: further widened thresholds
    """
    if ndx_df.empty or entry_ndx <= 0:
        return "OK"
    rows = ndx_df[ndx_df.index <= current_dt]
    if rows.empty:
        return "OK"
    current_ndx     = float(rows["Close"].iloc[-1])
    drop_from_entry = (current_ndx - entry_ndx) / entry_ndx
    if drop_from_entry <= BEAR_ALERT_DROP:
        if drop_from_entry <= BEAR_CONFIRMED_DROP:
            return "BEAR_CONFIRMED"
        if len(rows) >= 200:
            ma200 = rows["Close"].rolling(200).mean().iloc[-1]
            if pd.notna(ma200) and current_ndx < float(ma200) * BEAR_MA200_THRESHOLD:
                return "BEAR_CONFIRMED"
        return "BEAR_ALERT"
    return "OK"

def fetch_entry_price(ticker: str, entry_date: str) -> float | None:
    """Get the closing price on or just after entry_date."""
    try:
        start = (datetime.strptime(entry_date, "%Y-%m-%d") - timedelta(days=3)).strftime("%Y-%m-%d")
        end   = (datetime.strptime(entry_date, "%Y-%m-%d") + timedelta(days=7)).strftime("%Y-%m-%d")
        t     = yf.Ticker(ticker)
        df    = t.history(start=start, end=end, interval="1d", auto_adjust=True)
        if df.empty:
            return None
        df.index = pd.to_datetime(df.index).tz_localize(None)
        target   = pd.Timestamp(entry_date)
        future   = df[df.index >= target]
        if future.empty:
            return float(df["Close"].iloc[-1])
        return float(future["Close"].iloc[0])
    except Exception:
        return None

# ---------------------------------------------
#  POSITION SIZING
# ---------------------------------------------
def compute_position_sizes(positions: list, portfolio_value: float) -> dict:
    """
    Assign dollar allocation to each position based on confidence and score.
    Returns {ticker: dollar_amount}.
    """
    mults = {}
    for p in positions:
        conf  = p.get("confidence", "MED")
        score = p.get("weighted_score", 0)
        thr   = p.get("pass_threshold", 5.5)
        if conf == "HIGH" and score >= thr + 2.0:
            mults[p["ticker"]] = SIZE_MULT["HIGH_PLUS"]
        elif conf == "HIGH":
            mults[p["ticker"]] = SIZE_MULT["HIGH"]
        elif conf == "MED":
            mults[p["ticker"]] = SIZE_MULT["MED"]
        else:
            mults[p["ticker"]] = SIZE_MULT["LOW"]

    total_mult = sum(mults.values())
    base       = portfolio_value / total_mult

    sizes = {}
    for ticker, mult in mults.items():
        raw     = base * mult
        capped  = min(raw, portfolio_value * MAX_POSITION_PCT)
        sizes[ticker] = capped

    total = sum(sizes.values())
    scale = portfolio_value / total if total > 0 else 1.0
    for ticker in sizes:
        sizes[ticker] = round(sizes[ticker] * scale, 2)

    return sizes

# ---------------------------------------------
#  SIMULATE ONE POSITION
# ---------------------------------------------
def simulate_position(stock: dict, entry_date: str, end_date: str,
                      allocated_dollars: float,
                      ndx_df: pd.DataFrame | None = None,
                      market_regime: str = "BULL_STRONG") -> dict:
    """
    Simulate one position with daily stop-loss, trailing stop, MA exit, time stop,
    and drawdown duration stop.

    V5 changes vs V4:
    - Hard stops widened: HIGH -25% (was -20%), MED -20% (was -16%), LOW -16% (was -14%)
    - Trailing: activate at +8% (was +5%), trail -15% from peak (was -12%)
    - Three-tier trail: tighten at +20%/-11% and +50%/-7% (was +15%/-9% and +40%/-6%)
    - Drawdown duration: -8% floor for 60d (was -5%/50d)
    - MA confirm: 5 consecutive days (was 3)
    - BEAR_CONFIRMED: -20% NDX drop (was -18%), grace HIGH=120d MED=75d
    - Universe-aware: semi stocks get 15% wider stops
    - Score-margin: high-score stocks get wider stops (up to +4%)
    - BULL_WEAK regime: stops 20% wider (was 15%)
    """
    ticker     = stock["ticker"]
    confidence = stock.get("confidence", "MED")
    score      = stock.get("weighted_score", 0)
    thr        = stock.get("pass_threshold", 5.5)
    universe   = stock.get("universe", "").lower()

    time.sleep(FETCH_DELAY)
    df = fetch_daily_prices(ticker, entry_date, end_date)
    log(f"    {ticker}: fetched {len(df)} rows ({entry_date} -> {end_date[:10]})")

    if df.empty:
        try:
            t_obj  = yf.Ticker(ticker)
            cur_px = float(t_obj.fast_info["last_price"])
        except Exception:
            cur_px = None
        return {
            "ticker":             ticker,
            "sector":             stock.get("sector", "?"),
            "universe":           stock.get("universe", "?"),
            "confidence":         confidence,
            "score":              round(score, 2),
            "threshold":          round(thr, 2),
            "allocated":          round(allocated_dollars, 2),
            "shares":             round(allocated_dollars / cur_px, 4) if cur_px else 0,
            "entry_price":        round(cur_px, 2) if cur_px else None,
            "exit_price":         round(cur_px, 2) if cur_px else None,
            "entry_date":         entry_date,
            "exit_date":          entry_date,
            "days_held":          0,
            "peak_gain":          0.0,
            "pnl_pct":            0.0,
            "pnl_dollars":        0.0,
            "exit_reason":        "NO_DATA",
            "stop_pct":           STOP_LOSS.get(confidence, STOP_LOSS["MED"]) * 100,
            "trailing_activated": False,
        }

    # V5: compute effective stop with regime, universe, and score-margin adjustments
    stop_pct   = STOP_LOSS.get(confidence, STOP_LOSS["MED"])

    # Regime-aware widening
    if market_regime == "BULL_WEAK":
        stop_pct = stop_pct * BULL_WEAK_STOP_MULT  # V5: 20% wider in BULL_WEAK (was 15%)

    # V5: NEW -- Universe-aware widening for semi (cyclical, higher vol)
    if universe == "semi":
        stop_pct = stop_pct * SEMI_STOP_MULT  # 15% wider for semi stocks

    # V5: NEW -- Score-margin widening: stocks with score >> threshold get wider stops
    score_margin = max(0, score - thr)
    score_bonus = min(score_margin * SCORE_MARGIN_STOP_BONUS, 0.04)  # cap at 4% extra
    stop_pct = stop_pct - score_bonus  # stop_pct is negative, so subtract to widen

    entry_px   = float(df["Close"].iloc[0])
    stop_price = entry_px * (1 + stop_pct)
    peak_price = entry_px
    trail_active = False
    exit_price = None
    exit_date  = None
    exit_reason = None
    entry_day  = df.index[0]

    consecutive_drawdown_days = 0

    # Regime monitor init
    entry_ndx = 0.0
    if ndx_df is not None and not ndx_df.empty:
        ndx_rows = ndx_df[ndx_df.index >= entry_day]
        if not ndx_rows.empty:
            entry_ndx = float(ndx_rows["Close"].iloc[0])
    regime_tightened  = False
    effective_stop_px = stop_price
    bear_alert_date   = None

    # Compute MA for gate deterioration proxy
    if len(df) >= MA_DETERIORATION_DAYS:
        df["MA"] = df["Close"].rolling(MA_DETERIORATION_DAYS).mean()
    else:
        df["MA"] = None

    # V5: track consecutive days below MA for 5-day confirm (was 3)
    consecutive_below_ma = 0

    for i, (dt, row) in enumerate(df.iterrows()):
        price     = float(row["Close"])
        gain_pct  = (price - entry_px) / entry_px

        # Update peak
        if price > peak_price:
            peak_price = price

        # Track consecutive days below drawdown floor (V5: -8% floor, was -5%)
        if gain_pct < DRAWDOWN_FLOOR_PCT:
            consecutive_drawdown_days += 1
        else:
            consecutive_drawdown_days = 0

        # Activate trailing stop (V5: higher threshold to avoid noise)
        if not trail_active and gain_pct >= TRAIL_ACTIVATE_PCT:
            trail_active = True

        # Trailing stop check (V5: three-tier with wider trails)
        if trail_active:
            peak_gain_pct = (peak_price - entry_px) / entry_px
            if peak_gain_pct >= TRAIL_ULTRA_PCT:
                current_trail_pct = TRAIL_ULTRA_STOP_PCT   # V5: -7% from peak for big winners
            elif peak_gain_pct >= TRAIL_TIGHTEN_PCT:
                current_trail_pct = TRAIL_TIGHT_STOP_PCT   # V5: -11% from peak
            else:
                current_trail_pct = TRAIL_STOP_PCT          # V5: -15% from peak
            trail_stop = peak_price * (1 + current_trail_pct)
            if price <= trail_stop:
                exit_price  = price
                exit_date   = dt
                exit_reason = f"TRAIL_STOP (peak={peak_price:.2f}, stop={trail_stop:.2f}, trail={current_trail_pct*100:.0f}%)"
                break

        # Hard stop loss
        if price <= stop_price:
            exit_price  = max(price, stop_price)
            exit_date   = dt
            exit_reason = f"HARD_STOP_{confidence} ({stop_pct*100:.0f}%)"
            break

        # Drawdown duration stop (V5: -8% floor for 60d, was -5%/50d)
        if consecutive_drawdown_days >= DRAWDOWN_DURATION_DAYS:
            exit_price  = price
            exit_date   = dt
            exit_reason = f"DRAWDOWN_DURATION ({consecutive_drawdown_days}d below {DRAWDOWN_FLOOR_PCT*100:.0f}%)"
            break

        # Gate deterioration proxy: V5 requires MA_CONFIRM_DAYS=5 consecutive days below MA
        ma_val = row.get("MA")
        if pd.notna(ma_val) and i > MA_DETERIORATION_DAYS:
            if price < float(ma_val):
                consecutive_below_ma += 1
            else:
                consecutive_below_ma = 0
            if consecutive_below_ma >= MA_CONFIRM_DAYS:  # V5: 5 days (was 3)
                exit_price  = price
                exit_date   = dt
                exit_reason = f"MA{MA_DETERIORATION_DAYS}_CROSS ({MA_CONFIRM_DAYS}-day confirmed break)"
                break
        else:
            consecutive_below_ma = 0

        # Time stop (V5: 210d, was 180d)
        days_held = (dt - entry_day).days
        if days_held >= TIME_STOP_DAYS and abs(gain_pct) < TIME_STOP_BAND:
            exit_price  = price
            exit_date   = dt
            exit_reason = f"TIME_STOP ({TIME_STOP_DAYS}d flat within ±{TIME_STOP_BAND*100:.0f}%)"
            break

        # Regime monitor (weekly check: every 5 trading days)
        if ndx_df is not None and i % 5 == 0 and i > 0:
            regime = check_market_regime_shift(entry_ndx, ndx_df, dt)
            if regime == "BEAR_CONFIRMED":
                days_since_entry = (dt - entry_day).days
                # V5: extended grace periods
                if confidence == "HIGH" and days_since_entry < BEAR_GRACE_DAYS_HIGH:
                    if not regime_tightened:
                        regime_tightened  = True
                        bear_alert_date   = dt
                        effective_stop_px = entry_px * (1 + stop_pct * BEAR_ALERT_STOP_MULT)
                        log(f"    {ticker}: BEAR_CONFIRMED at {dt.date()} but HIGH conf grace ({days_since_entry}d < {BEAR_GRACE_DAYS_HIGH}d) -- tightening stop only", "yellow")
                elif confidence == "MED" and days_since_entry < BEAR_GRACE_DAYS_MED:
                    # V5: MED grace extended to 75d (was 45d)
                    if not regime_tightened:
                        regime_tightened  = True
                        bear_alert_date   = dt
                        effective_stop_px = entry_px * (1 + stop_pct * BEAR_ALERT_STOP_MULT)
                        log(f"    {ticker}: BEAR_CONFIRMED at {dt.date()} but MED conf grace ({days_since_entry}d < {BEAR_GRACE_DAYS_MED}d) -- tightening stop only", "yellow")
                else:
                    exit_price  = price
                    exit_date   = dt
                    exit_reason = "BEAR_CONFIRMED (NDX structural bear)"
                    break
            elif regime == "BEAR_ALERT" and not regime_tightened:
                regime_tightened  = True
                bear_alert_date   = dt
                effective_stop_px = entry_px * (1 + stop_pct * BEAR_ALERT_STOP_MULT)
                log(f"    {ticker}: BEAR_ALERT on {dt.date()} -- stop tightened to {stop_pct*BEAR_ALERT_STOP_MULT*100:.0f}%", "yellow")

        # Tightened stop check (BEAR_ALERT)
        if regime_tightened and price <= effective_stop_px:
            exit_price  = max(price, effective_stop_px)
            exit_date   = dt
            exit_reason = f"HARD_STOP_{confidence} (BEAR_ALERT tightened)"
            break

    # If no exit triggered: sell at last price
    if exit_price is None:
        exit_price  = float(df["Close"].iloc[-1])
        exit_date   = df.index[-1]
        exit_reason = "END_OF_PERIOD"

    days_held  = (exit_date - entry_day).days
    pnl_pct    = (exit_price - entry_px) / entry_px * 100
    pnl_dollars = allocated_dollars * (pnl_pct / 100)
    peak_gain   = (peak_price - entry_px) / entry_px * 100

    return {
        "ticker":       ticker,
        "sector":       stock.get("sector", "?"),
        "universe":     stock.get("universe", "?"),
        "confidence":   confidence,
        "score":        round(score, 2),
        "threshold":    round(thr, 2),
        "allocated":    round(allocated_dollars, 2),
        "shares":       round(allocated_dollars / entry_px, 4) if entry_px else 0,
        "entry_price":  round(entry_px, 2),
        "exit_price":   round(exit_price, 2),
        "entry_date":   str(entry_day.date()),
        "exit_date":    str(exit_date.date()) if exit_date else None,
        "days_held":    days_held,
        "peak_gain":    round(peak_gain, 1),
        "pnl_pct":      round(pnl_pct, 2),
        "pnl_dollars":  round(pnl_dollars, 2),
        "exit_reason":       exit_reason,
        "stop_pct":          stop_pct * 100,
        "trailing_activated": trail_active,
        "bear_alert_date":   str(bear_alert_date.date()) if bear_alert_date is not None else None,
    }

# ---------------------------------------------
#  PORTFOLIO REPORT
# ---------------------------------------------
def print_report(results: list, starting_capital: float, entry_date: str,
                 market_regime: str):
    wins   = [r for r in results if r["pnl_pct"] >= 0]
    losses = [r for r in results if r["pnl_pct"] <  0]
    total_pnl  = sum(r["pnl_dollars"] for r in results)
    final_val  = starting_capital + total_pnl
    total_pct  = total_pnl / starting_capital * 100

    avg_win  = sum(r["pnl_pct"] for r in wins)   / len(wins)   if wins   else 0
    avg_loss = sum(r["pnl_pct"] for r in losses) / len(losses) if losses else 0
    ev       = (len(wins)/len(results) * avg_win +
                len(losses)/len(results) * avg_loss) if results else 0
    pf       = (sum(r["pnl_dollars"] for r in wins) /
                abs(sum(r["pnl_dollars"] for r in losses))) if losses else float("inf")

    exit_reasons = {}
    for r in results:
        reason = r["exit_reason"].split(" ")[0]
        exit_reasons[reason] = exit_reasons.get(reason, 0) + 1

    log("\n" + "=" * 78, "bold")
    log(f"  VIRTUAL PORTFOLIO FINAL REPORT  (V5)", "bold")
    log(f"  Market Regime: {market_regime}  |  Entry Date: {entry_date}", "bold")
    log("=" * 78, "bold")
    log(f"\n  Starting Capital : ${starting_capital:>12,.2f}")
    log(f"  Final Value      : ${final_val:>12,.2f}")

    col = "green" if total_pnl >= 0 else "red"
    log(f"  Total P&L        : ${total_pnl:>+12,.2f}  ({total_pct:+.2f}%)", col)
    log(f"\n  Positions        : {len(results)}")
    log(f"  Winners          : {len(wins)}  ({len(wins)/len(results)*100:.1f}%)", "green")
    log(f"  Losers           : {len(losses)}  ({len(losses)/len(results)*100:.1f}%)", "red")
    log(f"  Avg Win          : {avg_win:+.1f}%")
    log(f"  Avg Loss         : {avg_loss:+.1f}%")
    log(f"  Expected Value   : {ev:+.2f}% per trade")
    log(f"  Profit Factor    : {pf:.2f}x")

    log(f"\n  Exit Reasons:")
    for reason, count in sorted(exit_reasons.items(), key=lambda x: -x[1]):
        log(f"    {reason:<30}: {count}")

    # Per-confidence breakdown
    log(f"\n  By Confidence Tier:", "bold")
    for conf in ["HIGH", "MED", "LOW"]:
        tier = [r for r in results if r["confidence"] == conf]
        if not tier: continue
        t_pnl = sum(r["pnl_dollars"] for r in tier)
        t_pct = sum(r["pnl_pct"] for r in tier) / len(tier)
        t_wins = sum(1 for r in tier if r["pnl_pct"] >= 0)
        col = "green" if t_pnl >= 0 else "red"
        log(f"    {conf:<6}: {len(tier):>3} positions  "
            f"avg={t_pct:+.1f}%  total=${t_pnl:>+8,.0f}  "
            f"win_rate={t_wins/len(tier)*100:.0f}%", col)

    # Per-universe breakdown
    log(f"\n  By Universe:", "bold")
    for uni in ["tech", "semi", "medtech", "energy"]:
        tier = [r for r in results if r.get("universe","").lower() == uni]
        if not tier: continue
        t_pnl = sum(r["pnl_dollars"] for r in tier)
        t_pct = sum(r["pnl_pct"] for r in tier) / len(tier)
        col = "green" if t_pnl >= 0 else "red"
        log(f"    {uni.upper():<8}: {len(tier):>3} positions  "
            f"avg={t_pct:+.1f}%  total=${t_pnl:>+8,.0f}", col)

    # Top 5 and bottom 5
    sorted_r = sorted(results, key=lambda x: x["pnl_pct"], reverse=True)
    log(f"\n  TOP 5 POSITIONS:", "green")
    for r in sorted_r[:5]:
        log(f"    {r['ticker']:<7} {r['pnl_pct']:>+7.1f}%  "
            f"${r['pnl_dollars']:>+8,.0f}  "
            f"[{r['confidence']}  score={r['score']}  "
            f"{r['days_held']}d  exit:{r['exit_reason'].split('(')[0].strip()}]", "green")

    log(f"\n  BOTTOM 5 POSITIONS:", "red")
    for r in sorted_r[-5:]:
        log(f"    {r['ticker']:<7} {r['pnl_pct']:>+7.1f}%  "
            f"${r['pnl_dollars']:>+8,.0f}  "
            f"[{r['confidence']}  score={r['score']}  "
            f"{r['days_held']}d  exit:{r['exit_reason'].split('(')[0].strip()}]", "red")

    # Max holding time across all positions
    max_days = max((r["days_held"] for r in results), default=0)
    avg_days = int(sum(r["days_held"] for r in results) / len(results)) if results else 0
    log(f"  Max Hold         : {max_days}d  (avg {avg_days}d per position)")

    log("\n" + "=" * 78, "bold")

    # Verdict
    if total_pct > 20:
        verdict = "STRONG ALPHA"
        vcol    = "green"
    elif total_pct > 5:
        verdict = "ALPHA GENERATING"
        vcol    = "green"
    elif total_pct > -5:
        verdict = "NEUTRAL (SPY-like)"
        vcol    = "yellow"
    else:
        verdict = "VALUE DESTROYING"
        vcol    = "red"
    log(f"  VERDICT: {verdict}  |  Portfolio return: {total_pct:+.2f}%  |  Max hold: {max_days}d", vcol)
    log("=" * 78, "bold")

    return total_pct, max_days

# ---------------------------------------------
#  EXCEL OUTPUT
# ---------------------------------------------
def write_excel(results: list, out_path: Path, starting_capital: float,
                total_pct: float):
    if not EXCEL_OK:
        log("  openpyxl not installed -- skipping Excel output", "yellow")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Virtual Trader Results"

    # Header
    thin  = Side(style="thin", color="BFBFBF")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr_font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cols = ["Ticker","Sector","Universe","Confidence","Score","Threshold",
            "Allocated $","Shares","Entry Price","Entry Date",
            "Exit Price","Exit Date","Days Held","Peak Gain %",
            "P&L %","P&L $","Exit Reason","Stop %","Trailing"]

    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill  = hdr_fill
        c.font  = hdr_font
        c.alignment = hdr_align
        c.border = bdr

    ws.row_dimensions[1].height = 24

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill   = PatternFill("solid", fgColor="FFC7CE")

    for ri, r in enumerate(sorted(results, key=lambda x: x["pnl_pct"], reverse=True), 2):
        fill = green_fill if r["pnl_pct"] >= 0 else red_fill
        vals = [
            r["ticker"], r["sector"], r.get("universe","?"), r["confidence"],
            r["score"], r["threshold"],
            r["allocated"], r["shares"],
            r["entry_price"], r["entry_date"],
            r["exit_price"], r["exit_date"], r["days_held"],
            r["peak_gain"],
            r["pnl_pct"], r["pnl_dollars"],
            r["exit_reason"], r["stop_pct"], r["trailing_activated"],
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill  = fill
            c.font  = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="center")
            c.border = bdr
            if ci in (7, 16):
                c.number_format = '"$"#,##0.00'
            elif ci in (14, 15):
                c.number_format = '+0.00%;-0.00%'
                if ci == 15: c.value = val / 100

        ws.row_dimensions[ri].height = 14

    # Column widths
    widths = [8,20,10,11,7,10,12,9,11,12,11,12,10,11,9,12,35,8,10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    summary = [
        ("Starting Capital",   f"${starting_capital:,.2f}"),
        ("Final Value",        f"${starting_capital * (1 + total_pct/100):,.2f}"),
        ("Total Return",       f"{total_pct:+.2f}%"),
        ("Positions",          len(results)),
        ("Winners",            sum(1 for r in results if r["pnl_pct"] >= 0)),
        ("Losers",             sum(1 for r in results if r["pnl_pct"] < 0)),
    ]
    for ri, (k, v) in enumerate(summary, 1):
        ws2.cell(ri, 1, k).font = Font(bold=True)
        ws2.cell(ri, 2, str(v))
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15

    wb.save(out_path)
    log(f"  Excel saved -> {out_path}", "green")

# ---------------------------------------------
#  MAIN
# ---------------------------------------------
def main(json_path: Path, capital: float):
    log("=" * 70, "bold")
    log("  VIRTUAL TRADER V5 -- Position Simulator", "bold")
    log(f"  Starting capital: ${capital:,.0f}", "bold")
    log(f"  V5: wider stops (HIGH={STOP_LOSS['HIGH']*100:.0f}% MED={STOP_LOSS['MED']*100:.0f}% LOW={STOP_LOSS['LOW']*100:.0f}%), "
        f"trail@{TRAIL_ACTIVATE_PCT*100:.0f}%/-{abs(TRAIL_STOP_PCT)*100:.0f}%, "
        f"bear_confirm@{BEAR_CONFIRMED_DROP*100:.0f}%, "
        f"drawdown_dur={DRAWDOWN_DURATION_DAYS}d@{DRAWDOWN_FLOOR_PCT*100:.0f}%, "
        f"MA_confirm={MA_CONFIRM_DAYS}d, "
        f"semi_mult={SEMI_STOP_MULT}x", "bold")
    log("=" * 70, "bold")

    # Load gate report
    if not json_path.exists():
        log(f"ERROR: JSON not found: {json_path}", "red")
        log("Run tester.py first to generate gate_report_latest.json", "yellow")
        sys.exit(1)

    with open(json_path, encoding="utf-8") as f:
        report = json.load(f)

    meta    = report.get("meta", {})
    stocks  = report.get("stocks", [])
    regime  = meta.get("nasdaq_regime", "UNKNOWN")

    # Filter to BUY signals only (strategy_passed=True, not vetoed)
    buys = [s for s in stocks if s.get("strategy_passed") and not s.get("veto")]
    log(f"\n  Loaded {len(stocks)} stocks from JSON")
    log(f"  BUY signals: {len(buys)}  |  Market regime: {regime}", "cyan")

    if not buys:
        log("ERROR: No BUY signals found in gate report.", "red")
        sys.exit(1)

    # Entry date = strategy START_DATE
    _raw_start = meta.get("start_date", "") or ""
    _m = re.search(r"(\d{4}-\d{2}-\d{2})", _raw_start)
    entry_date = (
        _m.group(1) if _m
        else (meta.get("generated", "")[:10] if meta.get("generated") else None)
        or datetime.today().strftime("%Y-%m-%d")
    )
    # yfinance end is EXCLUSIVE -- add 1 day so today's close is included
    end_date = (datetime.today() + timedelta(days=1)).strftime("%Y-%m-%d")

    log(f"  Gate period: {entry_date} -> {end_date[:10]}")

    # Position sizing
    log(f"\n  Computing position sizes ...", "yellow")
    sizes = compute_position_sizes(buys, capital)

    log(f"\n  POSITIONS TO OPEN:", "bold")
    log(f"  {'Ticker':<8} {'Conf':<6} {'Score':>6} {'Allocated':>12} {'Mult':>5}")
    log("  " + "-" * 45)
    for s in sorted(buys, key=lambda x: x.get("weighted_score", 0), reverse=True):
        t      = s["ticker"]
        conf   = s.get("confidence", "?")
        score  = s.get("weighted_score", 0)
        alloc  = sizes.get(t, 0)
        mult   = alloc / (capital / len(buys))
        log(f"  {t:<8} {conf:<6} {score:>6.2f} ${alloc:>11,.0f} {mult:>4.2f}x")

    # Fetch NDX once for regime monitoring
    log(f"\n  Fetching NDX (QQQ) prices for regime monitor ...", "yellow")
    ndx_df = fetch_ndx_prices(entry_date, end_date)
    if ndx_df.empty:
        log("  [WARN] NDX prices unavailable -- regime monitor disabled", "yellow")
    else:
        log(f"  NDX: {len(ndx_df)} days fetched", "cyan")

    # Simulate all positions
    log(f"\n  Simulating {len(buys)} positions (fetching daily prices) ...", "yellow")
    results = []
    earliest_bear_alert: "pd.Timestamp | None" = None
    for i, stock in enumerate(buys, 1):
        t      = stock["ticker"]
        alloc  = sizes.get(t, capital / len(buys))

        # PAUSE_ENTRIES_ON_BEAR_ALERT: skip if entry is within pause window
        if PAUSE_ENTRIES_ON_BEAR_ALERT > 0 and earliest_bear_alert is not None:
            pause_end = earliest_bear_alert + pd.Timedelta(days=PAUSE_ENTRIES_ON_BEAR_ALERT * 7 // 5)
            pos_entry = pd.Timestamp(entry_date)
            if pos_entry >= earliest_bear_alert and pos_entry <= pause_end:
                log(f"  [{i:>3}/{len(buys)}] {t:<8} SKIPPED -- BEAR_ALERT pause "
                    f"(alert={earliest_bear_alert.date()}, "
                    f"pause_end={pause_end.date()})", "yellow")
                results.append({
                    "ticker": t, "sector": stock.get("sector","?"),
                    "universe": stock.get("universe","?"),
                    "confidence": stock.get("confidence","?"),
                    "score": round(stock.get("weighted_score",0), 2),
                    "threshold": round(stock.get("pass_threshold",5.5), 2),
                    "allocated": round(alloc, 2), "shares": 0,
                    "entry_price": None, "exit_price": None,
                    "entry_date": entry_date, "exit_date": entry_date,
                    "days_held": 0, "peak_gain": 0.0,
                    "pnl_pct": 0.0, "pnl_dollars": 0.0,
                    "exit_reason": "SKIPPED_BEAR_ALERT",
                    "stop_pct": 0.0, "trailing_activated": False,
                    "bear_alert_date": str(earliest_bear_alert.date()),
                })
                continue

        log(f"  [{i:>3}/{len(buys)}] {t:<8} alloc=${alloc:>8,.0f} ...", "cyan")
        result = simulate_position(stock, entry_date, end_date, alloc,
                                   ndx_df=ndx_df if not ndx_df.empty else None,
                                   market_regime=regime)

        # Track earliest BEAR_ALERT date seen across all positions
        ba = result.get("bear_alert_date")
        if ba is not None:
            ba_ts = pd.Timestamp(ba)
            if earliest_bear_alert is None or ba_ts < earliest_bear_alert:
                earliest_bear_alert = ba_ts

        results.append(result)
        col = "green" if result["pnl_pct"] >= 0 else "red"
        log(f"           exit={result['exit_reason'].split(' ')[0]:<15} "
            f"P&L={result['pnl_pct']:>+7.1f}%  "
            f"${result['pnl_dollars']:>+8,.0f}  "
            f"({result['days_held']}d)", col)

    # Print report
    total_pct, max_days = print_report(results, capital, entry_date, regime)

    # Write vt_results_latest.json for auto_optimizer context
    exit_breakdown: dict[str, int] = {}
    for r in results:
        key = r["exit_reason"].split(" ")[0]
        exit_breakdown[key] = exit_breakdown.get(key, 0) + 1

    wins   = [r for r in results if r["pnl_pct"] >= 0]
    losses = [r for r in results if r["pnl_pct"] <  0]
    vt_json = {
        "generated":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "entry_date":     entry_date,
        "market_regime":  regime,
        "n_positions":    len(results),
        "n_wins":         len(wins),
        "n_losses":       len(losses),
        "win_rate_pct":   round(len(wins) / len(results) * 100, 1) if results else 0,
        "portfolio_return_pct": round(total_pct, 2),
        "avg_win_pct":    round(sum(r["pnl_pct"] for r in wins)   / len(wins),   2) if wins   else 0,
        "avg_loss_pct":   round(sum(r["pnl_pct"] for r in losses) / len(losses), 2) if losses else 0,
        "exit_breakdown": exit_breakdown,
        "positions":      results,
    }
    vt_out = PROJECT_ROOT / "reports" / "vt_results_latest.json"
    with open(vt_out, "w", encoding="utf-8") as f:
        json.dump(vt_json, f, indent=2, default=str)
    log(f"\n  JSON saved -> {vt_out}", "cyan")

    # Write Excel if possible
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx = OUTPUT_DIR / f"virtual_trader_report_{ts}.xlsx"
    write_excel(results, xlsx, capital, total_pct)

    return total_pct, max_days


# ---------------------------------------------
#  ENTRY POINT
# ---------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Virtual Portfolio Trader V5")
    parser.add_argument("--json",     type=Path, default=DEFAULT_JSON,
                        help="Path to gate_report_latest.json")
    parser.add_argument("--capital",  type=float, default=STARTING_CAPITAL,
                        help="Starting capital in USD (default: 100000)")
    args = parser.parse_args()

    main(json_path=args.json, capital=args.capital)