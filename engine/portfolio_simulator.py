#!/usr/bin/env python3
"""
portfolio_simulator.py  --  Cash-managed portfolio backtest
============================================================
Simulates a real trading account day-by-day using point-in-time
price/momentum data and PIT fundamentals from pit_fundamentals.csv.

Usage:
  python portfolio_simulator.py
  python portfolio_simulator.py --start 2024-04-19 --end 2026-06-13
"""

import sys, json, math, argparse, statistics, time
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd

try:
    import yfinance as yf
except ImportError:
    print("ERROR: yfinance not installed"); sys.exit(1)

# Import DB module from data_pipeline/
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "data_pipeline"))
import db as _db

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False

# ============================================================================
#  PATHS
# ============================================================================
PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR  = PROJECT_ROOT / "reports"
PERF_HIST   = PROJECT_ROOT / "logs" / "performance_history.txt"

# ============================================================================
#  PARAMS  --  loaded from strategy_params.json (single source of truth)
# ============================================================================
_PARAMS_PATH = PROJECT_ROOT / "config" / "strategy_params.json"
def _load_params():
    with open(_PARAMS_PATH, "r") as f:
        return json.load(f)
_P = _load_params()

# ============================================================================
#  CONFIG  --  auto_optimizer agents edit only this block
# ============================================================================
STARTING_CAPITAL        = 100_000          # USD
PER_BUY_FRACTION        = _P["sizing"]["per_buy_fraction"]
CONVICTION_MULT         = _P["sizing"]["conviction_mult"]
MAX_POSITION_PCT_EQUITY = _P["sizing"]["max_position_pct_equity"]
MAX_POSITIONS           = _P["sizing"]["max_positions"]
TARGET_POSITIONS        = 20              # informational
REGIME_POSITION_MULT    = _P["sizing"]["regime_position_mult"]
REGIME_MAX_POSITIONS    = _P["sizing"]["regime_max_positions"]
REGIME_EXPOSURE_CAP     = _P["sizing"].get("regime_exposure_cap", {
    "BULL_STRONG": 1.00, "BULL_WEAK": 0.85,
    "BEAR_GRIND":  0.60, "BEAR_VOLATILE": 0.40,
})
MIN_CASH_TO_TRADE       = _P["sizing"]["min_cash_to_trade"]
FRACTIONAL_SHARES       = True
ALLOW_PYRAMIDING        = _P["sizing"]["allow_pyramiding"]
MAX_ADDS_PER_POSITION   = _P["sizing"]["max_adds_per_position"]
ADD_ON_MIN_GAIN_PCT     = _P["sizing"]["add_on_min_gain_pct"]
TRAILING_STOP_PCT       = _P["exits"]["trailing_stop_pct"]
TRAIL_ACTIVATE_GAIN_PCT = _P["exits"]["trail_activate_gain_pct"]
TAKE_PROFIT_PCT         = _P["exits"]["take_profit_pct"]
MOMENTUM_EXIT_MA        = _P["exits"]["momentum_exit_ma"]
MA_CONFIRM_DAYS         = _P["exits"]["ma_confirm_days"]
MA_BREAKDOWN_PCT        = _P["exits"]["ma_breakdown_pct"]
GM_EROSION_CYCLICAL_THR = _P["exits"]["gm_erosion_cyclical_thr"]
GM_EROSION_NONCYC_THR   = _P["exits"]["gm_erosion_noncyc_thr"]
MAX_HOLD_DAYS           = _P["exits"]["max_hold_days"]
MIN_HOLD_DAYS           = _P["exits"]["min_hold_days"]
BELOW_MA_TREND_FLOOR    = _P["exits"]["below_ma_trend_floor"]  # drives BELOW_MA_DECLINING exit
MA100_BREAKDOWN_DAYS    = _P["exits"]["ma100_breakdown_days"]
COMMISSION_MIN_USD      = 2.50           # minimum commission per transaction (buy or sell)
COMMISSION_PER_SHARE    = 0.01           # per-share commission rate
REPORTING_LAG_DAYS      = 90             # calendar days after fiscal year end (matches EARNINGS_LAG_DAYS)
SIM_START               = "2024-04-19"
SIM_END                 = "2026-06-13"
# ============================================================================
#  END CONFIG
# ============================================================================

# Import gate logic from tester.py (safe now that log setup is in main())
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))
import tester as _t

# ============================================================================
#  TICKER UNIVERSE  (mirrors data_fetcher.py SECTORS)
# ============================================================================
SECTORS = {
    "Enterprise SaaS & AI":    ["MSFT","PLTR","ORCL","CRM","PATH","AI","GTLB","APPN","WIX","GDDY","YEXT","ZETA","BOX","DBX","ASAN","S","TUYA"],
    "Cybersecurity":           ["CRWD","PANW","FTNT","ZS","OKTA","TENB","VRNS","CHKP","QLYS","RDWR","NET","BB","RPD"],
    "FinTech & Payments":      ["PYPL","XYZ","STNE","PAGS","DLO","CPAY","WEX","FLYW","TOST","FOUR","PAYO","MQ","PSFE","GCT","EVTC"],
    "Data & Infrastructure":   ["MDB","SNOW","DDOG","NTNX","NTAP","RXT","IOT","DOX","TDC","AVPT","VRSN"],
    "Communications & Ops":    ["TWLO","BAND","RAMP","BLZE","OSPN","CSGS","NTSK"],
    "Solar Hardware":          ["ENPH","SEDG","FSLR","CSIQ","JKS","SHLS","NXT","SPWR"],
    "Solar Installation":      ["RUN","ARRY","SUUN"],
    "Renewable Utilities":     ["BEPC","BEP","CWEN","ORA","NRGV","FLNC"],
    "CleanTech / Emerging":    ["TURB","ASTI","SMXT","BEEM","BNRG"],
    "Medical Devices (Heavy)": ["MDT","SYK","ZBH","ISRG","BSX","EW","GEHC","PHG","STE","GMED","HAE"],
    "Monitoring & Specialized":["DXCM","PODD","TNDM","MASI","INSP","GKOS","IRTC","SENS","OSIX"],
    "Diagnostics & Lab Tech":  ["ABT","CODX","BRKR","PACB","QDEL","NEOG","BFLY","QSI","CTKB"],
    "Emerging & Biotech Med":  ["LMRI","LAB","MDAI","ATEC","CERS","STIM","TMCI","LUNG"],
    "Major Processors":        ["NVDA","AMD","INTC","AVGO","ARM"],
    "Memory & Storage":        ["MU","RMBS","WOLF"],
    "Foundries":               ["TSM","GFS","UMC","TSEM","ASX"],
    "Connectivity":            ["QCOM","MRVL","ALAB","CRDO","QRVO","SWKS","SYNA"],
    "Analog & Power":          ["TXN","ADI","NXPI","MCHP","ON","STM","VSH","ALGM","POWI"],
    "Emerging/Small Cap":      ["NVTS","POET","LAES","INDI","LSCC","MXL","SKYT","HIMX","MTSI"],
}
# Delisted / failed names -- included so survivorship bias is explicit.
# The simulator will detect when their price series ends and force-close as DELISTED.
DELISTED_TICKERS = {
    "SPCE": "CleanTech / Emerging",
    "RIDE": "CleanTech / Emerging",
    "NKLA": "CleanTech / Emerging",
    "WKHS": "CleanTech / Emerging",
    "GOEV": "CleanTech / Emerging",
    "PTRA": "CleanTech / Emerging",
    "SPWR": "Solar Hardware",
    "HYLN": "CleanTech / Emerging",
    "XPEV": "CleanTech / Emerging",
}
# Extend SECTORS map so sector lookup works for delisted names
for _tk, _sec in DELISTED_TICKERS.items():
    if _sec not in SECTORS:
        SECTORS[_sec] = []
    if _tk not in SECTORS[_sec]:
        SECTORS[_sec].append(_tk)

ALL_TICKERS = [t for tickers in SECTORS.values() for t in tickers]

CYCLICAL_KEYWORDS = [
    "Semi","Solar","Semiconductor","Major Proc","Memory","Foundry",
    "Connectivity","Analog","Emerging/Small","CleanTech",
]

def _is_cyclical(sector_str):
    return any(k in sector_str for k in CYCLICAL_KEYWORDS)

# ============================================================================
#  PRICE DOWNLOAD  (DB-backed with yfinance fallback)
# ============================================================================
FETCH_DELAY = 0.3
# Candidates for one-time dead-ticker seeding. Not a hardcoded blocklist --
# the dead_tickers DB table is the source of truth and can self-correct.
_SEED_DEAD_CANDIDATES = ["RIDE", "NKLA", "GOEV", "PTRA", "OSIX"]

# Counter for None returns from get_fundamentals_asof (reported at end)
_FUND_NONE_LOG = []   # list of (ticker, date_str) tuples

# Pre-loaded fundamentals: populated once in main(), then used by get_pit_fundamentals().
# {ticker: [(avail_date_str, captured_at_str, metrics_dict), ...]} sorted by (avail, cap) ASC
_FUND_PRELOADED = {}


def _preload_fundamentals():
    """
    Bulk-load ALL PIT snapshots from the fundamentals table into memory.
    One SQLite read replaces ~105,000 per-call queries during the day loop.
    Returns the same dict format as _FUND_PRELOADED.
    """
    import sqlite3 as _sqlite3, json as _json
    db_file = str(PROJECT_ROOT / "data" / "market_data.db")
    try:
        conn = _sqlite3.connect(db_file)
        conn.execute("PRAGMA journal_mode=WAL")
        rows = conn.execute(
            """SELECT ticker, availability_date, captured_at, sector, metrics_json
               FROM fundamentals
               ORDER BY ticker, availability_date, captured_at"""
        ).fetchall()
        conn.close()
    except Exception as exc:
        print(f"  [WARN] _preload_fundamentals: DB read failed: {exc}")
        return {}

    out = {}
    for ticker, avail, cap, sector, mj in rows:
        try:
            metrics = _json.loads(mj) if mj else {}
        except Exception:
            metrics = {}
        metrics.setdefault("Ticker", ticker)
        metrics.setdefault("Sector", sector)
        metrics.setdefault("availability_date", avail)
        if ticker not in out:
            out[ticker] = []
        out[ticker].append((avail, cap, metrics))
    return out


def _seed_dead_tickers(start_str, end_str):
    """
    One-time seed: for each candidate in _SEED_DEAD_CANDIDATES, if not already
    confirmed dead (fail_count >= 2), do a quick yfinance check. If it returns
    empty, mark dead twice so the very next run skips the network entirely.
    No-op for candidates already in the dead table -- runs in microseconds.
    """
    for ticker in _SEED_DEAD_CANDIDATES:
        if _db.is_dead(ticker):
            continue
        try:
            t      = yf.Ticker(ticker)
            df_raw = t.history(start=start_str, end=end_str,
                               interval="1d", auto_adjust=True)
            if df_raw.empty:
                _db.mark_dead(ticker)   # fail_count -> 1
                _db.mark_dead(ticker)   # fail_count -> 2  (is_dead() now True)
                print(f"  [SEED-DEAD] {ticker}: confirmed no data -> dead (fail_count=2)")
            else:
                print(f"  [SEED-DEAD] {ticker}: returned data ({len(df_raw)} days) -- not marking dead")
        except Exception as exc:
            print(f"  [SEED-DEAD] {ticker}: error ({exc}) -> dead (fail_count=2)")
            try:
                _db.mark_dead(ticker)
                _db.mark_dead(ticker)
            except Exception:
                pass


def _fetch_and_cache_prices(ticker, start_str, end_str):
    """
    Fetch prices from market_data.db; if missing, download from yfinance,
    store to DB, then return from DB.
    Returns (DataFrame, used_network: bool).
    """
    try:
        df = _db.get_prices(ticker, start=start_str, end=end_str)
        if not df.empty:
            return df, False  # served from DB -- no network
    except Exception as _dbe:
        print(f"    [DB WARN] get_prices {ticker}: {_dbe}")

    # Not in DB -- fetch from yfinance and write back so the next run is cached
    try:
        t  = yf.Ticker(ticker)
        df_raw = t.history(start=start_str, end=end_str, interval="1d", auto_adjust=True)
        if df_raw.empty:
            try:
                _db.mark_dead(ticker)
            except Exception:
                pass
            return pd.DataFrame(columns=["Close", "High", "Low", "Volume"]), True
        df_raw.index = pd.to_datetime(df_raw.index).tz_localize(None)
        ohlcv = df_raw[[c for c in ["Close", "High", "Low", "Volume"] if c in df_raw.columns]].copy()
        try:
            _db.upsert_prices(ticker, ohlcv)
        except Exception as _dbe:
            print(f"    [DB WARN] upsert_prices {ticker}: {_dbe}")
        try:
            _db.clear_dead(ticker)   # un-mark if it was previously flagged dead
        except Exception:
            pass
        return _db.get_prices(ticker, start=start_str, end=end_str), True
    except Exception as exc:
        print(f"    [FETCH ERROR] {ticker}: {exc}")
        try:
            _db.mark_dead(ticker)
        except Exception:
            pass
        return pd.DataFrame(columns=["Close", "High", "Low", "Volume"]), True


def download_prices(tickers, start_str, end_str):
    """DB-backed price loader; fetches from yfinance only if missing. Returns {ticker: DataFrame}."""
    cache      = {}
    net_hits   = 0
    dead_skips = 0
    n = len(tickers)
    for i, ticker in enumerate(tickers, 1):
        print(f"  [{i:>3}/{n}] {ticker:<8}", end=" ", flush=True)
        if _db.is_dead(ticker):
            print("[SKIP-DEAD]")
            dead_skips += 1
            continue
        df, used_net = _fetch_and_cache_prices(ticker, start_str, end_str)
        if df.empty:
            print(f"no data [{'NET' if used_net else 'DB'}]")
        else:
            cache[ticker] = df
            print(f"OK ({len(df)} days) [{'NET' if used_net else 'DB'}]")
        if used_net:
            net_hits += 1
            time.sleep(FETCH_DELAY)   # rate-limit only actual HTTP requests
    db_hits = n - net_hits - dead_skips
    print(f"  Price loading: {net_hits} network fetch(es), "
          f"{db_hits} from DB cache, {dead_skips} skipped (dead)")
    return cache

# ============================================================================
#  PRECOMPUTE MOMENTUM
# ============================================================================

def precompute_momentum(price_cache):
    """
    Build rolling MA series for every ticker.
    Returns {ticker: {closes, ma_exit, ma100, ma200, ret6m, ret20}}.
    """
    out = {}
    for ticker, df in price_cache.items():
        c = df["Close"]
        out[ticker] = {
            "closes":   c,
            "ma_exit":  c.rolling(MOMENTUM_EXIT_MA, min_periods=MOMENTUM_EXIT_MA).mean(),
            "ma100":    c.rolling(100, min_periods=100).mean(),
            "ma200":    c.rolling(200, min_periods=200).mean(),
            "ret6m":    c.pct_change(126),
            "ret20":    c.pct_change(20),    # for BELOW_MA_DECLINING exit
        }
    return out


def _get_last_idx(series, d):
    """Return the last index in series where index <= d, or None."""
    mask = series.index <= d
    if not mask.any():
        return None
    return series.index[mask][-1]


def get_pit_momentum(ticker, date, precomp):
    """
    Point-in-time momentum dict for ticker as of date.
    Returns dict or None if insufficient history.
    """
    if ticker not in precomp:
        return None
    m = precomp[ticker]
    d = pd.Timestamp(date)
    last_idx = _get_last_idx(m["closes"], d)
    if last_idx is None:
        return None

    cur = float(m["closes"][last_idx])
    if cur <= 0:
        return None

    def pct(ma_series):
        v = ma_series.get(last_idx)
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return 0.0
        fv = float(v)
        return round((cur / fv - 1) * 100, 1) if fv != 0 else 0.0

    ma100_v = m["ma100"].get(last_idx)
    above_ma100 = (cur > float(ma100_v)) if (ma100_v is not None and not math.isnan(float(ma100_v))) else False

    ret6m_v = m["ret6m"].get(last_idx)
    ret6m_pct = 0.0
    if ret6m_v is not None and not (isinstance(ret6m_v, float) and math.isnan(ret6m_v)):
        try:
            ret6m_pct = round(float(ret6m_v) * 100, 1)
        except Exception:
            pass

    return {
        "Price_vs_MA50_%":   pct(m["ma_exit"]),
        "Price_vs_MA100_%":  pct(m["ma100"]),
        "Price_Above_MA100": above_ma100,
        "Price_vs_MA200_%":  pct(m["ma200"]),
        "Return_6M_%":       ret6m_pct,
    }


def compute_rs_scores(date, precomp):
    """
    Cross-sectional relative-strength percentile (1-100) for all tickers.
    Matches the formula in data_fetcher.py: rank(pct=True)*99+1.
    """
    returns = {}
    d = pd.Timestamp(date)
    for ticker in ALL_TICKERS:
        if ticker not in precomp:
            continue
        m = precomp[ticker]
        last_idx = _get_last_idx(m["ret6m"], d)
        if last_idx is None:
            continue
        v = m["ret6m"].get(last_idx)
        if v is not None and not (isinstance(v, float) and math.isnan(v)):
            try:
                returns[ticker] = float(v)
            except Exception:
                pass

    if not returns:
        return {}

    vals = list(returns.values())
    n = len(vals)
    scores = {}
    for ticker, r in returns.items():
        rank_pct = sum(1 for v in vals if v < r) / max(n - 1, 1)
        scores[ticker] = round(rank_pct * 99 + 1, 0)
    return scores

# ============================================================================
#  PIT FUNDAMENTALS  (DB-backed)
# ============================================================================

def get_pit_fundamentals(ticker, as_of_date, pit_db=None):
    """
    Return the most recent PIT snapshot for ticker where
    availability_date <= as_of_date and captured_at <= as_of_date.

    Fast path: if _FUND_PRELOADED is populated (populated in main() before simulation),
    a simple linear scan of the ticker's sorted snapshot list replaces a SQLite round-trip.
    This turns ~105,000 per-call DB queries into one bulk load + in-memory lookups.

    Fallback: DB query via _db.get_fundamentals_asof() if ticker not preloaded.
    """
    date_str = str(pd.Timestamp(as_of_date).date())

    snaps = _FUND_PRELOADED.get(ticker)
    if snaps is not None:
        # Linear scan ascending by (avail, cap); keep last entry where both <= date_str.
        # Matches db.py ORDER BY availability_date DESC, captured_at DESC LIMIT 1.
        best = None
        for avail, cap, metrics in snaps:
            if avail > date_str:
                break
            if cap <= date_str:
                best = metrics
        if best is None:
            _FUND_NONE_LOG.append((ticker, date_str))
        return best

    # Fallback: ticker absent from preloaded dict (e.g., preload failed or partial)
    try:
        result = _db.get_fundamentals_asof(ticker, as_of_date)
        if result is None:
            _FUND_NONE_LOG.append((ticker, date_str))
        return result
    except Exception as exc:
        print(f"  [DB WARN] get_fundamentals_asof({ticker}): {exc}")
        return None


# ============================================================================
#  NDX REGIME
# ============================================================================

def compute_ndx_regime(ndx_closes, date):
    """4-state regime using QQQ vs MA100 + 20-day realized vol."""
    HIGH_VOL_THR = 20.0
    d = pd.Timestamp(date)
    avail = ndx_closes[ndx_closes.index <= d]
    if len(avail) < 25:
        return "BULL_STRONG"
    window = min(100, len(avail))
    ma100 = float(avail.iloc[-window:].mean())
    above = float(avail.iloc[-1]) > ma100
    daily_rets = avail.pct_change().dropna()
    if len(daily_rets) < 5:
        return "BULL_STRONG"
    vol20 = float(daily_rets.iloc[-20:].std()) * (252 ** 0.5) * 100
    if above:
        return "BULL_STRONG" if vol20 < HIGH_VOL_THR else "BULL_WEAK"
    return "BEAR_VOLATILE" if vol20 >= HIGH_VOL_THR else "BEAR_GRIND"

# ============================================================================
#  SCORING
# ============================================================================

def score_stock_pit(ticker, as_of_date, fund_row, pit_mom, rs_score, ndx_regime):
    """
    Point-in-time gate score for one stock.
    Returns dict {ticker, sector, universe, score, threshold, passed, confidence, veto}
    or None if sector is not mapped or gates error.
    """
    _t._ndx_regime = ndx_regime

    row = dict(fund_row)
    row.update(pit_mom)
    row["MARKET_REGIME"] = ndx_regime
    row["Relative_Strength_Score"] = rs_score

    try:
        vetoed, veto_reason = _t.check_veto(row, 50.0)
        universe, sub, gate_results = _t.run_gates(row, 50.0)
        if gate_results is None:
            return None
        rec = _t.build_record(row, universe, sub, gate_results,
                              price_chg=0.0, veto=vetoed,
                              veto_reason=veto_reason, sector_pct_rank=50.0)
    except Exception:
        return None

    gate_margins = {}
    for _gn, _gd in rec.get("gates", {}).items():
        _sc = _gd.get("score", 0.0)
        _mw = _gd.get("weight", 1.0)
        if _sc > 0 and _mw > 0:
            gate_margins[_gn] = {
                "score":      round(_sc, 2),
                "max_weight": round(_mw, 2),
                "barely":     _sc <= _mw * 0.70,
            }

    return {
        "ticker":        ticker,
        "sector":        rec["sector"],
        "universe":      universe,
        "score":         rec["weighted_score"],
        "threshold":     rec["pass_threshold"],
        "rescue_bonus":  round(rec.get("rescue_bonus", 0.0), 2),
        "passed":        rec["strategy_passed"],
        "confidence":    rec["confidence"],
        "veto":          rec["veto"],
        "gate_margins":  gate_margins,
    }


def score_universe(date, fund_rows, precomp, ndx_regime, pit_db=None):
    """
    Score all tickers as of date.
    Uses DB-backed PIT fundamentals via get_pit_fundamentals().
    Returns list of passing (non-vetoed) dicts, sorted by score descending.
    """
    rs_scores = compute_rs_scores(date, precomp)
    results   = []

    for ticker in ALL_TICKERS:
        fund_row = get_pit_fundamentals(ticker, date, pit_db)
        if fund_row is None:
            continue

        pit_mom = get_pit_momentum(ticker, date, precomp)
        if pit_mom is None:
            continue

        rs     = rs_scores.get(ticker, 50.0)
        result = score_stock_pit(ticker, date, fund_row, pit_mom, rs, ndx_regime)
        if result and result["passed"] and not result["veto"]:
            results.append(result)

    results.sort(key=lambda x: x["score"], reverse=True)
    return results

# ============================================================================
#  EXIT CONDITIONS
# ============================================================================

DELIST_STALE_DAYS = 15   # trading days: if price data stops this many days before sim end -> DELISTED

def check_exits(positions, date, precomp, fund_rows, pit_db=None, sim_end_ts=None):
    """
    Evaluate all exit conditions for every held position.
    Returns list of {ticker, exit_price, exit_reason}.
    Updates peak_price in-place.
    Conditions (checked in order):
      0. DELISTED: price series ended >DELIST_STALE_DAYS before sim end
      0.5 MA100_BREAKDOWN: N consecutive closes > MA_BREAKDOWN_PCT below MA100
      1. TRAIL_STOP: current < peak * (1 - TRAILING_STOP_PCT), only after trail_activated
      1.5 TAKE_PROFIT: gain from avg_cost >= TAKE_PROFIT_PCT
      (the following only fire once trading_days_held >= MIN_HOLD_DAYS)
      2. MA_CROSS: N consecutive closes below MA50 (only after ever-above since entry)
      3. BELOW_MA_DECLINING: below MA50+MA100 AND 20-day return < -BELOW_MA_TREND_FLOOR
      4. GM_EROSION_VETO: from as-of PIT snapshot
      5. MAX_HOLD: trading_days_held >= MAX_HOLD_DAYS
    """
    d   = pd.Timestamp(date)
    exits = []

    for ticker, pos in list(positions.items()):
        if ticker not in precomp:
            continue
        m = precomp[ticker]
        last_idx = _get_last_idx(m["closes"], d)

        # 0. Delisted: price series ended mid-sim.
        # Fire once: when today d is past the absolute last price AND the series
        # ended significantly before sim_end (>=DELIST_STALE_DAYS calendar days early).
        # Active stocks have data through fetch_end (sim_end+2d), so their gap is 0.
        if sim_end_ts is not None:
            series_last = m["closes"].index[-1]
            gap_to_sim_end = (sim_end_ts - series_last).days
            if gap_to_sim_end >= DELIST_STALE_DAYS and d > series_last:
                delist_price = float(m["closes"].iloc[-1])
                exits.append({
                    "ticker":      ticker,
                    "exit_price":  delist_price,
                    "exit_reason": "DELISTED",
                })
                continue

        if last_idx is None:
            continue

        current_price = float(m["closes"][last_idx])
        if current_price <= 0:
            continue

        peak_price = pos["peak_price"]
        if current_price > peak_price:
            pos["peak_price"] = current_price
            peak_price = current_price

        exit_reason = None

        # Update MA100 consecutive-below counter (runs every day, before exit tests)
        ma100_avail_chk = m["ma100"][m["ma100"].index <= d]
        if len(ma100_avail_chk):
            ma100_chk = float(ma100_avail_chk.iloc[-1])
            if not math.isnan(ma100_chk):
                if current_price < ma100_chk * (1 - MA_BREAKDOWN_PCT):
                    pos["days_below_ma100"] = pos.get("days_below_ma100", 0) + 1
                else:
                    pos["days_below_ma100"] = 0

        # 0.5 MA100 Breakdown: sustained close below MA100 (punctual backstop, no ever-above needed)
        if pos.get("days_below_ma100", 0) >= MA100_BREAKDOWN_DAYS:
            exit_reason = f"MA100_BREAKDOWN ({MA100_BREAKDOWN_DAYS}d below MA100)"

        gain_from_cost = (current_price - pos["avg_cost"]) / pos["avg_cost"]

        # Activate trailing stop once position has earned it
        if not pos.get("trail_activated", False) and gain_from_cost >= TRAIL_ACTIVATE_GAIN_PCT:
            pos["trail_activated"] = True

        # 1. Trailing stop -- only fires after activation
        trail_stop = peak_price * (1.0 - TRAILING_STOP_PCT)
        if pos.get("trail_activated", False) and current_price <= trail_stop:
            exit_reason = f"TRAIL_STOP (peak={peak_price:.2f} stop={trail_stop:.2f})"

        # 1.5 Take-profit: hard exit when gain exceeds threshold
        if exit_reason is None and gain_from_cost >= TAKE_PROFIT_PCT:
            exit_reason = f"TAKE_PROFIT ({gain_from_cost*100:.1f}% gain from avg_cost)"

        # Minimum hold period: only TRAIL_STOP, TAKE_PROFIT, and MA100_BREAKDOWN
        # (above) may fire before this. All checks below are gated.
        past_min_hold = pos["trading_days_held"] >= MIN_HOLD_DAYS

        # 2. MA cross (genuine cross-below only)
        if exit_reason is None and past_min_hold and pos["trading_days_held"] >= MA_CONFIRM_DAYS:
            close_avail = m["closes"][m["closes"].index <= d]
            ma_avail    = m["ma_exit"][m["ma_exit"].index <= d]
            if len(close_avail) >= MA_CONFIRM_DAYS and len(ma_avail) >= MA_CONFIRM_DAYS:
                close_tail = close_avail.iloc[-MA_CONFIRM_DAYS:]
                ma_tail    = ma_avail.iloc[-MA_CONFIRM_DAYS:]
                n_below = sum(
                    1 for c_v, m_v in zip(close_tail.values, ma_tail.values)
                    if not math.isnan(float(m_v)) and float(c_v) < float(m_v)
                       and float(c_v) < float(m_v) * (1 - MA_BREAKDOWN_PCT)
                )
                if n_below >= MA_CONFIRM_DAYS:
                    entry_ts = pd.Timestamp(pos["entry_date"])
                    cs_entry = close_avail[close_avail.index >= entry_ts]
                    ms_entry = ma_avail[ma_avail.index >= entry_ts]
                    ever_above = any(
                        not math.isnan(float(mv)) and float(c) > float(mv)
                        for c, mv in zip(cs_entry.values, ms_entry.values)
                    )
                    if ever_above:
                        exit_reason = f"MA{MOMENTUM_EXIT_MA}_CROSS ({MA_CONFIRM_DAYS}d confirm)"

        # 3. Below-MAs-and-declining exit (replaces BELOW_MA_STALLED)
        # Exit when: below both MA50 AND MA100, AND 20-day return is negative
        # (confirmed medium-term downtrend, not just a noisy dip below the MAs)
        if exit_reason is None and past_min_hold:
            ma50_avail  = m["ma_exit"][m["ma_exit"].index <= d]
            ma100_avail = m["ma100"][m["ma100"].index <= d]
            if len(ma50_avail) and len(ma100_avail):
                ma50_v  = float(ma50_avail.iloc[-1])
                ma100_v = float(ma100_avail.iloc[-1])
                if not math.isnan(ma50_v) and not math.isnan(ma100_v):
                    below_both = current_price < ma50_v and current_price < ma100_v
                    if below_both:
                        ret20_series = m["ret20"]
                        last_ret20_idx = _get_last_idx(ret20_series, d)
                        if last_ret20_idx is not None:
                            ret20 = ret20_series.get(last_ret20_idx)
                            if ret20 is not None and not (isinstance(ret20, float) and math.isnan(ret20)):
                                if float(ret20) < -BELOW_MA_TREND_FLOOR:
                                    exit_reason = (
                                        f"BELOW_MA_DECLINING (below MA50+MA100, "
                                        f"20d_ret={float(ret20)*100:.1f}% < -{BELOW_MA_TREND_FLOOR*100:.0f}%)"
                                    )

        # 4. GM erosion veto -- read from as-of PIT snapshot when available
        if exit_reason is None and past_min_hold:
            fund_row = get_pit_fundamentals(ticker, date, pit_db) or {}
            gm_erosion = fund_row.get("GM Erosion", 0)
            try:
                gm_erosion = float(gm_erosion) if gm_erosion is not None else 0.0
                if math.isnan(gm_erosion):
                    gm_erosion = 0.0
            except Exception:
                gm_erosion = 0.0
            is_cyc = _is_cyclical(str(fund_row.get("Sector", "")))
            gm_thr = GM_EROSION_CYCLICAL_THR if is_cyc else GM_EROSION_NONCYC_THR
            if gm_erosion > gm_thr:
                exit_reason = f"GM_EROSION_VETO ({gm_erosion:.1f}%>{gm_thr:.0f}%)"

        # 5. Max hold days
        if exit_reason is None and past_min_hold and pos["trading_days_held"] >= MAX_HOLD_DAYS:
            exit_reason = f"MAX_HOLD ({MAX_HOLD_DAYS}d)"

        if exit_reason:
            exits.append({
                "ticker":      ticker,
                "exit_price":  current_price,
                "exit_reason": exit_reason,
            })

    return exits

# ============================================================================
#  DAY LOOP
# ============================================================================

def _calc_commission(shares, price):
    """Commission for one transaction: max(COMMISSION_MIN_USD, shares * COMMISSION_PER_SHARE)."""
    return max(COMMISSION_MIN_USD, shares * COMMISSION_PER_SHARE)


def run_simulation(sim_start_str, sim_end_str,
                   price_cache, fund_rows,
                   spy_closes, ndx_closes,
                   pit_db=None):
    """
    Main day-by-day simulation.
    Returns (closed_trades, open_positions, equity_curve, total_commissions, n_transactions).
    """
    sim_s = pd.Timestamp(sim_start_str)
    sim_e = pd.Timestamp(sim_end_str)

    # Trading dates from SPY price history inside [sim_start, sim_end]
    trading_dates = [d for d in spy_closes.index if sim_s <= d <= sim_e]
    if not trading_dates:
        print("ERROR: no trading dates in range"); sys.exit(1)

    print(f"\n  Sim range : {sim_start_str} -> {sim_end_str}")
    print(f"  Trading days: {len(trading_dates)}")

    precomp          = precompute_momentum(price_cache)
    cash             = float(STARTING_CAPITAL)
    positions        = {}     # {ticker: pos_dict}
    closed_trades    = []
    equity_curve     = []
    total_commissions = 0.0   # cumulative commissions paid (buys + sells)
    n_transactions   = 0     # count of individual buy and sell transactions
    # Exposure tracking for avg_exposure_by_regime reporting
    _regime_exposure_sum   = {}   # regime -> cumulative exposure fraction
    _regime_exposure_count = {}   # regime -> day count

    for day_idx, today in enumerate(trading_dates):
        today_str = str(today.date())
        today_dt  = today.to_pydatetime()

        # --- mark positions to today's close ---
        for ticker, pos in positions.items():
            if ticker in precomp:
                last = _get_last_idx(precomp[ticker]["closes"], today)
                if last is not None:
                    pos["current_price"] = float(precomp[ticker]["closes"][last])

        # --- exits ---
        exits = check_exits(positions, today, precomp, fund_rows,
                            pit_db=pit_db, sim_end_ts=sim_e)
        for ex in exits:
            ticker = ex["ticker"]
            if ticker not in positions:
                continue
            pos       = positions.pop(ticker)
            exit_px   = ex["exit_price"]
            shares    = pos["shares"]
            gross_proceeds = shares * exit_px
            sell_comm      = _calc_commission(shares, exit_px)
            proceeds       = gross_proceeds - sell_comm
            total_commissions += sell_comm
            n_transactions    += 1
            pnl_pct     = (exit_px - pos["avg_cost"]) / pos["avg_cost"] * 100
            pnl_dollars = proceeds - pos["total_cost"]
            cash += proceeds
            closed_trades.append({
                "ticker":            ticker,
                "sector":            pos["sector"],
                "universe":          pos["universe"],
                "entry_date":        pos["entry_date"],
                "exit_date":         today_str,
                "entry_price":       round(pos["avg_cost"],    4),
                "exit_price":        round(exit_px,            4),
                "peak_price":        round(pos["peak_price"],  4),
                "shares":            round(shares,             6),
                "total_cost":        round(pos["total_cost"],  2),
                "proceeds":          round(proceeds,           2),
                "sell_commission":   round(sell_comm,          2),
                "pnl_pct":           round(pnl_pct,            2),
                "pnl_dollars":       round(pnl_dollars,        2),
                "exit_reason":       ex["exit_reason"],
                "confidence":        pos["confidence"],
                "conviction":        pos["conviction"],
                "score":             pos["score"],
                "threshold":         pos.get("threshold",      0.0),
                "rescue_bonus":      pos.get("rescue_bonus",   0.0),
                "trading_days_held": pos["trading_days_held"],
                "gate_margins":      pos.get("gate_margins",   {}),
            })

        # --- increment day counter ---
        for pos in positions.values():
            pos["trading_days_held"] += 1

        # --- rescore universe ---
        ndx_regime      = compute_ndx_regime(ndx_closes, today)
        regime_pos_mult = REGIME_POSITION_MULT.get(ndx_regime, 1.0)
        regime_max_pos  = REGIME_MAX_POSITIONS.get(ndx_regime, MAX_POSITIONS)
        exposure_cap    = REGIME_EXPOSURE_CAP.get(ndx_regime, 1.0)
        candidates      = score_universe(today_dt, fund_rows, precomp, ndx_regime, pit_db=pit_db)

        # --- deploy free cash (exposure-capped) ---
        mtm = sum(p["shares"] * p["current_price"] for p in positions.values())
        total_equity     = cash + mtm
        # current_exposure is updated inside the loop after each fill
        current_exposure = mtm / total_equity if total_equity > 0 else 0.0

        cand_idx = 0
        while (cash > MIN_CASH_TO_TRADE
               and len(positions) < regime_max_pos
               and cand_idx < len(candidates)):

            # ── Exposure cap gate (Brake 3) ──────────────────────────────────
            # Stop opening/adding positions once current MTM exposure >= cap.
            if current_exposure >= exposure_cap:
                break

            cand   = candidates[cand_idx]
            cand_idx += 1
            ticker = cand["ticker"]
            conf   = cand["confidence"]

            # Is it already held?
            is_held = ticker in positions
            if is_held:
                if not ALLOW_PYRAMIDING:
                    continue
                # Regime-aware pyramid gating
                if ndx_regime == "BEAR_VOLATILE":
                    continue  # no adds in confirmed bear volatile
                pos = positions[ticker]
                if ndx_regime == "BEAR_GRIND" and pos["adds"] >= 1:
                    continue  # max 1 add in bear grind
                if pos["adds"] >= MAX_ADDS_PER_POSITION:
                    continue
                gain = (pos["current_price"] - pos["avg_cost"]) / pos["avg_cost"]
                if gain < ADD_ON_MIN_GAIN_PCT:
                    continue

            # Get current price
            if ticker not in precomp:
                continue
            last_idx = _get_last_idx(precomp[ticker]["closes"], today)
            if last_idx is None:
                continue
            cur_px = float(precomp[ticker]["closes"][last_idx])
            if cur_px <= 0:
                continue

            # Buy size (Brake 1: regime_pos_mult shapes individual position size)
            mult        = CONVICTION_MULT.get(conf, 1.0)
            buy_dollars = cash * PER_BUY_FRACTION * mult * regime_pos_mult

            # Cap by max position pct of total equity (existing per-position cap)
            max_position_dollars = MAX_POSITION_PCT_EQUITY * total_equity
            if is_held:
                existing_val = positions[ticker]["shares"] * cur_px
                max_position_dollars = max(0.0, max_position_dollars - existing_val)

            # ── Exposure headroom cap (Brake 3 per-buy limit) ─────────────────
            # Clip buy so the purchase cannot push exposure above the regime cap.
            headroom_dollars = (exposure_cap - current_exposure) * total_equity
            if headroom_dollars < MIN_CASH_TO_TRADE:
                break  # no meaningful headroom left -- stop the loop for today

            buy_dollars = min(buy_dollars, max_position_dollars, headroom_dollars, cash)
            if buy_dollars < MIN_CASH_TO_TRADE:
                continue

            # Compute provisional shares; then check if cash covers cost + commission.
            # If not, reduce shares so that actual_cost + commission fits in cash.
            shares = buy_dollars / cur_px
            if not FRACTIONAL_SHARES:
                shares = math.floor(shares)
            if shares <= 0:
                continue
            actual_cost = shares * cur_px
            buy_comm    = _calc_commission(shares, cur_px)
            # If cost + commission exceeds cash, reduce shares to fit
            if actual_cost + buy_comm > cash + 0.01:
                # Solve: shares * cur_px + max(MIN, shares * PER_SHARE) <= cash
                # First try fitting within MIN commission band:
                shares_min_band = (cash - COMMISSION_MIN_USD) / cur_px
                shares_per_band = cash / (cur_px + COMMISSION_PER_SHARE)
                shares = max(shares_min_band, shares_per_band)
                if not FRACTIONAL_SHARES:
                    shares = math.floor(shares)
                actual_cost = shares * cur_px
                buy_comm    = _calc_commission(shares, cur_px)
            if shares <= 0 or actual_cost <= 0:
                continue
            if actual_cost + buy_comm > cash + 0.01:
                continue   # still can't fit even after reduction -- skip

            cash -= actual_cost + buy_comm
            total_commissions += buy_comm
            n_transactions    += 1

            if is_held:
                pos = positions[ticker]
                pos["total_cost"] += actual_cost + buy_comm  # commission folded into cost basis
                pos["shares"]     += shares
                pos["avg_cost"]    = pos["total_cost"] / pos["shares"]
                pos["adds"]       += 1
            else:
                positions[ticker] = {
                    "ticker":            ticker,
                    "sector":            cand["sector"],
                    "universe":          cand["universe"],
                    "entry_date":        today_str,
                    "shares":            shares,
                    "total_cost":        actual_cost + buy_comm,  # cost basis includes buy commission
                    "avg_cost":          cur_px,
                    "peak_price":        cur_px,
                    "current_price":     cur_px,
                    "confidence":        conf,
                    "conviction":        mult,
                    "score":             cand["score"],
                    "threshold":         cand.get("threshold",    0.0),
                    "rescue_bonus":      cand.get("rescue_bonus", 0.0),
                    "trading_days_held": 0,
                    "adds":              0,
                    "days_below_ma100":  0,
                    "trail_activated":   False,
                    "gate_margins":      cand.get("gate_margins", {}),
                }

            # Recompute exposure after this fill so the next candidate sees fresh headroom
            new_mtm          = sum(p["shares"] * p["current_price"] for p in positions.values())
            current_exposure = new_mtm / total_equity if total_equity > 0 else 0.0

        # --- record equity and regime exposure ---
        mtm2 = sum(p["shares"] * p["current_price"] for p in positions.values())
        total_equity = cash + mtm2
        day_exposure = mtm2 / total_equity if total_equity > 0 else 0.0
        _regime_exposure_sum[ndx_regime]    = _regime_exposure_sum.get(ndx_regime, 0.0) + day_exposure
        _regime_exposure_count[ndx_regime]  = _regime_exposure_count.get(ndx_regime, 0) + 1
        equity_curve.append({
            "date":        today_str,
            "equity":      round(total_equity, 2),
            "cash":        round(cash, 2),
            "n_positions": len(positions),
            "exposure":    round(day_exposure, 4),
            "regime":      ndx_regime,
        })

        if day_idx % 63 == 0 or day_idx == len(trading_dates) - 1:
            print(f"  [{today_str}]  equity=${total_equity:>10,.0f}"
                  f"  cash=${cash:>8,.0f}  pos={len(positions):>3}/{regime_max_pos}"
                  f"  exp={day_exposure*100:>5.1f}%/{exposure_cap*100:.0f}%"
                  f"  closed={len(closed_trades):>4}"
                  f"  regime={ndx_regime} (x{regime_pos_mult:.2f})")

    # Build avg_exposure_by_regime summary
    avg_exp_by_regime = {
        r: round(_regime_exposure_sum[r] / _regime_exposure_count[r], 4)
        for r in _regime_exposure_sum
        if _regime_exposure_count.get(r, 0) > 0
    }
    return closed_trades, positions, equity_curve, round(total_commissions, 2), n_transactions, avg_exp_by_regime

# ============================================================================
#  DIAGNOSTIC SIGNALS
# ============================================================================

def compute_diagnostics(closed_trades, price_cache):
    """
    Compute three optimizer-facing diagnostic signals from closed trades.
    Called by compute_metrics() with the full price_cache.

    Signal 1 -- BAD ENTRIES:   losses where price fell below entry within 15 days.
    Signal 2 -- PREMATURE EXITS: exits where price was >8% higher 20 days later.
    Signal 3 -- EXIT QUALITY:  scoreboard of good / premature / rode-it-down exits.
    """
    BAD_ENTRY_DAYS  = 15     # trading days window for bad-entry detection
    PREMATURE_DAYS  = 20     # trading days lookahead for premature exit
    PREMATURE_THR   = 8.0    # pct rise after exit that flags premature
    BARELY_RATIO    = 0.70   # gate score <= 70% of max_weight -> "barely passed"
    RODE_DOWN_PEAK  = 0.10   # peak must be >=10% above entry to qualify as "rode it down"

    # ---- Signal 1: bad entries -----------------------------------------
    bad_trades = []
    for t in closed_trades:
        if t["pnl_pct"] >= 0:
            continue
        tk = t["ticker"]
        if tk not in price_cache:
            continue
        closes = price_cache[tk]["Close"]
        entry_ts = pd.Timestamp(t["entry_date"])
        post = closes[closes.index > entry_ts].iloc[:BAD_ENTRY_DAYS]
        if post.empty:
            continue
        if any(float(c) < t["entry_price"] for c in post.values):
            bad_trades.append(t)

    gate_counts = {}
    n_with_gates = 0
    for t in bad_trades:
        gm = t.get("gate_margins", {})
        if not gm:
            continue
        n_with_gates += 1
        for gname, info in gm.items():
            if info.get("barely", False):
                gate_counts[gname] = gate_counts.get(gname, 0) + 1

    ranked = sorted(gate_counts.items(), key=lambda x: -x[1])
    bad_pnl  = round(sum(t["pnl_dollars"] for t in bad_trades), 2)
    n_losers = sum(1 for t in closed_trades if t["pnl_pct"] < 0)
    top_gate = ranked[0][0] if ranked else "N/A"

    sig1 = {
        "n_bad_entries":     len(bad_trades),
        "n_total_losers":    n_losers,
        "total_pnl_dollars": bad_pnl,
        "n_with_gate_data":  n_with_gates,
        "top_barely_gates":  [
            {
                "gate":         g,
                "count":        c,
                "pct_of_bad":   round(c / n_with_gates * 100, 1) if n_with_gates else 0.0,
            }
            for g, c in ranked[:5]
        ],
        "hint": (
            f"HINT: bad entries cluster on marginal passes of {top_gate}. "
            f"Consider raising that gate threshold or weight in tester.py."
        ) if ranked and n_with_gates > 0 else
        "HINT: no gate data -- run sim once to populate gate_margins on trades.",
    }

    # ---- Signal 2: premature exits -------------------------------------
    premature    = []
    premature_idx = set()
    by_reason    = {}

    for i, t in enumerate(closed_trades):
        tk = t["ticker"]
        if tk not in price_cache:
            continue
        closes   = price_cache[tk]["Close"]
        exit_ts  = pd.Timestamp(t["exit_date"])
        after    = closes[closes.index > exit_ts]
        if len(after) < 10:
            continue
        idx        = min(PREMATURE_DAYS - 1, len(after) - 1)
        future_px  = float(after.iloc[idx])
        exit_px    = t["exit_price"]
        pct_after  = (future_px / exit_px - 1) * 100
        if pct_after > PREMATURE_THR:
            missed = round((future_px - exit_px) * t.get("shares", 0), 2)
            rkey   = t["exit_reason"].split("(")[0].strip()
            premature.append({
                "i": i, "ticker": tk,
                "exit_date":      t["exit_date"],
                "exit_reason":    rkey,
                "pct_after":      round(pct_after, 2),
                "dollars_missed": missed,
            })
            premature_idx.add(i)
            if rkey not in by_reason:
                by_reason[rkey] = {"count": 0, "dollars": 0.0}
            by_reason[rkey]["count"]   += 1
            by_reason[rkey]["dollars"] = round(by_reason[rkey]["dollars"] + missed, 2)

    total_missed = round(sum(p["dollars_missed"] for p in premature), 2)
    sorted_reasons = sorted(by_reason.items(), key=lambda x: -x[1]["dollars"])
    top_reason = sorted_reasons[0][0] if sorted_reasons else "N/A"

    sig2 = {
        "n_premature":          len(premature),
        "n_total_closed":       len(closed_trades),
        "total_dollars_missed": total_missed,
        "by_exit_reason":       {
            r: {"count": d["count"], "dollars_missed": d["dollars"]}
            for r, d in sorted_reasons
        },
        "top_reason": top_reason,
        "hint": (
            f"HINT: {len(premature)} exits left ${total_missed:,.0f} on the table, "
            f"mostly via {top_reason}. Consider loosening that specific exit "
            f"(e.g. widen TRAILING_STOP_PCT or raise MA_CONFIRM_DAYS) in CONFIG."
        ) if premature else "HINT: no premature exits detected at the 8% threshold.",
    }

    # ---- Signal 3: exit quality summary --------------------------------
    n_prem      = len(premature)
    n_rode_down = 0
    n_good      = 0

    for i, t in enumerate(closed_trades):
        if i in premature_idx:
            continue
        peak     = t.get("peak_price")
        entry_px = t["entry_price"]
        pnl      = t["pnl_pct"]
        if (peak is not None and peak > 0
                and peak > entry_px * (1.0 + RODE_DOWN_PEAK)
                and pnl < 0):
            n_rode_down += 1
        else:
            n_good += 1

    n_total = len(closed_trades)
    sig3 = {
        "total_closed": n_total,
        "n_good_exits":  n_good,
        "n_premature":   n_prem,
        "n_rode_down":   n_rode_down,
        "pct_good":      round(n_good       / n_total * 100, 1) if n_total else 0.0,
        "pct_premature": round(n_prem       / n_total * 100, 1) if n_total else 0.0,
        "pct_rode_down": round(n_rode_down  / n_total * 100, 1) if n_total else 0.0,
    }

    return {
        "signal1_bad_entries":     sig1,
        "signal2_premature_exits": sig2,
        "signal3_exit_quality":    sig3,
    }


# ============================================================================
#  METRICS
# ============================================================================

def compute_metrics(closed_trades, open_positions, equity_curve,
                    sim_start_str, sim_end_str, spy_closes,
                    price_cache=None, total_commissions=0.0, n_transactions=0,
                    avg_exposure_by_regime=None):
    m = {}
    equities  = [e["equity"] for e in equity_curve]
    final_eq  = equities[-1] if equities else STARTING_CAPITAL
    m["final_equity"]     = round(final_eq, 2)
    m["total_return_pct"] = round((final_eq / STARTING_CAPITAL - 1) * 100, 2)

    # CAGR
    n_years = len(equities) / 252.0
    if n_years > 0:
        m["cagr"] = round(((final_eq / STARTING_CAPITAL) ** (1.0 / n_years) - 1) * 100, 2)
    else:
        m["cagr"] = 0.0

    # Max drawdown
    peak_eq = equities[0]
    max_dd  = 0.0
    for eq in equities:
        peak_eq = max(peak_eq, eq)
        dd = (eq - peak_eq) / peak_eq * 100
        max_dd = min(max_dd, dd)
    m["max_drawdown"] = round(max_dd, 2)

    # Trade stats (closed only)
    wins   = [t for t in closed_trades if t["pnl_pct"] >= 0]
    losses = [t for t in closed_trades if t["pnl_pct"] <  0]
    n_cl   = len(closed_trades)
    m["n_closed"]  = n_cl
    m["n_open"]    = len(open_positions)
    m["win_rate"]  = round(len(wins) / n_cl * 100, 1) if n_cl else 0.0
    m["avg_win"]   = round(sum(t["pnl_pct"] for t in wins)   / len(wins),   2) if wins   else 0.0
    m["avg_loss"]  = round(sum(t["pnl_pct"] for t in losses) / len(losses), 2) if losses else 0.0

    tot_gain  = sum(t["pnl_dollars"] for t in wins)             if wins   else 0.0
    tot_loss  = abs(sum(t["pnl_dollars"] for t in losses))      if losses else 1e-9
    m["profit_factor"] = round(tot_gain / tot_loss, 2)

    # Sharpe (annualized from daily equity returns)
    if len(equities) > 2:
        daily_rets = [(equities[i] - equities[i-1]) / equities[i-1]
                      for i in range(1, len(equities))]
        mu  = sum(daily_rets) / len(daily_rets)
        std = statistics.stdev(daily_rets) if len(daily_rets) > 1 else 0.0
        m["sharpe"] = round(mu / std * (252 ** 0.5), 2) if std > 0 else 0.0
    else:
        m["sharpe"] = 0.0

    # Average cash idle fraction -- used by auto_optimizer STAY_INVESTED guardrail
    if equity_curve:
        _fracs = [e["cash"] / e["equity"] for e in equity_curve if e.get("equity", 0) > 0]
        m["avg_cash_pct"] = round(sum(_fracs) / len(_fracs), 4) if _fracs else 0.0
    else:
        m["avg_cash_pct"] = 0.0

    # Average market exposure by regime (new: verifies exposure cap is working)
    m["avg_exposure_by_regime"] = avg_exposure_by_regime or {}

    # SPY stats over sim period
    t_s   = pd.Timestamp(sim_start_str)
    t_e   = pd.Timestamp(sim_end_str)
    spy_window = spy_closes[(spy_closes.index >= t_s) & (spy_closes.index <= t_e)]
    if len(spy_window) >= 2:
        spy_ret = (float(spy_window.iloc[-1]) / float(spy_window.iloc[0]) - 1) * 100
        m["spy_return"] = round(spy_ret, 2)
        m["alpha"]      = round(m["total_return_pct"] - spy_ret, 2)

        # SPY Sharpe (annualized from daily returns over same window)
        spy_daily = spy_window.pct_change().dropna()
        if len(spy_daily) > 1:
            spy_mu  = float(spy_daily.mean())
            spy_std = float(spy_daily.std())
            m["spy_sharpe"] = round(spy_mu / spy_std * (252 ** 0.5), 2) if spy_std > 0 else 0.0
        else:
            m["spy_sharpe"] = 0.0

        # SPY max drawdown over same window
        spy_peak = float(spy_window.iloc[0])
        spy_mdd  = 0.0
        for v in spy_window:
            spy_peak = max(spy_peak, float(v))
            dd = (float(v) - spy_peak) / spy_peak * 100
            spy_mdd = min(spy_mdd, dd)
        m["spy_max_drawdown"] = round(spy_mdd, 2)
    else:
        m["spy_return"]      = None
        m["alpha"]           = None
        m["spy_sharpe"]      = None
        m["spy_max_drawdown"]= None

    # Per-sector breakdown (closed trades)
    sec = {}
    for t in closed_trades:
        s = t["sector"]
        if s not in sec:
            sec[s] = {"n": 0, "wins": 0, "pnl_pct_list": [], "pnl_dollars": 0.0}
        sec[s]["n"] += 1
        if t["pnl_pct"] >= 0:
            sec[s]["wins"] += 1
        sec[s]["pnl_pct_list"].append(t["pnl_pct"])
        sec[s]["pnl_dollars"] += t["pnl_dollars"]
    m["sector_stats"] = sec

    # Concentration analysis
    ticker_pnl = {}
    for t in closed_trades:
        ticker_pnl[t["ticker"]] = ticker_pnl.get(t["ticker"], 0.0) + t["pnl_dollars"]
    total_gross_pnl = sum(abs(v) for v in ticker_pnl.values()) or 1.0
    total_net_pnl   = sum(ticker_pnl.values())
    sorted_tickers  = sorted(ticker_pnl.items(), key=lambda x: -abs(x[1]))
    def _top_share(n):
        top = sorted_tickers[:n]
        return round(sum(abs(v) for _, v in top) / total_gross_pnl * 100, 1)
    m["concentration"] = {
        "top1_pct":  _top_share(1),
        "top3_pct":  _top_share(3),
        "top5_pct":  _top_share(5),
        "top1_name": sorted_tickers[0][0]  if sorted_tickers else "",
        "top3_names":[tk for tk, _ in sorted_tickers[:3]],
        "top5_names":[tk for tk, _ in sorted_tickers[:5]],
    }
    # Return excl. top-2 names (concentration stress test)
    if len(sorted_tickers) >= 2:
        excl2_set    = {tk for tk, _ in sorted_tickers[:2]}
        excl2_trades = [t for t in closed_trades if t["ticker"] not in excl2_set]
        excl2_gain   = sum(t["pnl_dollars"] for t in excl2_trades if t["pnl_dollars"] >= 0) or 0.0
        excl2_loss   = abs(sum(t["pnl_dollars"] for t in excl2_trades if t["pnl_dollars"] < 0)) or 1e-9
        m["concentration"]["excl_top2_pf"]     = round(excl2_gain / excl2_loss, 2)
        m["concentration"]["excl_top2_trades"]  = len(excl2_trades)
    else:
        m["concentration"]["excl_top2_pf"]    = None
        m["concentration"]["excl_top2_trades"] = 0

    # Delisted trade summary
    m["delisted_trades"] = [t for t in closed_trades if t["exit_reason"] == "DELISTED"]

    # Optimizer diagnostic signals (requires price_cache for post-exit lookahead)
    m["diagnostics"] = compute_diagnostics(closed_trades, price_cache or {})

    # Commission accounting
    m["total_commissions"] = round(total_commissions, 2)
    m["n_transactions"]    = n_transactions

    return m

# ============================================================================
#  REPORTING
# ============================================================================

def _safe_pct_str(v):
    if v is None:
        return "    N/A"
    return f"{v:>+7.2f}%"


def write_text_report(metrics, closed_trades, open_positions, config_dict, out_path):
    L = []
    def p(s=""): L.append(s)

    p("=" * 78)
    p("  PORTFOLIO SIMULATOR REPORT")
    p(f"  Generated : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    p(f"  Period    : {config_dict['SIM_START']} -> {config_dict['SIM_END']}")
    p("=" * 78)
    p()
    p("  SUMMARY")
    p("  " + "-" * 44)
    p(f"  Starting Capital : ${STARTING_CAPITAL:>12,.2f}")
    p(f"  Final Equity     : ${metrics['final_equity']:>12,.2f}")
    p(f"  Total Return     : {metrics['total_return_pct']:>+8.2f}%")
    p(f"  CAGR             : {metrics['cagr']:>+8.2f}%")
    p(f"  Max Drawdown     : {metrics['max_drawdown']:>+8.2f}%")
    p(f"  Win Rate         : {metrics['win_rate']:>8.1f}%  ({metrics['n_closed']} closed / {metrics['n_open']} open)")
    p(f"  Avg Win          : {metrics['avg_win']:>+8.2f}%")
    p(f"  Avg Loss         : {metrics['avg_loss']:>+8.2f}%")
    p(f"  Profit Factor    : {metrics['profit_factor']:>8.2f}x")
    p(f"  Portfolio Sharpe : {metrics['sharpe']:>8.2f}  (annualized, daily equity)")
    p(f"  SPY Sharpe       : {metrics['spy_sharpe'] if metrics['spy_sharpe'] is not None else 'N/A':>8}  (same window)")
    p(f"  SPY Max Drawdown : {_safe_pct_str(metrics['spy_max_drawdown'])}")
    p(f"  SPY Return       : {_safe_pct_str(metrics['spy_return'])}")
    p(f"  Alpha vs SPY     : {_safe_pct_str(metrics['alpha'])}")
    tot_comm   = metrics.get('total_commissions', 0.0)
    n_txn      = metrics.get('n_transactions', 0)
    comm_pct   = tot_comm / STARTING_CAPITAL * 100
    p(f"  Total Commissions: ${tot_comm:>9,.2f}  ({n_txn} transactions,  {comm_pct:.2f}% of starting capital)")
    p()

    # Regime exposure table
    exp_by_regime = metrics.get("avg_exposure_by_regime", {})
    if exp_by_regime:
        caps = REGIME_EXPOSURE_CAP
        p("  REGIME EXPOSURE  (avg daily MTM / total_equity, vs configured cap)")
        p("  " + "-" * 56)
        p(f"  {'Regime':<16} {'AvgExposure':>12}  {'Cap':>6}  {'Days':>6}")
        p("  " + "-" * 56)
        regime_order = ["BULL_STRONG", "BULL_WEAK", "BEAR_GRIND", "BEAR_VOLATILE"]
        for rg in regime_order:
            if rg not in exp_by_regime:
                continue
            avg_exp = exp_by_regime[rg]
            cap_val = caps.get(rg, 1.0)
            # Count days from equity_curve
            n_days = sum(1 for e in metrics.get("_equity_curve_ref", []) if e.get("regime") == rg)
            cap_str = f"{cap_val*100:.0f}%"
            p(f"  {rg:<16} {avg_exp*100:>11.1f}%  {cap_str:>6}  {n_days:>6}")
        p("  " + "-" * 56)
        p()

    # Concentration block
    con = metrics.get("concentration", {})
    p("  CONCENTRATION (gross P&L share by name)")
    p("  " + "-" * 44)
    p(f"  Top 1 name  ({con.get('top1_name',''):<6}): {con.get('top1_pct', 0):>5.1f}% of gross P&L")
    p(f"  Top 3 names ({','.join(con.get('top3_names',[])):<18}): {con.get('top3_pct', 0):>5.1f}%")
    p(f"  Top 5 names ({','.join(con.get('top5_names',[])):<22}): {con.get('top5_pct', 0):>5.1f}%")
    excl2_pf = con.get("excl_top2_pf")
    excl2_n  = con.get("excl_top2_trades", 0)
    if excl2_pf is not None:
        p(f"  Excl. top-2: profit factor = {excl2_pf:.2f}x  ({excl2_n} trades)")
    p()

    # Diagnostic signals
    diag = metrics.get("diagnostics", {})
    s1   = diag.get("signal1_bad_entries",     {})
    s2   = diag.get("signal2_premature_exits", {})
    s3   = diag.get("signal3_exit_quality",    {})
    if diag:
        p("  DIAGNOSTIC SIGNALS  (optimizer guidance)")
        p("  " + "-" * 68)

        p("  SIGNAL 1 -- BAD ENTRIES  (lever: tighten gates in tester.py)")
        n_bad  = s1.get("n_bad_entries",  0)
        n_loss = s1.get("n_total_losers", 0)
        bad_pl = s1.get("total_pnl_dollars", 0.0)
        n_gd   = s1.get("n_with_gate_data", 0)
        p(f"    {n_bad} of {n_loss} losing trades fell below entry price within 15 trading days")
        p(f"    Total P&L from bad entries : ${bad_pl:>+,.0f}")
        if n_gd > 0:
            p(f"    Gate barely-passed ranking  ({n_gd} bad-entry trades with gate data):")
            for rank_i, rec in enumerate(s1.get("top_barely_gates", []), 1):
                gname = rec.get("gate", "?")[:24]
                cnt   = rec.get("count", 0)
                pct   = rec.get("pct_of_bad", 0.0)
                p(f"      {rank_i}. {gname:<24} : {cnt:>3}/{n_gd}  ({pct:.1f}%)")
        else:
            p("    (no gate data yet -- will populate on next run)")
        p(f"    {s1.get('hint', '')}")
        p()

        p("  SIGNAL 2 -- PREMATURE EXITS  (lever: loosen/adjust stop in CONFIG)")
        n_prem  = s2.get("n_premature",          0)
        n_cl    = s2.get("n_total_closed",        0)
        missed  = s2.get("total_dollars_missed",  0.0)
        p(f"    {n_prem} of {n_cl} exits left ${missed:,.0f} on the table"
          f"  (20d lookahead, >8% threshold)")
        p(f"    By exit reason  (ranked by dollars missed):")
        for rkey, rdata in s2.get("by_exit_reason", {}).items():
            p(f"      {rkey[:22]:<22} : {rdata['count']:>4} exits  / ${rdata['dollars_missed']:>+9,.0f} missed")
        p(f"    {s2.get('hint', '')}")
        p()

        p("  SIGNAL 3 -- EXIT QUALITY SUMMARY")
        n_tot = s3.get("total_closed",  1)
        p(f"    Good exits  (sold near top / no clear missed gain) : "
          f"{s3.get('n_good_exits', 0):>4}  ({s3.get('pct_good', 0):.1f}%)")
        p(f"    Premature   (price rose >8% in 20d after exit)     : "
          f"{s3.get('n_premature', 0):>4}  ({s3.get('pct_premature', 0):.1f}%)")
        p(f"    Rode it down (peak >=10% above entry, exited loss)  : "
          f"{s3.get('n_rode_down', 0):>4}  ({s3.get('pct_rode_down', 0):.1f}%)")
        p("  " + "-" * 68)
        p()

    # Delisted summary
    dl = metrics.get("delisted_trades", [])
    if dl:
        p("  DELISTED NAMES (forced-close at last price)")
        p("  " + "-" * 68)
        dl_net = sum(t["pnl_dollars"] for t in dl)
        for t in sorted(dl, key=lambda x: x["entry_date"]):
            p(f"  {t['ticker']:<7}  entry={t['entry_date']}  exit={t['exit_date']}"
              f"  ret={t['pnl_pct']:>+7.2f}%  P&L=${t['pnl_dollars']:>+9,.0f}")
        p(f"  Total delisted P&L: ${dl_net:>+,.0f}  ({len(dl)} trades)")
        p()

    p("  KNOWN LIMITATIONS:")
    p("  - PIT fundamentals go back ~4-5 years (yfinance limit); earlier periods")
    p("    marked INSUFFICIENT_PIT_DATA in walk_forward robustness table.")
    p("  - Tickers with only 1 PIT snapshot flagged SHALLOW in pit_coverage.txt.")
    p("  - Delisted tickers (SPCE NKLA RIDE etc.) are included; forced-closed above.")
    p()
    p("  PER-SECTOR BREAKDOWN (closed trades)")
    p("  " + "-" * 68)
    p(f"  {'Sector':<32} {'N':>4}  {'WinR%':>6}  {'AvgRet%':>8}  {'P&L $':>10}")
    p("  " + "-" * 68)
    for sector, st in sorted(metrics["sector_stats"].items(),
                              key=lambda x: -x[1]["pnl_dollars"]):
        n   = st["n"]
        wr  = st["wins"] / n * 100 if n else 0.0
        avg = sum(st["pnl_pct_list"]) / n if n else 0.0
        p(f"  {sector[:32]:<32} {n:>4}  {wr:>6.1f}  {avg:>+8.2f}  {st['pnl_dollars']:>+10,.0f}")
    p()
    p("  TRADE LOG (closed trades, sorted by entry date)")
    p("  " + "-" * 108)
    p(f"  {'Ticker':<7}  {'Sector':<26}  {'EntryDate':<12}  {'ExitDate':<12}"
      f"  {'Entry':>7}  {'Exit':>7}  {'Ret%':>7}  {'Conf':<5}  Exit Reason")
    p("  " + "-" * 108)
    for t in sorted(closed_trades, key=lambda x: x["entry_date"]):
        reason = t["exit_reason"].split("(")[0].strip()[:22]
        p(f"  {t['ticker']:<7}  {t['sector'][:26]:<26}  {t['entry_date']:<12}  {t['exit_date']:<12}"
          f"  {t['entry_price']:>7.2f}  {t['exit_price']:>7.2f}  {t['pnl_pct']:>+7.2f}"
          f"  {t['confidence']:<5}  {reason}")
    p()
    if open_positions:
        p("  OPEN POSITIONS (valued at last close)")
        p("  " + "-" * 72)
        for ticker, pos in sorted(open_positions.items(),
                                   key=lambda x: -(x[1]["shares"] * x[1]["current_price"])):
            val  = pos["shares"] * pos["current_price"]
            gain = (pos["current_price"] - pos["avg_cost"]) / pos["avg_cost"] * 100
            p(f"  {ticker:<7}  avg_cost={pos['avg_cost']:.2f}  cur={pos['current_price']:.2f}"
              f"  gain={gain:>+7.2f}%  val=${val:>10,.0f}  [{pos['confidence']}  d={pos['trading_days_held']}]")
        p()
    p("  CONFIG")
    p("  " + "-" * 44)
    for k, v in config_dict.items():
        p(f"  {k:<30} = {v}")
    p()
    p("=" * 78)

    text = "\n".join(L)
    out_path.write_text(text, encoding="ascii", errors="replace")
    return text


def write_json_report(metrics, closed_trades, open_positions, config_dict, out_path):
    _excl = {"sector_stats", "concentration", "delisted_trades", "_equity_curve",
             "_equity_curve_ref", "diagnostics"}
    data = {
        "generated":              datetime.now().isoformat(),
        "config":                 config_dict,
        "summary":                {k: v for k, v in metrics.items() if k not in _excl},
        "avg_exposure_by_regime": metrics.get("avg_exposure_by_regime", {}),
        "concentration":          metrics.get("concentration", {}),
        "delisted_trades":        metrics.get("delisted_trades", []),
        "diagnostics":            metrics.get("diagnostics", {}),
        "sector_stats":  {
            sec: {"n": st["n"], "wins": st["wins"],
                  "pnl_dollars": round(st["pnl_dollars"], 2),
                  "avg_ret_pct": round(sum(st["pnl_pct_list"]) / st["n"], 2) if st["n"] else 0}
            for sec, st in metrics["sector_stats"].items()
        },
        "closed_trades": [
            dict(t)
            for t in closed_trades
        ],
        "open_positions": [
            {**{k: v for k, v in pos.items() if k not in ("peak_price", "gate_margins")},
             "ticker": tk}
            for tk, pos in open_positions.items()
        ],
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"  JSON  -> {out_path}")


def write_xlsx_report(metrics, closed_trades, config_dict, out_path):
    if not _XLSX_OK:
        print("  [WARN] openpyxl not installed -- skipping XLSX"); return
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Summary"
    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    rows = [
        ("Final Equity",   f"${metrics['final_equity']:,.2f}"),
        ("Total Return %", f"{metrics['total_return_pct']:+.2f}%"),
        ("CAGR %",         f"{metrics['cagr']:+.2f}%"),
        ("Max Drawdown %", f"{metrics['max_drawdown']:+.2f}%"),
        ("Win Rate %",     f"{metrics['win_rate']:.1f}%"),
        ("Closed Trades",  metrics["n_closed"]),
        ("Open Positions", metrics["n_open"]),
        ("Avg Win %",      f"{metrics['avg_win']:+.2f}%"),
        ("Avg Loss %",     f"{metrics['avg_loss']:+.2f}%"),
        ("Profit Factor",  f"{metrics['profit_factor']:.2f}x"),
        ("Sharpe",         f"{metrics['sharpe']:.2f}"),
        ("SPY Return %",   f"{metrics['spy_return']:+.2f}%" if metrics["spy_return"] is not None else "N/A"),
        ("Alpha vs SPY %", f"{metrics['alpha']:+.2f}%"     if metrics["alpha"] is not None else "N/A"),
    ]
    for ri, (k, v) in enumerate(rows, 1):
        c = ws.cell(ri, 1, k); c.font = Font(bold=True)
        ws.cell(ri, 2, str(v))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    # Trades sheet
    ws2  = wb.create_sheet("Trades")
    cols = ["Ticker","Sector","Universe","Entry Date","Exit Date","Avg Cost $",
            "Exit Price $","Return %","P&L $","Confidence","Conviction","Score",
            "Exit Reason","Days Held"]
    hf   = PatternFill("solid", fgColor="1F3864")
    hfont= Font(name="Arial", size=10, bold=True, color="FFFFFF")
    for ci, col in enumerate(cols, 1):
        c = ws2.cell(1, ci, col)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = bdr
    ws2.row_dimensions[1].height = 24

    gfill = PatternFill("solid", fgColor="C6EFCE")
    rfill = PatternFill("solid", fgColor="FFC7CE")
    for ri, t in enumerate(sorted(closed_trades, key=lambda x: x["pnl_pct"], reverse=True), 2):
        fill = gfill if t["pnl_pct"] >= 0 else rfill
        vals = [t["ticker"], t["sector"], t["universe"],
                t["entry_date"], t["exit_date"],
                t["entry_price"], t["exit_price"],
                t["pnl_pct"] / 100, t["pnl_dollars"],
                t["confidence"], t["conviction"], t["score"],
                t["exit_reason"], t["trading_days_held"]]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(ri, ci, val)
            c.fill = fill
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="center")
            c.border = bdr
            if ci == 8:  c.number_format = "+0.00%;-0.00%"
            if ci == 9:  c.number_format = '"$"#,##0.00'
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws2.freeze_panes = "A2"

    # Equity curve sheet
    ws3  = wb.create_sheet("Equity Curve")
    ws3.append(["Date", "Equity $", "Cash $", "Positions"])
    for e in metrics.get("_equity_curve", []):
        ws3.append([e["date"], e["equity"], e["cash"], e["n_positions"]])
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 14

    wb.save(out_path)
    print(f"  XLSX  -> {out_path}")


def append_perf_history(metrics, sim_start, sim_end):
    ts  = datetime.now().strftime("%Y-%m-%d %H:%M")
    spy = f"{metrics['spy_return']:+.1f}%" if metrics["spy_return"] is not None else "N/A"
    alp = f"{metrics['alpha']:+.1f}%"      if metrics["alpha"]      is not None else "N/A"
    line = (
        f"[{ts}] PORTFOLIO_SIM  {sim_start}->{sim_end}"
        f"  return={metrics['total_return_pct']:+.2f}%"
        f"  cagr={metrics['cagr']:+.2f}%"
        f"  winrate={metrics['win_rate']:.1f}%"
        f"  maxdd={metrics['max_drawdown']:+.2f}%"
        f"  sharpe={metrics['sharpe']:.2f}"
        f"  pf={metrics['profit_factor']:.2f}"
        f"  spy={spy}  alpha={alp}"
        f"  n_closed={metrics['n_closed']}  n_open={metrics['n_open']}\n"
    )
    with open(PERF_HIST, "a", encoding="ascii", errors="replace") as f:
        f.write(line)

# ============================================================================
#  MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Portfolio Simulator")
    parser.add_argument("--start",        default=SIM_START)
    parser.add_argument("--end",          default=SIM_END)
    parser.add_argument("--recheck-dead", action="store_true",
                        help="Clear dead-ticker table and force a fresh network attempt on all of them")
    args = parser.parse_args()

    sim_start = args.start
    sim_end   = args.end

    # Warmup: 350 calendar days before start so MA200 is valid on day 1
    warmup_dt  = datetime.strptime(sim_start, "%Y-%m-%d") - timedelta(days=350)
    warmup     = warmup_dt.strftime("%Y-%m-%d")
    fetch_end  = (datetime.strptime(sim_end, "%Y-%m-%d") + timedelta(days=2)).strftime("%Y-%m-%d")

    print("=" * 70)
    print("  PORTFOLIO SIMULATOR")
    print(f"  Period     : {sim_start} -> {sim_end}")
    print(f"  Warmup from: {warmup}")
    print(f"  Capital    : ${STARTING_CAPITAL:,.0f}")
    print(f"  Per-buy    : {PER_BUY_FRACTION*100:.0f}% of free cash")
    print(f"  Conviction : HIGH={CONVICTION_MULT['HIGH']}x  MED={CONVICTION_MULT['MED']}x  LOW={CONVICTION_MULT['LOW']}x")
    print(f"  Regime mult: " + "  ".join(f"{k}={v}x" for k, v in REGIME_POSITION_MULT.items()))
    print("=" * 70)

    t0_total = time.time()

    # Init DB
    try:
        _db.init_db()
    except Exception as _dbe:
        print(f"  [DB WARN] init_db failed: {_dbe}")

    # Dead-ticker handling: either force a recheck pass or seed known-dead candidates.
    if args.recheck_dead:
        dead = _db.list_dead()
        if dead:
            for entry in dead:
                _db.clear_dead(entry["ticker"])
            print(f"  [RECHECK-DEAD] Cleared {len(dead)} entries -- will retry all this run")
        else:
            print("  [RECHECK-DEAD] Dead-ticker table is empty -- nothing to clear")
    else:
        _seed_dead_tickers(warmup, fetch_end)

    # Pre-load ALL fundamentals into memory (one bulk read replaces ~105k per-call queries)
    print("\n  Using DB-backed PIT fundamentals (market_data.db)")
    cov = _db.list_pit_coverage()
    print(f"  PIT coverage: {len(cov)} tickers in DB")
    print("  Pre-loading fundamentals into memory ...")
    t0_fund = time.time()
    global _FUND_PRELOADED
    _FUND_PRELOADED = _preload_fundamentals()
    t_fund = time.time() - t0_fund
    n_snaps = sum(len(v) for v in _FUND_PRELOADED.values())
    print(f"  Preloaded: {n_snaps} snapshots, {len(_FUND_PRELOADED)} tickers  ({t_fund:.1f}s)")

    fund_rows = {}   # kept for call-site signature compat; not used in DB path

    # Download / read prices (DB-backed with yfinance fallback)
    print(f"\n  Loading prices for {len(ALL_TICKERS)+2} tickers ...")
    t0_price = time.time()
    all_fetch   = ALL_TICKERS + ["SPY", "QQQ"]
    price_cache = download_prices(all_fetch, warmup, fetch_end)
    t_price = time.time() - t0_price
    print(f"  Price load elapsed: {t_price:.1f}s")

    spy_df = price_cache.pop("SPY", None)
    ndx_df = price_cache.pop("QQQ", None)

    if spy_df is None or spy_df.empty:
        print("  ERROR: SPY prices unavailable"); sys.exit(1)
    spy_closes = spy_df["Close"]

    if ndx_df is None or ndx_df.empty:
        print("  WARN: QQQ unavailable -- regime defaults to BULL_STRONG")
        ndx_closes = spy_closes
    else:
        ndx_closes = ndx_df["Close"]

    available = set(price_cache.keys())
    print(f"\n  Price data available: {len(price_cache)} tickers")

    # pass pit_db=None since fundamentals now come from DB directly
    pit_db = None

    # Run simulation
    t0_sim = time.time()
    closed_trades, open_positions, equity_curve, total_commissions, n_transactions, avg_exposure_by_regime = run_simulation(
        sim_start, sim_end,
        price_cache, fund_rows,
        spy_closes, ndx_closes,
        pit_db=pit_db,
    )
    t_sim = time.time() - t0_sim
    print(f"  Sim loop elapsed: {t_sim:.1f}s")

    # Compute metrics
    metrics = compute_metrics(
        closed_trades, open_positions, equity_curve,
        sim_start, sim_end, spy_closes,
        price_cache=price_cache,
        total_commissions=total_commissions,
        n_transactions=n_transactions,
        avg_exposure_by_regime=avg_exposure_by_regime,
    )
    metrics["_equity_curve"]     = equity_curve  # for XLSX sheet
    metrics["_equity_curve_ref"] = equity_curve  # for regime exposure day counts in report

    config_dict = {
        "SIM_START":               sim_start,
        "SIM_END":                 sim_end,
        "STARTING_CAPITAL":        STARTING_CAPITAL,
        "PER_BUY_FRACTION":        PER_BUY_FRACTION,
        "CONVICTION_MULT":         str(CONVICTION_MULT),
        "MAX_POSITION_PCT_EQUITY": MAX_POSITION_PCT_EQUITY,
        "MAX_POSITIONS":           MAX_POSITIONS,
        "REGIME_POSITION_MULT":    str(REGIME_POSITION_MULT),
        "REGIME_MAX_POSITIONS":    str(REGIME_MAX_POSITIONS),
        "MIN_CASH_TO_TRADE":       MIN_CASH_TO_TRADE,
        "TRAILING_STOP_PCT":       TRAILING_STOP_PCT,
        "TRAIL_ACTIVATE_GAIN_PCT": TRAIL_ACTIVATE_GAIN_PCT,
        "TAKE_PROFIT_PCT":         TAKE_PROFIT_PCT,
        "MA_CONFIRM_DAYS":         MA_CONFIRM_DAYS,
        "MA_BREAKDOWN_PCT":        MA_BREAKDOWN_PCT,
        "MAX_HOLD_DAYS":           MAX_HOLD_DAYS,
        "MIN_HOLD_DAYS":           MIN_HOLD_DAYS,
        "BELOW_MA_TREND_FLOOR":    BELOW_MA_TREND_FLOOR,
        "REPORTING_LAG_DAYS":      REPORTING_LAG_DAYS,
        "ALLOW_PYRAMIDING":        ALLOW_PYRAMIDING,
        "MAX_ADDS_PER_POSITION":   MAX_ADDS_PER_POSITION,
        "ADD_ON_MIN_GAIN_PCT":     ADD_ON_MIN_GAIN_PCT,
        "GM_EROSION_CYCLICAL_THR": GM_EROSION_CYCLICAL_THR,
        "GM_EROSION_NONCYC_THR":   GM_EROSION_NONCYC_THR,
        "MA100_BREAKDOWN_DAYS":    MA100_BREAKDOWN_DAYS,
        "REGIME_EXPOSURE_CAP":     str(REGIME_EXPOSURE_CAP),
        "COMMISSION_MIN_USD":       COMMISSION_MIN_USD,
        "COMMISSION_PER_SHARE":     COMMISSION_PER_SHARE,
    }

    print("\n  Writing reports ...")
    t0_rep = time.time()
    rpt_txt  = OUTPUT_DIR / "portfolio_report.txt"
    rpt_json = OUTPUT_DIR / "portfolio_report.json"
    rpt_xlsx = OUTPUT_DIR / "portfolio_report.xlsx"

    summary_text = write_text_report(metrics, closed_trades, open_positions, config_dict, rpt_txt)
    write_json_report(metrics, closed_trades, open_positions, config_dict, rpt_json)
    metrics["_equity_curve"] = equity_curve
    write_xlsx_report(metrics, closed_trades, config_dict, rpt_xlsx)
    append_perf_history(metrics, sim_start, sim_end)
    t_rep = time.time() - t0_rep

    print(f"  TXT   -> {rpt_txt}")

    # Print summary block to console
    print()
    for line in summary_text.split("\n"):
        if line.startswith("  TRADE LOG"):
            break
        print(line)

    # Report get_fundamentals_asof None stats
    none_count = len(_FUND_NONE_LOG)
    if none_count == 0:
        print("\n  [DB] get_fundamentals_asof returned None: 0 times")
    else:
        # Deduplicate by (ticker, date) and show the first 20 unique entries
        seen = {}
        for tk, dt in _FUND_NONE_LOG:
            seen.setdefault(dt, set()).add(tk)
        earliest = sorted(seen.keys())[:5]
        print(f"\n  [DB] get_fundamentals_asof returned None: {none_count} times")
        print("       (first occurrences by date -- expected for early sim dates)")
        for dt in earliest:
            tks = sorted(seen[dt])[:8]
            print(f"       {dt}: {', '.join(tks)}{' ...' if len(seen[dt]) > 8 else ''}")

    print("  [DB] No static-snapshot fallback (multi_sector_trend_latest.csv) was used.")

    t_total = time.time() - t0_total
    print(f"\n  PHASE TIMINGS (seconds):")
    print(f"    fund_preload : {t_fund:.1f}s")
    print(f"    price_load   : {t_price:.1f}s")
    print(f"    sim_loop     : {t_sim:.1f}s")
    print(f"    reporting    : {t_rep:.1f}s")
    print(f"    TOTAL        : {t_total:.1f}s")

    return metrics


if __name__ == "__main__":
    main()
