# tester.py V30 -- tech G8 fundamentals override 0.3->0.5, G7 high-GM escape tightened, G3 no bear adjustment
import sys, json, re, statistics
from datetime import datetime, timedelta
from pathlib import Path

import json as _json
_PARAMS_PATH = Path(__file__).resolve().parents[1] / "config" / "strategy_params.json"
def _load_params():
    with open(_PARAMS_PATH, "r") as f:
        return _json.load(f)
_P = _load_params()

import pandas as pd
try:
    import yfinance as yf
    _YF_AVAILABLE = True
except ImportError:
    _YF_AVAILABLE = False

def sf(val, default=0.0):
    try:
        if val is None: return default
        v = float(val)
        return default if (v != v) else v
    except (TypeError, ValueError):
        return default

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

PROJECT_ROOT = Path(__file__).resolve().parents[1]
MULTI_CSV  = PROJECT_ROOT / "data" / "multi_sector_trend_latest.csv"
OUTPUT_DIR = PROJECT_ROOT / "reports"
LOG_PATH   = PROJECT_ROOT / "logs" / "gate_log.txt"

PASS_THRESHOLD_SEMI_BASE    = 5.0
PASS_THRESHOLD_TECH_BASE    = 6.3
PASS_THRESHOLD_MEDTECH_BASE = 5.7
PASS_THRESHOLD_ENERGY_BASE  = 5.5
PASS_THRESHOLD_DEFAULT_BASE = 5.5

DIRECT_WEIGHT  = 1.0
PROXY_WEIGHT   = 0.90
NA_WEIGHT      = 0.0

W_G1 = 0.8
W_G2 = 1.0
W_G3 = 1.0
W_G4 = 1.0
W_G5 = 1.0
W_G6 = 1.0
W_G7 = 1.0
W_G8 = 1.5
W_G8_TECH = 0.8

MIN_DIRECT_GATES_TECH    = 3
MIN_DIRECT_GATES_MEDTECH = 2
MIN_DIRECT_GATES_DEFAULT = 2

REGIME_ADJUSTMENTS = _P["gates"]["regime_adjustments"]

_ndx_regime = "BULL_STRONG"
COUNTER_CYCLICAL = {"semi"}

# Sectors banned from entry during specific bear regimes.
# These sectors fail in bear markets due to macro factors (rates, risk-off)
# that fundamental gates cannot filter. Regime-conditional veto is the only
# reliable fix -- threshold tuning just selects slightly better bad stocks.
BEAR_BANNED_SECTORS = {
    "BEAR_VOLATILE": {
        "Cybersecurity",
        "Solar Installation",
        "Renewable Utilities",
        "CleanTech / Emerging",
        "FinTech & Payments",  # pilot add: 13.0% WR, -$3,084 P&L (was absent)
    },
    "BEAR_GRIND": {
        "Solar Installation",
        "CleanTech / Emerging",
        "FinTech & Payments",  # pilot add: 13.0% WR, -$3,084 P&L (was absent)
    },
}

def pass_threshold(universe: str) -> float:
    if universe == "semi":     base = PASS_THRESHOLD_SEMI_BASE
    elif universe == "tech":   base = PASS_THRESHOLD_TECH_BASE
    elif universe == "medtech":base = PASS_THRESHOLD_MEDTECH_BASE
    elif universe == "energy": base = PASS_THRESHOLD_ENERGY_BASE
    else:                      base = PASS_THRESHOLD_DEFAULT_BASE
    if universe in COUNTER_CYCLICAL: return base
    return base + REGIME_ADJUSTMENTS.get(_ndx_regime, 0.0)

def min_direct_gates(universe: str) -> int:
    if universe == "tech":    return MIN_DIRECT_GATES_TECH
    elif universe == "medtech": return MIN_DIRECT_GATES_MEDTECH
    return MIN_DIRECT_GATES_DEFAULT

class _Tee:
    _ansi = re.compile(r'\033\[[0-9;]*m')
    def __init__(self, console, logfile):
        self.console = console; self.logfile = logfile
    def write(self, obj):
        try: self.console.write(obj); self.console.flush()
        except (UnicodeEncodeError, UnicodeDecodeError):
            try: self.console.write(obj.encode('ascii','replace').decode('ascii')); self.console.flush()
            except Exception: pass
        try: self.logfile.write(self._ansi.sub('', obj)); self.logfile.flush()
        except Exception: pass
    def flush(self):
        try: self.console.flush()
        except Exception: pass
        try: self.logfile.flush()
        except Exception: pass

GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
CYAN   = "\033[96m"
BLUE   = "\033[94m"
MAGENTA= "\033[95m"
BOLD   = "\033[1m"
RESET  = "\033[0m"

def gate(passed: bool, note: str, proxy: bool = False, wt: float = 1.0):
    base = PROXY_WEIGHT if proxy else DIRECT_WEIGHT
    w = round(base * wt, 2)
    tag = " [P]" if proxy else ""
    return (passed, w, note + tag, proxy, None)

def gate_scored(score: float, max_score: float, note: str, proxy: bool = False):
    w = PROXY_WEIGHT if proxy else DIRECT_WEIGHT
    passed = score > 0
    return (passed, w, note, proxy, score)

def gate_na(note: str):
    return (False, NA_WEIGHT, note + " [N/A]", False, 0.0)

def gate_valuation(ps, thr: float, rev_growth: float,
                   label: str, proxy: bool = False) -> tuple:
    try: ps_f = float(ps)
    except (TypeError, ValueError): ps_f = float('nan')
    if pd.isna(ps_f) or ps_f == 999.0:
        return gate_na(f"{label} data unavailable (999/N/A) -- excluded from score")
    thr_adj = dynamic_g1_threshold(thr, rev_growth)
    ok = ps_f < thr_adj
    note = (f"P/S={ps_f:.2f} < {thr_adj:.2f} (base {thr}, rev={rev_growth:.0f}%) PASS"
            if ok else f"P/S={ps_f:.2f} >= {thr_adj:.2f} FAIL")
    return gate(ok, note, proxy=proxy, wt=W_G1)

def score_gates(gate_results):
    weighted = max_pos = 0.0
    nd = np_ = 0
    for _, (passed, weight, _, is_proxy, score_override) in gate_results.items():
        if weight == NA_WEIGHT: continue
        if score_override is not None:
            weighted += score_override
            if score_override > 0:
                max_pos += weight
            elif passed and score_override == 0:
                pass
            else:
                max_pos += weight
            if score_override > 0:
                if is_proxy: np_ += 1
                else: nd += 1
        elif passed:
            weighted += weight
            max_pos += weight
            if is_proxy: np_ += 1
            else: nd += 1
        else:
            max_pos += weight
    return weighted, max_pos, int(nd), int(np_)

def check_veto(row, sector_pct_rank: float) -> tuple:
    gm_erosion = sf(row.get("GM Erosion", 0))
    universe_hint = str(row.get("Sector", ""))
    is_cyclical = any(k in universe_hint for k in
                      ["Semi","Solar","Semiconductor","Major Proc",
                       "Memory","Foundry","Connectivity","Analog",
                       "Emerging/Small","CleanTech"])
    veto_thr = 20.0 if is_cyclical else 12.0
    if gm_erosion > veto_thr:
        return True, (f"VETO -- GM Erosion={gm_erosion:.1f}% > {veto_thr:.0f}% "
                      f"(moat collapse kill-switch)")
    # NEW: sector-specific veto gates for Cybersecurity / FinTech verticals
    _sec = str(row.get("Sector", ""))
    _g2 = sf(row.get("GM %", 0)) / 100.0  # G2 gross-margin proxy as 0-1 score
    _g3_eff = sf(row.get("Rule 40", 0))   # G3 efficiency proxy (Rule of 40)
    _ma200 = sf(row.get("Price_vs_MA200_%", -999))
    _g8_ma200_override = _ma200 > 0
    _g6_debt = sf(row.get("Share Growth %", 0)) / 100.0  # G6 debt/dilution proxy
    if _sec == "Cybersecurity":
        if ((_g2 < 0.65 or _g3_eff < 22.0) and not _g8_ma200_override):  # half-strength: was _g2<0.8 or _g3_eff<28.0
            return True, (f"SECTOR VETO (Cyber) -- G2={_g2:.2f}<0.65 or G3eff={_g3_eff:.1f}<22 "  # was 0.8 / 28
                          f"and no MA200 override")
    if _sec == "Solar Hardware":  # NEW half-strength Solar Hardware sector veto
        _g8_solar = sf(row.get("Price_vs_MA200_%", -999)) > 0  # MA200 override escape
        if ((_g2 < 0.20 or _g3_eff < 0.0) and not _g8_solar):  # half-strength of GM<40% / trend<0.5
            return True, (f"SECTOR VETO (Solar HW) -- G2={_g2:.2f}<0.20 or G3eff={_g3_eff:.1f}<0 "
                          f"and no MA200 override")
    if _sec == "FinTech & Payments":
        if (_g2 < 0.7 or _g6_debt > 0.35):  # sector veto gate
            return True, (f"SECTOR VETO (FinTech) -- G2={_g2:.2f}<0.7 or G6debt={_g6_debt:.2f}>0.35")
    # NEW: momentum-decay veto for cyclical/high-decay sectors at MED-and-below conviction
    decay, decay_reason = momentum_decay_check(row)
    if decay:
        return True, decay_reason
    # Regime-conditional sector ban: sectors that fail due to macro (rates, risk-off)
    # that fundamental gates cannot detect. Bans are defined in BEAR_BANNED_SECTORS.
    _banned = BEAR_BANNED_SECTORS.get(_ndx_regime, set())
    if _sec in _banned:
        return True, (
            f"REGIME VETO ({_ndx_regime}): sector '{_sec}' banned during this regime "
            f"(macro-driven failure; gates cannot filter rate/risk-off exposure)"
        )
    return False, ""

def momentum_decay_check(row) -> tuple:
    # Flags flattened/reversed momentum (Price/SMA-20 compressed) for higher-decay cyclical sectors.
    _sec = str(row.get("Sector", ""))
    _decay_sectors = {"Cybersecurity", "FinTech & Payments", "Data & Infrastructure"}
    if _sec not in _decay_sectors:
        return False, ""
    _sma20 = sf(row.get("SMA20", 0))
    _price = sf(row.get("Price", 0))
    if _sma20 <= 0 or _price <= 0:
        return False, ""
    _ratio = _price / _sma20
    if _ratio < 1.03:  # was 1.02 per critic revision -- more pronounced decay required
        return True, (f"MOMENTUM DECAY VETO ({_sec}) -- Price/SMA20={_ratio:.3f} < 1.03 "
                      f"(momentum flattened/reversed, cyclical inflection risk)")
    return False, ""

SECTOR_MAP = {
    "Solar Hardware":           ("energy",  "solar_hw"),
    "Solar Installation":       ("energy",  "solar_install"),
    "Renewable Utilities":      ("energy",  "renewables"),
    "CleanTech / Emerging":     ("energy",  "solar_hw"),
    "Cybersecurity":            ("tech",    "cyber"),
    "Enterprise SaaS & AI":     ("tech",    "infra_saas"),
    "Data & Infrastructure":    ("tech",    "infra_saas"),
    "Communications & Ops":     ("tech",    "infra_saas"),
    "FinTech & Payments":       ("tech",    "fintech"),
    "Medical Devices (Heavy)":  ("medtech", "surgical"),
    "Monitoring & Specialized": ("medtech", "monitoring"),
    "Diagnostics & Lab Tech":   ("medtech", "implants"),
    "Emerging & Biotech Med":   ("medtech", "monitoring"),
    "Major Processors":         ("semi",    "proc_ai"),
    "Connectivity":             ("semi",    "connectivity"),
    "Foundries":                ("semi",    "foundry_analog"),
    "Analog & Power":           ("semi",    "foundry_analog"),
    "Memory & Storage":         ("semi",    "memory_smallcap"),
    "Emerging/Small Cap":       ("semi",    "memory_smallcap"),
}

SUB_LABELS = {
    "solar_hw":        "Solar Hardware",
    "solar_install":   "Solar Installer",
    "renewables":      "Renewable Utilities",
    "cyber":           "Cybersecurity",
    "infra_saas":      "Infra & Data SaaS",
    "fintech":         "FinTech & Payments",
    "surgical":        "Surgical / Heavy MedTech",
    "monitoring":      "Patient Monitoring",
    "implants":        "Implants & Consumables",
    "proc_ai":         "Semi -- Processors & AI",
    "connectivity":    "Semi -- Connectivity & RF",
    "foundry_analog":  "Semi -- Foundries & Analog",
    "memory_smallcap": "Semi -- Memory & Small Cap",
}

def dynamic_g1_threshold(base_thr: float, rev_growth: float) -> float:
    if rev_growth > 80:  return base_thr * 2.0
    if rev_growth > 40:  return base_thr * 1.5
    if rev_growth > 20:  return base_thr * 1.2
    return base_thr

def gm_band_score(gm: float, top: float, mid: float, proxy: bool = False) -> tuple:
    w = PROXY_WEIGHT if proxy else DIRECT_WEIGHT
    if gm >= top:   return w, "full", w
    elif gm >= mid: return round(w * 0.7, 2), "mid", w
    else:           return 0.0, "fail", w

def gm_gate(gm: float, gm_erosion: float,
            top: float, mid: float, proxy: bool = False) -> tuple:
    score, band, w = gm_band_score(gm, top, mid, proxy)
    bonus = 0.0
    if gm_erosion < 0 and score > 0:
        bonus = min(0.2, w - score)
        score = round(score + bonus, 2)
    if band == "full":     note = f"GM={gm:.1f}% >= {top}% FULL"
    elif band == "mid":    note = f"GM={gm:.1f}% in {mid}-{top}% MID (x0.7)"
    else:                  note = f"GM={gm:.1f}% < {mid}% FAIL"
    if bonus > 0:
        note += f" +{bonus:.1f} delta bonus (improving, erosion={gm_erosion:+.1f}%)"
    elif gm_erosion >= 0 and score > 0:
        note += f" (erosion={gm_erosion:+.1f}% no bonus)"
    return (score > 0, w, note, proxy, score)

def gate_momentum(row, sector_pct_rank: float, wt: float = 1.5) -> tuple:
    ma100_pct  = sf(row.get("Price_vs_MA100_%", 0))
    ma200_pct  = sf(row.get("Price_vs_MA200_%", 0))
    above_ma100= ma100_pct > 0
    above_ma200= ma200_pct > 0
    ret_6m     = sf(row.get("Return_6M_%", 0))
    rs_score   = sf(row.get("Relative_Strength_Score", 50), 50.0)
    _subsec    = SECTOR_MAP.get(str(row.get("Sector", "")), (None, None))[1]  # NEW: subsector lookup for solar veto
    # Bear-regime block: during BEAR_VOLATILE or BEAR_GRIND, only stocks already
    # above MA200 qualify for entry. Partial credit for "below MA200 but positive 6M"
    # was tuned for bull conditions and produces falling-knife entries in bear markets.
    # 97.5% of 2022 losing trades fell below entry within 15 days -- this is the fix.
    if _ndx_regime in ("BEAR_VOLATILE", "BEAR_GRIND") and not above_ma200:
        w_actual = round(DIRECT_WEIGHT * wt, 2)
        return (False, w_actual,
                f"BEAR REGIME BLOCK: must be above MA200 in {_ndx_regime} "
                f"(MA200={ma200_pct:+.1f}%) -- no partial credit below MA200 in bear",
                False, 0.0)
    if (not above_ma200) and _subsec in ("solar_hw", "solar_install") and ret_6m < 16.0:  # NEW: green-energy Below-MA200 veto, ret_6m<16% (critic-revised from 18%)
        raw = 0.0; label = f"VETO: Green-energy subsector ({_subsec}) Below-MA200 entry needs 6M return >= 16% (got {ret_6m:+.1f}%)"  # NEW
        w_actual = round(DIRECT_WEIGHT * wt, 2)  # NEW
        label += f" | RS={rs_score:.0f}"  # NEW
        return (False, w_actual, label, False, 0.0)  # NEW
    if above_ma200:
        raw = 1.0; label = f"Above MA200 (+{ma200_pct:.1f}%) PASS"
    elif above_ma100:
        if rs_score >= 70:
            raw = 0.75; label = f"Above MA100(+{ma100_pct:.1f}%) below MA200({ma200_pct:+.1f}%) RS={rs_score:.0f}>=70 -- recovering 0.75pt"
        else:
            raw = 0.6; label = f"Above MA100(+{ma100_pct:.1f}%) below MA200({ma200_pct:+.1f}%) -- recovering 0.6pt"
    elif ret_6m > 0:
        if ret_6m >= 12.5 and rs_score >= 65:  # was ret_6m>0 (G8 6M floor 11%->12.5%)
            raw = 0.43; label = f"Below MAs but 6M>=12.5% ({ret_6m:+.1f}%) RS={rs_score:.0f}>=65 -- nascent 0.43pt"  # was 0.4 (+8% G8 floor)
        else:
            raw = 0.33 if ret_6m >= 12.5 else 0.0; label = f"Below MAs 6M return ({ret_6m:+.1f}%) vs 12.5% floor -- nascent {raw}pt"  # was 0.3 (G8 6M floor 11%->12.5%)
    else:
        raw = 0.0; label = f"Below MA100({ma100_pct:+.1f}%) MA200({ma200_pct:+.1f}%) 6M={ret_6m:+.1f}% -- 0pts"
    w_actual = round(DIRECT_WEIGHT * wt, 2)
    score_w = round(raw * wt, 2)
    label += f" | RS={rs_score:.0f}"
    return (raw > 0, w_actual, label, False, score_w)

def gate_momentum_tech(row, sector_pct_rank: float, wt: float = 0.8,
                       fundamentals_override: bool = False) -> tuple:
    ma100_pct  = sf(row.get("Price_vs_MA100_%", 0))
    ma200_pct  = sf(row.get("Price_vs_MA200_%", 0))
    above_ma100= ma100_pct > 0
    above_ma200= ma200_pct > 0
    ret_6m     = sf(row.get("Return_6M_%", 0))
    rs_score   = sf(row.get("Relative_Strength_Score", 50), 50.0)
    # Bear-regime block: same logic as gate_momentum() -- MA200 required in bear.
    if _ndx_regime in ("BEAR_VOLATILE", "BEAR_GRIND") and not above_ma200:
        w_actual = round(DIRECT_WEIGHT * wt, 2)
        return (False, w_actual,
                f"BEAR REGIME BLOCK: must be above MA200 in {_ndx_regime} "
                f"(MA200={ma200_pct:+.1f}%) -- tech universe, no partial credit in bear",
                False, 0.0)
    if above_ma200:
        raw = 1.0; label = f"Above MA200 (+{ma200_pct:.1f}%) PASS"
    elif above_ma100:
        if rs_score >= 70:
            raw = 0.75; label = f"Above MA100(+{ma100_pct:.1f}%) below MA200({ma200_pct:+.1f}%) RS={rs_score:.0f}>=70 -- recovering"
        else:
            raw = 0.6; label = f"Above MA100(+{ma100_pct:.1f}%) below MA200({ma200_pct:+.1f}%) -- recovering"
    elif ret_6m > 0:
        raw = 0.33; label = f"Below MAs but 6M positive ({ret_6m:+.1f}%) -- nascent"  # was 0.3 (+10% G8 floor)
    else:
        raw = 0.0; label = f"Below MAs 6M={ret_6m:+.1f}% -- 0pts"
    if ma200_pct < -15 and raw > 0.4:
        label += f" [CAPPED {raw:.2f}->0.4: MA200<-15%]"
        raw = 0.4
    # V30: fundamentals override 0.3->0.5 for BULL_WEAK/BEAR_GRIND
    # was 0.3 -- Period I TECH 38.6% accuracy, G8 blocks 79.7%
    if fundamentals_override and raw == 0.0 and _ndx_regime in ("BULL_WEAK", "BEAR_GRIND"):
        raw = 0.5  # was 0.3 -- FLYW/CPAY/NET missed at 0.3
        label += " [FUNDAMENTALS RESCUE: G2+G3+G5 strong, +0.5 partial]"
    w_actual = round(DIRECT_WEIGHT * wt, 2)
    score_w = round(raw * wt, 2)
    label += f" | RS={rs_score:.0f}"
    return (raw > 0, w_actual, label, False, score_w)

def compute_momentum_rescue(gate_results, w_score, thr, row=None, universe=""):
    if universe == "tech": return 0.0
    g8 = gate_results.get("G8 Momentum")
    if g8 is None: return 0.0
    _, weight, _, _, score_override = g8
    g8_score = score_override if score_override is not None else (weight if g8[0] else 0.0)
    bonus = 0.0
    if universe == "medtech":
        rescue_threshold = weight * 0.7; rescue_range = 1.2
    else:
        rescue_threshold = weight * 0.5; rescue_range = 1.5
    if g8_score >= rescue_threshold and w_score < thr and w_score >= (thr - rescue_range):
        gap = thr - w_score
        bonus = round(max(0.3, min(0.8, 0.8 * (1.0 - gap / rescue_range))), 2)
    return bonus

def tech_quality_kill(gate_results, rev_growth: float = 0.0, fcf_margin: float = 0.0) -> tuple:
    g3 = gate_results.get("G3 Rule of 40", (True, 0, "", False, None))
    g4 = gate_results.get("G4 Retention (NRR)", (True, 0, "", False, None))
    g5 = gate_results.get("G5 Op. Leverage", (True, 0, "", False, None))
    g3_pass = g3[0] or (g3[4] is not None and g3[4] > 0)
    g4_pass = g4[0] or (g4[4] is not None and g4[4] > 0)
    g5_pass = g5[0] or (g5[4] is not None and g5[4] > 0)
    fail_count = sum(1 for p in [g3_pass, g4_pass, g5_pass] if not p)
    if fail_count >= 2:
        failed_names = []
        if not g3_pass: failed_names.append("G3")
        if not g4_pass: failed_names.append("G4")
        if not g5_pass: failed_names.append("G5")
        return True, f"QUALITY KILL: {'+'.join(failed_names)} failed ({fail_count}/3 quality gates)"
    if not g3_pass and (rev_growth < 0 and fcf_margin < -5):
        return True, (f"QUALITY KILL: G3 failed + weak fundamentals "
                      f"(rev={rev_growth:.1f}%<0 AND fcf={fcf_margin:.1f}%<-5)")
    return False, ""

def tech_val_or_quality_check(gate_results) -> tuple:
    g1 = gate_results.get("G1 Valuation", (False, NA_WEIGHT, "", False, None))
    g3 = gate_results.get("G3 Rule of 40", (True, 0, "", False, None))
    g1_is_na = g1[1] == NA_WEIGHT
    g1_pass = g1[0] or (g1[4] is not None and g1[4] > 0)
    g3_pass = g3[0] or (g3[4] is not None and g3[4] > 0)
    if g1_is_na:
        if not g3_pass:
            return True, "VAL-OR-QUALITY BLOCK: G1=N/A and G3(R40) failed"
        return False, ""
    if not g1_pass and not g3_pass:
        return True, "VAL-OR-QUALITY BLOCK: both G1 and G3 failed"
    return False, ""

def tech_expensive_check(gate_results, w_score: float, thr: float, ps_growth: float) -> tuple:
    g1 = gate_results.get("G1 Valuation", (False, NA_WEIGHT, "", False, None))
    g1_is_na = g1[1] == NA_WEIGHT
    if g1_is_na or ps_growth == 999.0: return False, ""
    if ps_growth > 5.0 and w_score < (thr + 0.5):
        return True, (f"EXPENSIVE BLOCK: PS/Growth={ps_growth:.2f}>5.0, "
                      f"score={w_score:.2f} < {thr+0.5:.1f}")
    return False, ""

def tech_strong_arm_check(gate_results, row) -> tuple:
    pricing = str(row.get("Pricing Power", "Weak") or "Weak")
    fcf_margin = sf(row.get("FCF_Margin_%", 0))
    roic = sf(row.get("ROIC %", 0))
    g3 = gate_results.get("G3 Rule of 40", (True, 0, "", False, None))
    g3_pass = g3[0] or (g3[4] is not None and g3[4] > 0)
    if g3_pass: return False, ""
    if pricing == "Strong" or fcf_margin > 8: return False, ""
    if fcf_margin > 5: return False, ""
    g1 = gate_results.get("G1 Valuation", (False, NA_WEIGHT, "", False, None))
    g1_pass = g1[0] or (g1[4] is not None and g1[4] > 0)
    if g1_pass and g1[1] != NA_WEIGHT: return False, ""
    if roic > 8: return False, ""
    return True, (f"STRONG-ARM BLOCK: no strong signal "
                  f"(G3=fail, fcf={fcf_margin:.1f}%<=5, roic={roic:.1f}%<=8)")

def tech_ma200_check(row) -> tuple:
    ma200_pct = sf(row.get("Price_vs_MA200_%", 0))
    if ma200_pct <= 0:
        return True, f"TECH MA200 BLOCK: price vs MA200={ma200_pct:+.1f}% <= 0 (must be above)"
    return False, ""

TECH_REV_FLOOR = -5.0

def _raw_gm_erosion_is_na(row) -> bool:
    raw = row.get("GM Erosion")
    if raw is None: return True
    if isinstance(raw, float) and raw != raw: return True
    try:
        if pd.isna(raw): return True
    except (TypeError, ValueError):
        pass
    if isinstance(raw, str) and raw.strip() in ("", "nan", "<NA>", "None"):
        return True
    return False

# ============================================================================
#  UNIVERSE 1 -- ENERGY
# ============================================================================
def gates_energy(row, sub, sector_pct_rank):
    results = {}
    ps           = sf(row.get("PS_Ratio", 999), 999.0)
    ps_growth    = sf(row.get("PS/Growth", 999), 999.0)
    gm           = sf(row.get("GM %", 0))
    gm_erosion   = sf(row.get("GM Erosion", 0))
    roic         = sf(row.get("ROIC %", 0))
    inv_days     = sf(row.get("Inv Days", 0))
    inv_trend    = sf(row.get("Inv Trend", 0))
    rule40       = sf(row.get("Rule 40", 0))
    share_growth = sf(row.get("Share Growth %", 0))
    pricing      = str(row.get("Pricing Power", "Weak") or "Weak")
    rev_growth   = sf(row.get("Revenue_Growth_%", 0))
    fcf_margin   = sf(row.get("FCF_Margin_%", 0))
    erosion_is_na = _raw_gm_erosion_is_na(row)
    if sub == "solar_hw":
        results["G1 Valuation"] = gate_valuation(ps_growth, 5.0, rev_growth, "Solar HW P/S")
    elif sub == "solar_install":
        results["G1 Valuation"] = gate_valuation(ps_growth, 2.0, rev_growth, "Solar Install P/S")
    else:
        try: ps_val = float(ps)
        except (TypeError, ValueError): ps_val = float('nan')
        if pd.isna(ps_val) or ps_val == 999.0:
            results["G1 Valuation"] = gate_na("Renewables P/S data unavailable")
        else:
            ok = ps_val < 3.0
            results["G1 Valuation"] = gate(ok,
                f"P/S={ps_val:.2f} < 3.0 PASS" if ok else f"P/S={ps_val:.2f} >= 3.0 FAIL",
                proxy=False, wt=W_G1)
    gm_tops = _P["gates"]["gm_tops"]
    gm_mids = _P["gates"]["gm_mids"]
    results["G2 Gross Margin"] = gm_gate(gm, gm_erosion, gm_tops[sub], gm_mids[sub])
    if sub == "solar_hw":
        ok = roic > -5
        results["G3 Capital Eff."] = gate(ok,
            f"ROIC={roic:.1f}% > -5% PASS" if ok else f"ROIC={roic:.1f}% <= -5% FAIL")
    elif sub == "solar_install":
        ok = roic > 0 or fcf_margin > 0
        results["G3 Capital Eff."] = gate(ok,
            f"ROIC={roic:.1f}% or FCF_M={fcf_margin:.1f}% positive PASS" if ok
            else f"ROIC={roic:.1f}% FCF_M={fcf_margin:.1f}% both negative FAIL")
    else:
        ok = roic > -4 or fcf_margin > 3
        results["G3 Capital Eff."] = gate(ok,
            f"ROIC={roic:.1f}% > -4 or FCF_M={fcf_margin:.1f}% > 3 PASS" if ok
            else f"ROIC={roic:.1f}% and FCF_M={fcf_margin:.1f}% both poor FAIL")
    if sub == "solar_hw":
        ok = inv_days < 120 or inv_trend < 5
        results["G4 Inventory"] = gate(ok,
            f"Days={inv_days:.0f} trend={inv_trend:+.1f} PASS" if ok
            else f"Days={inv_days:.0f} trend={inv_trend:+.1f} FAIL")
    elif sub == "solar_install":
        ok = inv_trend <= 10
        results["G4 Inventory"] = gate(ok,
            f"Ops trend={inv_trend:+.1f} <= 10 PASS" if ok
            else f"Ops trend={inv_trend:+.1f} > 10 FAIL")
    else:
        ok = rule40 > -20
        results["G4 Ops Efficiency"] = gate(ok,
            f"Rule40={rule40:.1f} > -20 (ops proxy) PASS" if ok
            else f"Rule40={rule40:.1f} <= -20 FAIL")
    ok = rev_growth > -12
    results["G5 Growth Signal"] = gate(ok,
        f"Rev growth={rev_growth:.1f}% > -12% PASS" if ok
        else f"Rev growth={rev_growth:.1f}% <= -12% FAIL")
    thr = 10
    ok = share_growth < thr
    results["G6 Leverage"] = gate(ok,
        f"Share growth={share_growth:.2f}% < {thr}% PASS" if ok
        else f"Share growth={share_growth:.2f}% >= {thr}% FAIL")
    gm_tops_map = {"solar_hw": 35, "solar_install": 20, "renewables": 28}
    gm_floor_mult = 0.4 if sub == "solar_hw" else 0.5
    gm_floor = gm_tops_map.get(sub, 28) * gm_floor_mult
    if erosion_is_na:
        erosion_ok = False
        erosion_note = f"GM erosion=N/A (missing data, treated as FAIL)"
    else:
        erosion_ok = gm_erosion < 6
        erosion_note = f"GM erosion={gm_erosion:+.1f}% <6"
    gm_floor_ok = gm >= gm_floor
    if sub == "renewables":
        if erosion_is_na:
            ok = gm_floor_ok and (roic > 0 or fcf_margin > 0)
        elif erosion_ok:
            ok = gm_floor_ok
        else:
            ok = False
        note_logic = "OR(guarded)"
    elif sub == "solar_install":
        if erosion_is_na:
            ok = gm_floor_ok and (roic > 0 or fcf_margin > 0)
        else:
            erosion_strict_ok = gm_erosion < 4
            ok = erosion_strict_ok and gm_floor_ok
        note_logic = "AND(strict4)"
    else:
        if erosion_is_na:
            ok = gm_floor_ok and (roic > 0 or fcf_margin > 0)
        else:
            ok = erosion_ok and gm_floor_ok
        note_logic = "AND"
    results["G7 Margin Stability"] = gate(ok,
        f"{erosion_note} {note_logic} GM={gm:.1f}% >={gm_floor:.0f}% floor -- PASS" if ok
        else f"{erosion_note} (lim 6%) {note_logic} GM={gm:.1f}% (floor {gm_floor:.0f}%) FAIL")
    results["G8 Momentum"] = gate_momentum(row, sector_pct_rank, wt=W_G8)
    return results

# ============================================================================
#  UNIVERSE 2 -- TECH
# ============================================================================
def gates_tech(row, sub, sector_pct_rank):
    results = {}
    ps           = sf(row.get("PS/Growth", 999), 999.0)
    gm           = sf(row.get("GM %", 0))
    gm_erosion   = sf(row.get("GM Erosion", 0))
    roic         = sf(row.get("ROIC %", 0))
    rule40       = sf(row.get("Rule 40", 0))
    share_growth = sf(row.get("Share Growth %", 0))
    pricing      = str(row.get("Pricing Power", "Weak") or "Weak")
    rev_growth   = sf(row.get("Revenue_Growth_%", 0))
    fcf_margin   = sf(row.get("FCF_Margin_%", 0))
    base_thrs = {"cyber": 2.5, "infra_saas": 2.0, "fintech": 1.0}
    results["G1 Valuation"] = gate_valuation(
        ps, base_thrs[sub], rev_growth, f"Sales-PEG ({sub})")
    gm_configs = {k: tuple(v) for k, v in _P["gates"]["gm_configs"].items()}
    top, mid = gm_configs[sub]
    results["G2 Gross Margin"] = gm_gate(gm, gm_erosion, top, mid)
    # V30: no bear adjustment -- was -3 in bear
    # FLYW/CPAY/RAMP passed on weakened thresholds
    r40_bases = {"cyber": 33, "infra_saas": 28, "fintech": 20}
    r40_thr = r40_bases[sub]  # was r40_base-3 in bear
    ok = rule40 > r40_thr
    results["G3 Rule of 40"] = gate(ok,
        f"R40={rule40:.1f} > {r40_thr} PASS" if ok
        else f"R40={rule40:.1f} <= {r40_thr} FAIL", wt=W_G3)
    hard_floor = fcf_margin > -15
    strong_ok = (pricing == "Strong" or fcf_margin > 8)
    if not hard_floor:
        note = f"Retention HARD FAIL: fcf_m={fcf_margin:.1f}% <= -15%"
        ok = False; g4_score = 0.0
    elif strong_ok:
        note = f"Retention STRONG: pricing={pricing} fcf_m={fcf_margin:.1f}% PASS"
        ok = True; g4_score = PROXY_WEIGHT
    else:
        note = f"Retention: pricing={pricing} fcf_m={fcf_margin:.1f}% rev={rev_growth:.1f}% roic={roic:.1f}% FAIL"
        ok = False; g4_score = 0.0
    results["G4 Retention (NRR)"] = (ok, PROXY_WEIGHT, note + " [P]", True, g4_score)
    if fcf_margin < -10 and rev_growth < 5:
        ok = False
        note = f"HARD KILL: fcf_m={fcf_margin:.1f}%<-10 AND rev={rev_growth:.1f}%<5 FAIL"
    elif sub == "fintech":
        ok = pricing == "Strong" or rev_growth > 15 or fcf_margin > 5
        note = (f"Take Rate {pricing} or rev={rev_growth:.1f}%>15 or fcf={fcf_margin:.1f}%>5 PASS" if ok
                else f"Take Rate {pricing} rev={rev_growth:.1f}% fcf={fcf_margin:.1f}% FAIL")
    else:
        ok = fcf_margin > 5 or (rev_growth > 15 and fcf_margin > -5)
        note = (f"FCF_M={fcf_margin:.1f}%>5 or (rev={rev_growth:.1f}%>15+fcf>-5) PASS" if ok
                else f"FCF_M={fcf_margin:.1f}% rev={rev_growth:.1f}% FAIL")
    results["G5 Op. Leverage"] = gate(ok, note)
    dil_thrs = {"cyber": 10, "infra_saas": 8, "fintech": 5}
    thr = dil_thrs[sub]
    ok = share_growth < thr
    results["G6 Dilution"] = gate(ok,
        f"Share growth={share_growth:.2f}% < {thr}% PASS" if ok
        else f"Share growth={share_growth:.2f}% >= {thr}% FAIL")
    if sub == "fintech":
        ok = ((rev_growth > 5 and gm > 40) or (gm > 50 and rev_growth > -10) or
              pricing == "Strong" or (gm > 55 and rev_growth > 5))
        note = (f"Rev={rev_growth:.1f}% GM={gm:.1f}% pricing={pricing} PASS" if ok
                else f"Rev={rev_growth:.1f}% GM={gm:.1f}% pricing={pricing} FAIL")
        results["G7 Platform Power"] = gate(ok, note)
    else:
        # V30: tightened high-GM escape gm>68+rev>15
        # was gm>65+rev>5 -- FLYW(-32.1%) passed via old escape
        ok = (rev_growth > 5 and (gm > 50 or pricing == "Strong")) or (gm > 68 and rev_growth > 15)  # was gm>65, rev>5
        if ok and gm > 68 and rev_growth > 15 and not (gm > 50 or pricing == "Strong"):
            note = f"High-GM escape: GM={gm:.1f}%>68+rev={rev_growth:.1f}%>15 PASS"
        elif ok:
            note = f"Rev={rev_growth:.1f}%>5 + GM={gm:.1f}%>50/pricing={pricing} PASS"
        else:
            note = f"Rev={rev_growth:.1f}% GM={gm:.1f}% pricing={pricing} FAIL"
        results["G7 Platform Power"] = gate(ok, note)
    # G8 Momentum: absolute-value fundamentals override, then gate-score fallback
    g8_gm    = sf(row.get("GM %", 0))
    g8_r40   = sf(row.get("Rule 40", 0))
    g8_rev   = sf(row.get("Revenue_Growth_%", 0))
    g8_ma200 = sf(row.get("Price_vs_MA200_%", -999))
    g8_ma100 = sf(row.get("Price_vs_MA100_%", -999))
    strong_fundamentals = (g8_gm >= 65 and g8_r40 >= 30 and g8_rev >= 10)
    temporarily_weak = (g8_ma200 < 0 and g8_ma200 >= -15 and g8_ma100 >= -15)
    if strong_fundamentals and temporarily_weak:
        # V30: override score 0.5 -- was 0.4*wt
        g8_override_score = round(DIRECT_WEIGHT * W_G8_TECH * 0.5, 2)  # was 0.4
        g8_note = (f"Fundamentals override: GM={g8_gm:.0f}% R40={g8_r40:.0f} "
                   f"rev={g8_rev:.1f}% | MA200={g8_ma200:+.1f}% MA100={g8_ma100:+.1f}% "
                   f"(temporarily weak) | override=+{g8_override_score:.2f} [direct]")
        results["G8 Momentum"] = (True, DIRECT_WEIGHT * W_G8_TECH,
                                  g8_note, False, g8_override_score)
    else:
        g2_ok = results.get("G2 Gross Margin", (False,))[0]
        g3_ok = results.get("G3 Rule of 40", (False,))[0]
        g5_ok = results.get("G5 Op. Leverage", (False,))[0]
        g2_sc  = (results.get("G2 Gross Margin", (False, 0, "", False, 0.0))[4] or 0.0)
        g3_so  = results.get("G3 Rule of 40", (False, 0, "", False, None))[4]
        g3_wt  = results.get("G3 Rule of 40", (False, 0, "", False, None))[1]
        g3_eff = g3_so if g3_so is not None else (g3_wt if g3_ok else 0.0)
        fundamentals_override = (g2_ok and g2_sc >= 0.9 and g3_ok and g3_eff >= 0.9 and g5_ok)
        results["G8 Momentum"] = gate_momentum_tech(row, sector_pct_rank, wt=W_G8_TECH,
                                                     fundamentals_override=fundamentals_override)
    return results

# ============================================================================
#  UNIVERSE 3 -- MEDTECH
# ============================================================================
def gates_medtech(row, sub, sector_pct_rank):
    results = {}
    ps           = sf(row.get("PS/Growth", 999), 999.0)
    gm           = sf(row.get("GM %", 0))
    gm_erosion   = sf(row.get("GM Erosion", 0))
    roic         = sf(row.get("ROIC %", 0))
    inv_trend    = sf(row.get("Inv Trend", 0))
    rule40       = sf(row.get("Rule 40", 0))
    share_growth = sf(row.get("Share Growth %", 0))
    pricing      = str(row.get("Pricing Power", "Weak") or "Weak")
    rev_growth   = sf(row.get("Revenue_Growth_%", 0))
    fcf_margin   = sf(row.get("FCF_Margin_%", 0))
    base_thrs = {"surgical": 2.5, "monitoring": 2.0, "implants": 2.0}
    results["G1 Valuation"] = gate_valuation(
        ps, base_thrs[sub], rev_growth, f"MedTech P/S ({sub})", proxy=False)
    gm_configs = {"surgical": (65, 50), "monitoring": (70, 55), "implants": (55, 45)}  # was implants (60, 45) -- Diagnostics & Lab Tech pilot floor raise per critic one-sector-at-a-time
    top, mid = gm_configs[sub]
    results["G2 Gross Margin"] = gm_gate(gm, gm_erosion, top, mid)
    roic_thr = {"surgical": 12, "monitoring": 10, "implants": 10}[sub]
    ok = roic > roic_thr
    results["G3 ROIC"] = gate(ok,
        f"ROIC={roic:.1f}% > {roic_thr}% PASS" if ok else f"ROIC={roic:.1f}% <= {roic_thr}% FAIL")
    r40_thr = {"surgical": 10, "monitoring": 15, "implants": 5}[sub]
    ok = rule40 > r40_thr or roic > 20
    if roic > 20 and not (rule40 > r40_thr):
        note = f"ROIC={roic:.1f}%>20 bypass (R40={rule40:.1f} <= {r40_thr}) PASS"
    elif ok:
        note = f"R40={rule40:.1f} > {r40_thr} (R&D proxy) PASS"
    else:
        note = f"R40={rule40:.1f} <= {r40_thr} ROIC={roic:.1f}% FAIL"
    results["G4 Innovation (R&D)"] = gate(ok, note, proxy=True)
    neg_count = 0
    if rev_growth <= -20: neg_count += 1
    if roic <= 0:         neg_count += 1
    if fcf_margin <= 0:   neg_count += 1
    if neg_count >= 2:
        ok = False
        note = f"Hard fail: {neg_count}/3 negatives (rev={rev_growth:.1f}% roic={roic:.1f}% fcf={fcf_margin:.1f}%)"
    else:
        ok = rev_growth > 2 or (roic > 15 and rev_growth > -5)
        note = (f"Rev growth={rev_growth:.1f}% > 2% or (ROIC={roic:.1f}%>15 + rev>-5) -- recurring model PASS" if ok
                else f"Rev growth={rev_growth:.1f}% <= 2% ROIC={roic:.1f}% FAIL")
    results["G5 Recurring Rev."] = gate(ok, note)
    reg_thr = {"surgical": 5, "monitoring": 8, "implants": 5}[sub]
    ok = share_growth < reg_thr
    results["G6 Regulatory Risk"] = gate(ok,
        f"Share growth={share_growth:.2f}% < {reg_thr}% (FDA proxy) PASS" if ok
        else f"Share growth={share_growth:.2f}% >= {reg_thr}% FAIL", proxy=True)
    inv_thr = 25 if sub == "implants" else 15
    _gm_erosion_na = _raw_gm_erosion_is_na(row)
    moat_quality = (pricing == "Strong") or (not _gm_erosion_na and gm_erosion < 2)
    moat_ops = inv_trend < inv_thr
    if _gm_erosion_na:
        ok = moat_ops and roic > 15
        moat_note_extra = " [erosion=NA: require ops+ROIC>15]"
    else:
        ok = moat_quality and moat_ops
        moat_note_extra = ""
    results["G7 Market Moat"] = gate(ok,
        f"Moat: quality(erosion={gm_erosion:+.1f}<2)={moat_quality} ops(inv={inv_trend:+.1f}<{inv_thr})={moat_ops}{moat_note_extra} PASS" if ok
        else f"Moat FAIL: quality(erosion={gm_erosion:+.1f})={moat_quality} ops(inv={inv_trend:+.1f}<{inv_thr})={moat_ops}{moat_note_extra}",
        proxy=True)
    results["G8 Momentum"] = gate_momentum(row, sector_pct_rank, wt=W_G8)
    return results

# ============================================================================
#  UNIVERSE 4 -- SEMICONDUCTORS
# ============================================================================
def gates_semi(row, sub, sector_pct_rank):
    results = {}
    ps           = sf(row.get("PS/Growth", 999), 999.0)
    ps_ratio     = sf(row.get("PS_Ratio", 999), 999.0)
    gm           = sf(row.get("GM %", 0))
    gm_erosion   = sf(row.get("GM Erosion", 0))
    roic         = sf(row.get("ROIC %", 0))
    inv_days     = sf(row.get("Inv Days", 0))
    inv_trend    = sf(row.get("Inv Trend", 0))
    rule40       = sf(row.get("Rule 40", 0))
    share_growth = sf(row.get("Share Growth %", 0))
    pricing      = str(row.get("Pricing Power", "Weak") or "Weak")
    rev_growth   = sf(row.get("Revenue_Growth_%", 0))
    ma200_pct    = sf(row.get("Price_vs_MA200_%", 0))
    ma100_pct    = sf(row.get("Price_vs_MA100_%", 0))
    ret_6m       = sf(row.get("Return_6M_%", 0))
    capex_sales  = sf(row.get("Capex_Sales_%", 0))
    fcf_margin   = sf(row.get("FCF_Margin_%", 0))
    rs_csv       = sf(row.get("Relative_Strength_Score", 50), 50.0)
    if sub == "proc_ai":
        results["G1 Valuation"] = gate_valuation(ps, 2.2, rev_growth, "Semi Proc/AI Sales-PEG")
    elif sub == "connectivity":
        results["G1 Valuation"] = gate_valuation(ps, 2.2, rev_growth, "Semi Connectivity Sales-PEG")
    elif sub == "foundry_analog":
        try: ps_val = float(ps_ratio)
        except (TypeError, ValueError): ps_val = float('nan')
        if pd.isna(ps_val) or ps_val == 999.0:
            results["G1 Valuation"] = gate_na("Semi Foundry/Analog P/S data unavailable")
        else:
            ok = ps_val < 5.0
            results["G1 Valuation"] = gate(ok,
                f"P/S={ps_val:.2f} < 5.0 PASS" if ok else f"P/S={ps_val:.2f} >= 5.0 FAIL", wt=W_G1)
    else:
        results["G1 Valuation"] = gate_valuation(
            ps, 1.5, rev_growth, "Semi Memory/SmallCap P/S (P/B proxy)", proxy=True)
    gm_configs = {
        "proc_ai": (60, 45), "connectivity": (50, 35),
        "foundry_analog": (45, 30), "memory_smallcap": (30, 15),
    }
    top, mid = gm_configs[sub]
    results["G2 Gross Margin"] = gm_gate(gm, gm_erosion, top, mid)
    if sub == "proc_ai":
        ok = inv_trend <= 10 or inv_days < 130
        results["G3 Inventory"] = gate(ok,
            f"Inv trend={inv_trend:+.1f} days={inv_days:.0f} -- stable/down PASS" if ok
            else f"Inv trend={inv_trend:+.1f} days={inv_days:.0f} -- rising FAIL", wt=W_G3)
    elif sub == "connectivity":
        ok = inv_trend <= 15 or inv_days < 150
        results["G3 Inventory"] = gate(ok,
            f"Inv trend={inv_trend:+.1f} days={inv_days:.0f} -- stable/down PASS" if ok
            else f"Inv trend={inv_trend:+.1f} days={inv_days:.0f} -- rising FAIL", wt=W_G3)
    elif sub == "foundry_analog":
        ok = inv_days < 250 or inv_trend < 0
        results["G3 Inventory"] = gate(ok,
            f"Days={inv_days:.0f} trend={inv_trend:+.1f} (<250 or declining) PASS" if ok
            else f"Days={inv_days:.0f} trend={inv_trend:+.1f} FAIL", wt=W_G3)
    else:
        ok = inv_days < 160 or inv_trend < 0
        results["G3 Inventory"] = gate(ok,
            f"Days={inv_days:.0f} trend={inv_trend:+.1f} PASS (<160 or declining)" if ok
            else f"Days={inv_days:.0f} trend={inv_trend:+.1f} FAIL")
    if sub == "proc_ai":
        ok = roic > 3
        results["G4 FCF Conversion"] = gate(ok,
            f"ROIC={roic:.1f}% > 3% PASS" if ok else f"ROIC={roic:.1f}% <= 3% FAIL", proxy=True)
    elif sub == "connectivity":
        ok = roic > 5
        results["G4 FCF Conversion"] = gate(ok,
            f"ROIC={roic:.1f}% > 5% PASS" if ok else f"ROIC={roic:.1f}% <= 5% FAIL", proxy=True)
    elif sub == "foundry_analog":
        ok = roic > 7
        results["G4 FCF Conversion"] = gate(ok,
            f"ROIC={roic:.1f}% > 7% PASS" if ok else f"ROIC={roic:.1f}% <= 7% FAIL", proxy=True)
    else:
        ok = roic > -10 and gm_erosion < 10
        results["G4 FCF Conversion"] = gate(ok,
            f"ROIC={roic:.1f}% > -10 + erosion={gm_erosion:+.1f}% < 10% PASS" if ok
            else f"ROIC={roic:.1f}% erosion={gm_erosion:+.1f}% FAIL", proxy=True)
    if sub == "proc_ai":
        ok = rev_growth > 5 or pricing == "Strong"
        results["G5 R&D Efficiency"] = gate(ok,
            f"Rev growth={rev_growth:.1f}% > 5% or pricing={pricing} PASS" if ok
            else f"Rev growth={rev_growth:.1f}% <= 5% + pricing={pricing} FAIL", proxy=True)
    elif sub == "connectivity":
        ok = rev_growth > 0 or pricing == "Strong" or roic > 8
        results["G5 R&D Efficiency"] = gate(ok,
            f"Rev growth={rev_growth:.1f}%>0 or pricing={pricing} or ROIC={roic:.1f}%>8 PASS" if ok
            else f"Rev growth={rev_growth:.1f}% pricing={pricing} ROIC={roic:.1f}% FAIL", proxy=True)
    elif sub == "foundry_analog":
        ok = rev_growth > -15 or pricing == "Strong" or roic > 8
        results["G5 R&D Efficiency"] = gate(ok,
            f"Rev growth={rev_growth:.1f}%>-15 or pricing={pricing} or ROIC={roic:.1f}%>8 PASS" if ok
            else f"Rev growth={rev_growth:.1f}% pricing={pricing} ROIC={roic:.1f}% FAIL")
    else:
        ok = pricing == "Strong" or rev_growth > 0 or roic > 5
        results["G5 R&D Efficiency"] = gate(ok,
            f"Design wins proxy: pricing={pricing} rev={rev_growth:.1f}% roic={roic:.1f}% PASS" if ok
            else f"Design wins proxy FAIL (pricing={pricing} rev={rev_growth:.1f}% roic={roic:.1f}%) FAIL",
            proxy=True)
    if sub == "proc_ai":
        ok = share_growth < 5
        results["G6 Dilution"] = gate(ok,
            f"Share growth={share_growth:.2f}% < 5% PASS" if ok
            else f"Share growth={share_growth:.2f}% >= 5% FAIL")
    elif sub == "connectivity":
        ok = share_growth < 5
        results["G6 Dilution"] = gate(ok,
            f"Share growth={share_growth:.2f}% < 5% PASS" if ok
            else f"Share growth={share_growth:.2f}% >= 5% FAIL")
    elif sub == "foundry_analog":
        ok = share_growth < 3
        results["G6 Dilution"] = gate(ok,
            f"Share growth={share_growth:.2f}% < 3% PASS" if ok
            else f"Share growth={share_growth:.2f}% >= 3% FAIL", proxy=True)
    else:
        ok = share_growth < 15
        results["G6 Dilution"] = gate(ok,
            f"Share growth={share_growth:.2f}% < 15% PASS" if ok
            else f"Share growth={share_growth:.2f}% >= 15% FAIL", proxy=True)
    if sub == "proc_ai":
        ok = pricing == "Strong"
        results["G7 Moat/Pricing"] = gate(ok,
            "ASP trend Strong -- pricing power PASS" if ok else "ASP Weak FAIL")
    elif sub == "connectivity":
        ok = pricing == "Strong" or rev_growth > 5 or gm > 46
        results["G7 Moat/Pricing"] = gate(ok,
            f"Connectivity: pricing={pricing} rev={rev_growth:.1f}%>5 gm={gm:.1f}%>46 PASS" if ok
            else f"ASP {pricing} rev={rev_growth:.1f}% gm={gm:.1f}% FAIL")
    elif sub == "foundry_analog":
        ok = rev_growth > 5 or (pricing == "Strong") or gm > 55
        results["G7 Moat/Pricing"] = gate(ok,
            f"Market share: rev={rev_growth:.1f}% >5 or gm={gm:.1f}%>55 or pricing PASS" if ok
            else f"Market share FAIL (rev={rev_growth:.1f}% gm={gm:.1f}% pricing={pricing}) FAIL",
            proxy=True)
    else:
        ok = pricing == "Strong" or rev_growth > 5 or roic > 8
        results["G7 Moat/Pricing"] = gate(ok,
            f"IP proxy: pricing={pricing} rev={rev_growth:.1f}% roic={roic:.1f}% PASS" if ok
            else f"IP proxy FAIL (pricing={pricing} rev={rev_growth:.1f}% roic={roic:.1f}%) FAIL",
            proxy=True)
    if sub == "proc_ai":
        above_ma200 = ma200_pct > 0
        rs_strong = rs_csv >= 60.0
        if above_ma200 and rs_strong:
            score = 1.0; note = f"RS={rs_csv:.0f}% >= 60 AND above MA200 (+{ma200_pct:.1f}%) PASS"
        elif above_ma200:
            score = 0.7; note = f"Above MA200 (+{ma200_pct:.1f}%) RS={rs_csv:.0f}% <60 PARTIAL"
        elif rs_strong:
            score = 0.5; note = f"RS={rs_csv:.0f}% >= 60 but below MA200 ({ma200_pct:+.1f}%) PARTIAL"
        elif ret_6m > 0:
            score = 0.33; note = f"Below MA200 but 6M positive ({ret_6m:+.1f}%) RS={rs_csv:.0f}% -- nascent"  # was 0.3 (+10% G8 floor)
        else:
            score = 0.0; note = f"RS={rs_csv:.0f}% < 60 AND below MA200 ({ma200_pct:+.1f}%) FAIL"
        w_actual = round(DIRECT_WEIGHT * W_G8, 2)
        results["G8 Momentum"] = (score > 0, w_actual, note, False, round(score * W_G8, 2))
    elif sub == "connectivity":
        above_ma200 = ma200_pct > 0
        rs_strong = rs_csv >= 60.0
        if above_ma200 and rs_strong:
            score = 1.0; note = f"RS={rs_csv:.0f}% >= 60 AND above MA200 (+{ma200_pct:.1f}%) PASS"
        elif above_ma200:
            score = 0.7; note = f"Above MA200 (+{ma200_pct:.1f}%) RS={rs_csv:.0f}% <60 PARTIAL"
        elif rs_strong:
            score = 0.5; note = f"RS={rs_csv:.0f}% >= 60 but below MA200 ({ma200_pct:+.1f}%) PARTIAL"
        elif ret_6m > 0:
            score = 0.33; note = f"Below MA200 but 6M positive ({ret_6m:+.1f}%) RS={rs_csv:.0f}% -- nascent"  # was 0.3 (+10% G8 floor)
        else:
            score = 0.0; note = f"RS={rs_csv:.0f}% < 60 AND below MA200 ({ma200_pct:+.1f}%) FAIL"
        w_actual = round(DIRECT_WEIGHT * W_G8, 2)
        results["G8 Momentum"] = (score > 0, w_actual, note, False, round(score * W_G8, 2))
    elif sub == "foundry_analog":
        results["G8 Momentum"] = gate_momentum(row, sector_pct_rank, wt=W_G8)
    else:
        if ret_6m >= 11:  # was 10 (+10% G8 floor)
            raw = 1.0; note = f"6M return={ret_6m:+.1f}% >= 11% (strong momentum) PASS"
        elif ret_6m > 0:
            raw = 0.5; note = f"6M return={ret_6m:+.1f}% > 0 (weak positive) PARTIAL"
        else:
            raw = 0.0; note = f"6M return={ret_6m:+.1f}% <= 0 FAIL"
        w_actual = round(DIRECT_WEIGHT * W_G8, 2)
        score_w = round(raw * W_G8, 2)
        results["G8 Momentum"] = (raw > 0, w_actual, note, False, score_w)
    return results

def run_gates(row, sector_pct_rank):
    sector = row["Sector"]
    mapping = SECTOR_MAP.get(sector)
    if mapping is None: return None, None, None
    universe, sub = mapping
    if universe == "energy":  return universe, sub, gates_energy(row, sub, sector_pct_rank)
    elif universe == "tech":  return universe, sub, gates_tech(row, sub, sector_pct_rank)
    elif universe == "medtech": return universe, sub, gates_medtech(row, sub, sector_pct_rank)
    elif universe == "semi":  return universe, sub, gates_semi(row, sub, sector_pct_rank)
    return None, None, None

def compute_sector_percentiles(df):
    ranks = {}
    for sector, group in df.groupby("Sector"):
        col = "Return_6M_%"
        if col not in group.columns:
            for _, row in group.iterrows(): ranks[row["Ticker"]] = 50.0
            continue
        vals = pd.to_numeric(group[col], errors="coerce").fillna(float('nan'))
        for idx, row in group.iterrows():
            try:
                v = float(row[col])
                if v != v: ranks[row["Ticker"]] = 50.0; continue
                pct = float((vals < v).sum()) / max(len(vals) - 1, 1) * 100
                ranks[row["Ticker"]] = round(pct, 1)
            except (TypeError, ValueError):
                ranks[row["Ticker"]] = 50.0
    return ranks

def compute_confidence(n_direct, n_proxy, w_score, thr, universe):
    margin = w_score - thr
    if n_proxy == 0 and margin >= _P["gates"]["conviction_thresholds"]["high_margin"]:
        return "HIGH"
    if n_proxy <= 2 and (margin >= _P["gates"]["conviction_thresholds"]["med_margin"] or n_direct >= 4):
        return "MED"
    return "LOW"

def build_record(row, universe, sub, gate_results, price_chg,
                  veto=False, veto_reason="", sector_pct_rank=50.0):
    w_score, w_max, n_direct, n_proxy = score_gates(gate_results)
    thr = pass_threshold(universe)
    rescue = compute_momentum_rescue(gate_results, w_score, thr, row=row, universe=universe)
    w_score_adj = w_score + rescue

    _non_g8_failures = sum(
        1 for gname, gval in gate_results.items()
        if gname != "G8 Momentum"
        and gval[1] > 0.0
        and not gval[0]
        and (gval[4] is None or gval[4] == 0.0)
    )
    if _non_g8_failures >= 2:
        w_score_adj = round(w_score_adj - 0.3, 2)
    min_dg = min_direct_gates(universe)
    direct_gate_ok = n_direct >= min_dg
    quality_killed = False; quality_reason = ""
    rev_growth_val = sf(row.get("Revenue_Growth_%", 0))
    fcf_margin_val = sf(row.get("FCF_Margin_%", 0))
    if universe == "tech":
        quality_killed, quality_reason = tech_quality_kill(
            gate_results, rev_growth=rev_growth_val, fcf_margin=fcf_margin_val)
    val_quality_blocked = False; val_quality_note = ""
    if universe == "tech" and not quality_killed:
        val_quality_blocked, val_quality_note = tech_val_or_quality_check(gate_results)
    expensive_blocked = False; expensive_note = ""
    ps_growth_val = sf(row.get("PS/Growth", 999), 999.0)
    if universe == "tech" and not quality_killed and not val_quality_blocked:
        expensive_blocked, expensive_note = tech_expensive_check(
            gate_results, w_score_adj, thr, ps_growth_val)
    strong_arm_blocked = False; strong_arm_note = ""
    if (universe == "tech" and not quality_killed and not val_quality_blocked
            and not expensive_blocked):
        strong_arm_blocked, strong_arm_note = tech_strong_arm_check(gate_results, row)
    ma200_blocked = False; ma200_note = ""
    if universe == "tech":
        ma200_blocked, ma200_note = tech_ma200_check(row)
    tech_rev_blocked = False; tech_rev_note = ""
    if universe == "tech" and rev_growth_val < TECH_REV_FLOOR:
        tech_rev_blocked = True
        tech_rev_note = f"TECH REV FLOOR: rev={rev_growth_val:.1f}% < {TECH_REV_FLOOR}%"

    strategy_ok = ((w_score_adj >= thr) and not veto and direct_gate_ok
                   and not quality_killed and not val_quality_blocked
                   and not expensive_blocked and not strong_arm_blocked
                   and not ma200_blocked
                   and not tech_rev_blocked)
    if veto:             outcome = "VETO-FAIL"
    elif strategy_ok:    outcome = "WIN" if price_chg >= 0 else "LOSS"
    else:                outcome = "WIN (contrarian)" if price_chg < 0 else "LOSS (contrarian)"
    confidence = compute_confidence(n_direct, n_proxy, w_score_adj, thr, universe)
    if veto: confidence = "VETO"
    if universe == "energy" and strategy_ok and confidence != "VETO":
        g8_res = gate_results.get("G8 Momentum")
        if g8_res is not None:
            g8_passed, g8_weight, _, _, g8_override = g8_res
            g8_actual = g8_override if g8_override is not None else (g8_weight if g8_passed else 0.0)
            if g8_actual == 0.0:
                if _ndx_regime == "BULL_WEAK":
                    veto = True
                    veto_reason = "V26 VETO: energy G8=0 in BULL_WEAK (no price momentum, below MA200)"
                    strategy_ok = False
                    outcome = "VETO-FAIL"
                    confidence = "VETO"
                else:
                    confidence = "LOW"
    if universe == "medtech" and strategy_ok and _ndx_regime == "BULL_WEAK" and confidence != "VETO":
        g8_res_mt = gate_results.get("G8 Momentum")
        if g8_res_mt is not None:
            _g8p, _g8w, _, _, _g8ov = g8_res_mt
            _g8act = _g8ov if _g8ov is not None else (_g8w if _g8p else 0.0)
            if _g8act == 0.0:
                veto = True
                veto_reason = "V27 VETO: medtech G8=0 in BULL_WEAK (no price momentum)"
                strategy_ok = False
                outcome = "VETO-FAIL"
                confidence = "VETO"
    if universe == "semi" and strategy_ok and not veto:
        _semi_rev = sf(row.get("Revenue_Growth_%", 0))
        _semi_r40 = sf(row.get("Rule 40", 0))
        if _semi_rev < -30 and _semi_r40 < 0:
            veto = True
            veto_reason = f"V28 VETO: semi catastrophic decline (Rev={_semi_rev:.1f}%<-30, R40={_semi_r40:.1f}<0)"
            strategy_ok = False
            outcome = "VETO-FAIL"
            confidence = "VETO"
    gate_flat = {}
    for gname, (passed, weight, note, is_proxy, score_override) in gate_results.items():
        clean = re.sub(r'\s*\[(?:P|N/A|proxy|direct)\s*[^\]]*\]', '', note)
        actual_score = score_override if score_override is not None else (weight if passed else 0.0)
        gate_flat[gname] = {
            "passed": passed, "weight": weight, "is_proxy": is_proxy,
            "is_na": weight == NA_WEIGHT, "score": round(actual_score, 2), "note": clean,
        }
    fail_reasons = []
    if not direct_gate_ok: fail_reasons.append(f"min_direct={min_dg} (had {n_direct})")
    if quality_killed:      fail_reasons.append(quality_reason)
    if val_quality_blocked: fail_reasons.append(val_quality_note)
    if expensive_blocked:   fail_reasons.append(expensive_note)
    if strong_arm_blocked:  fail_reasons.append(strong_arm_note)
    if ma200_blocked:       fail_reasons.append(ma200_note)
    if tech_rev_blocked:    fail_reasons.append(tech_rev_note)
    return {
        "ticker": str(row["Ticker"]), "sector": str(row["Sector"]),
        "universe": universe, "industry": SUB_LABELS.get(sub, sub),
        "ps_growth": sf(row.get("PS/Growth", 999), 999.0),
        "gm_pct": sf(row.get("GM %", 0)), "gm_erosion": sf(row.get("GM Erosion", 0)),
        "roic_pct": sf(row.get("ROIC %", 0)), "inv_days": sf(row.get("Inv Days", 0)),
        "inv_trend": sf(row.get("Inv Trend", 0)), "rule40": sf(row.get("Rule 40", 0)),
        "share_growth": sf(row.get("Share Growth %", 0)),
        "pricing_power": str(row.get("Pricing Power", "Weak") or "Weak"),
        "rev_growth": sf(row.get("Revenue_Growth_%", 0)),
        "price_vs_ma200": sf(row.get("Price_vs_MA200_%", 0)),
        "return_6m": sf(row.get("Return_6M_%", 0)),
        "price_change": price_chg,
        "weighted_score": round(w_score_adj, 2),
        "weighted_score_raw": round(w_score, 2),
        "rescue_bonus": round(rescue, 2),
        "max_score": round(w_max, 2),
        "pass_threshold": thr,
        "gates_direct_passed": n_direct, "gates_proxy_passed": n_proxy,
        "min_direct_required": min_dg, "direct_gate_ok": direct_gate_ok,
        "quality_killed": quality_killed, "quality_reason": quality_reason,
        "val_quality_blocked": val_quality_blocked, "val_quality_note": val_quality_note,
        "expensive_blocked": expensive_blocked, "expensive_note": expensive_note,
        "strong_arm_blocked": strong_arm_blocked, "strong_arm_note": strong_arm_note,
        "ma200_blocked": ma200_blocked, "ma200_note": ma200_note,
        "tech_rev_blocked": tech_rev_blocked, "tech_rev_note": tech_rev_note,
        "strategy_passed": strategy_ok, "confidence": confidence,
        "outcome": outcome, "veto": veto, "veto_reason": veto_reason,
        "sector_pct_rank": sector_pct_rank,
        "gates": gate_flat, "_fail_reasons": fail_reasons,
    }

# ============================================================================
#  EXCEL STYLES
# ============================================================================
C = {
    "win_bg":"C6EFCE","win_fg":"276221","loss_bg":"FFC7CE","loss_fg":"9C0006",
    "cwin_bg":"FFEB9C","cwin_fg":"9C6500","closs_bg":"F4CCCC","closs_fg":"CC0000",
    "pass_cell":"D9EAD3","fail_cell":"FCE5CD","proxy_cell":"EAD1DC",
    "na_cell":"E8E8E8","partial_cell":"FFF2CC",
    "hdr_dark":"1F3864","hdr_mid":"2F5496","white":"FFFFFF",
    "gray_light":"F2F2F2","border":"BFBFBF","momentum":"D0E4F7",
}
_thin = Side(style="thin", color=C["border"])
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(size=10, color="000000", bold=False, italic=False):
    return Font(name="Arial", size=size, color=color, bold=bold, italic=italic)
def _align(h="center", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _hdr(ws, row, values, bg=None, size=10):
    bg = bg or C["hdr_mid"]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font=_font(size,C["white"],bold=True); c.fill=_fill(bg)
        c.alignment=_align("center",wrap=True); c.border=_BORDER
def _cw(ws, widths):
    for ltr, w in widths.items(): ws.column_dimensions[ltr].width = w

OUTCOME_BG = {
    "WIN":("C6EFCE","276221"),"LOSS":("FFC7CE","9C0006"),
    "WIN (contrarian)":("FFEB9C","9C6500"),"LOSS (contrarian)":("F4CCCC","CC0000"),
}
CONF_BG = {"HIGH":"C6EFCE","MED":"FFEB9C","LOW":"FFC7CE"}

def _gate_cell_bg(g):
    if g.get("is_na"): return C["na_cell"]
    if not g.get("passed"): return C["fail_cell"]
    score = g.get("score", 0); weight = g.get("weight", 1)
    if weight > 0 and score < weight: return C["partial_cell"]
    if g.get("is_proxy"): return C["proxy_cell"]
    return C["pass_cell"]

def _gate_label(g):
    if g.get("is_na"): return "N/A"
    score = g.get("score", 0); weight = g.get("weight", 1)
    passed = g.get("passed", False)
    suffix = " [P]" if g.get("is_proxy") else ""
    if not passed: return f"FAIL{suffix}"
    if weight > 0 and score < weight: return f"{score:.1f}/{weight:.1f}{suffix}"
    return f"PASS{suffix}"

def write_excel(records, out_path):
    wb = Workbook(); wb.remove(wb.active)
    _sheet_summary(wb, records); _sheet_detail(wb, records)
    _sheet_matrix(wb, records); _sheet_raw(wb, records)
    wb.save(out_path); print(f"  Excel saved  -> {out_path}")

def _sheet_summary(wb, records):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "70AD47"; ws.freeze_panes = "A5"
    ws.merge_cells("A1:P1")
    t=ws["A1"]; t.value=f"Gate Tester V30  |  NASDAQ: {_ndx_regime.upper()}  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    t.font=_font(13,C["white"],bold=True); t.fill=_fill(C["hdr_dark"])
    t.alignment=_align("center"); ws.row_dimensions[1].height=26
    ws.merge_cells("A2:P2")
    s=ws["A2"]
    s.value=(f"V30: G8 tech={W_G8_TECH} | thr: Semi={pass_threshold('semi'):.1f} Tech={pass_threshold('tech'):.1f} "
             f"MedTech={pass_threshold('medtech'):.1f} Energy={pass_threshold('energy'):.1f}")
    s.font=_font(9,C["white"],italic=True); s.fill=_fill(C["hdr_mid"])
    s.alignment=_align("center"); ws.row_dimensions[2].height=16
    hdrs=["Ticker","Universe","Sector","Industry","Score","/ Max",
          "Passed","Confidence","Return %","Outcome",
          "Direct","Proxy","Rev Growth %","vs MA200 %","6M Return %","Pricing"]
    _hdr(ws,3,hdrs); ws.row_dimensions[3].height=20
    for i,rec in enumerate(records,4):
        bg=C["gray_light"] if i%2==0 else C["white"]
        vals=[rec["ticker"],rec["universe"],rec["sector"],rec["industry"],
              rec["weighted_score"],rec["max_score"],
              "PASSED" if rec["strategy_passed"] else "FAILED",
              rec["confidence"],rec["price_change"]/100,rec["outcome"],
              rec["gates_direct_passed"],rec["gates_proxy_passed"],
              rec["rev_growth"],rec["price_vs_ma200"],rec["return_6m"],rec["pricing_power"]]
        for j,v in enumerate(vals,1):
            c=ws.cell(row=i,column=j,value=v); c.border=_BORDER; c.font=_font(9); c.fill=_fill(bg)
            c.alignment=_align("left" if j<=4 else "center")
            if j==5:
                c.number_format="0.00"
                c.fill=_fill(C["pass_cell"] if float(v)>=rec.get("pass_threshold",5.0) else C["fail_cell"])
            elif j==6: c.number_format="0.00"
            elif j==7:
                c.fill=_fill(C["pass_cell"] if rec["strategy_passed"] else C["fail_cell"])
                c.font=_font(9,bold=True)
            elif j==8: c.fill=_fill(CONF_BG.get(v,bg)); c.font=_font(9,bold=True)
            elif j==9:
                c.number_format="+0.00%;-0.00%"
                c.fill=_fill(C["pass_cell"] if rec["price_change"]>=0 else C["fail_cell"])
            elif j==10:
                obg,ofg=OUTCOME_BG.get(v,(bg,"000000"))
                c.fill=_fill(obg); c.font=_font(9,ofg,bold=True)
            elif j==14:
                c.number_format="+0.0%;-0.0%"; c.value=v/100
                c.fill=_fill(C["pass_cell"] if v>=0 else C["fail_cell"])
            elif j==15:
                c.number_format="+0.0%;-0.0%"; c.value=v/100
                c.fill=_fill(C["pass_cell"] if v>=0 else C["fail_cell"])
        ws.row_dimensions[i].height=15
    _cw(ws,{"A":7,"B":8,"C":22,"D":22,"E":9,"F":8,"G":9,"H":9,
            "I":10,"J":18,"K":8,"L":8,"M":11,"N":11,"O":11,"P":10})
    ws.auto_filter.ref=f"A3:{get_column_letter(len(hdrs))}3"

def _sheet_detail(wb, records):
    ws=wb.create_sheet("Full Detail"); ws.sheet_properties.tabColor="4472C4"; ws.freeze_panes="D4"
    gate_names=list(records[0]["gates"].keys()) if records else []
    base_hdrs=["Ticker","Sector","Industry","Score","Max","Passed",
               "Confidence","Return %","Outcome","Direct","Proxy"]
    inp_hdrs=["PS/Growth","GM %","GM Erosion","ROIC %","Inv Days",
              "Inv Trend","Rule 40","Share Growth %","Pricing",
              "Rev Growth %","vs MA200 %","6M Return %"]
    all_hdrs=base_hdrs+inp_hdrs+gate_names
    def _banner(c1,c2,text,bg):
        ws.merge_cells(start_row=2,start_column=c1,end_row=2,end_column=c2)
        c=ws.cell(2,c1); c.value=text
        c.font=_font(9,C["white"],bold=True); c.fill=_fill(bg); c.alignment=_align("center")
    _banner(1,len(base_hdrs),"SCORING",C["hdr_dark"])
    _banner(len(base_hdrs)+1,len(base_hdrs)+len(inp_hdrs),"RAW INPUTS",C["hdr_mid"])
    if gate_names:
        _banner(len(base_hdrs)+len(inp_hdrs)+1,len(all_hdrs),
                "GATE RESULTS  [P]=proxy  N/A=excluded  partial=graduated","7F6000")
    _hdr(ws,3,all_hdrs); ws.row_dimensions[3].height=40
    for i,rec in enumerate(records,4):
        bg=C["gray_light"] if i%2==0 else C["white"]; col=1
        def wc(val,bg_=None,bold_=False,num_fmt=None,fg_="000000"):
            nonlocal col
            c=ws.cell(row=i,column=col,value=val)
            c.fill=_fill(bg_ or bg); c.font=_font(9,fg_,bold=bold_)
            c.alignment=_align("left" if col<=3 else "center"); c.border=_BORDER
            if num_fmt: c.number_format=num_fmt
            col+=1
        wc(rec["ticker"],bold_=True); wc(rec["sector"]); wc(rec["industry"])
        sc_bg=C["pass_cell"] if rec["weighted_score"]>=rec.get("pass_threshold",5.0) else C["fail_cell"]
        wc(rec["weighted_score"],bg_=sc_bg,num_fmt="0.00")
        wc(rec["max_score"],num_fmt="0.00")
        sp_bg=C["pass_cell"] if rec["strategy_passed"] else C["fail_cell"]
        wc("PASSED" if rec["strategy_passed"] else "FAILED",bg_=sp_bg,bold_=True)
        wc(rec["confidence"],bg_=CONF_BG.get(rec["confidence"],bg),bold_=True)
        pct_bg=C["pass_cell"] if rec["price_change"]>=0 else C["fail_cell"]
        wc(rec["price_change"]/100,bg_=pct_bg,num_fmt="+0.00%;-0.00%")
        obg,ofg=OUTCOME_BG.get(rec["outcome"],(bg,"000000"))
        wc(rec["outcome"],bg_=obg,bold_=True,fg_=ofg)
        wc(rec["gates_direct_passed"]); wc(rec["gates_proxy_passed"])
        ps=rec["ps_growth"]
        wc("N/A" if ps==999.0 else ps,num_fmt="0.00" if ps!=999.0 else None)
        wc(rec["gm_pct"],num_fmt="0.0"); wc(rec["gm_erosion"],num_fmt="+0.0;-0.0")
        wc(rec["roic_pct"],num_fmt="0.0"); wc(rec["inv_days"],num_fmt="0")
        wc(rec["inv_trend"],num_fmt="+0.0;-0.0"); wc(rec["rule40"],num_fmt="0.0")
        wc(rec["share_growth"],num_fmt="+0.00;-0.00")
        wc(rec["pricing_power"],bg_=C["pass_cell"] if rec["pricing_power"]=="Strong" else bg)
        wc(rec["rev_growth"],num_fmt="+0.0;-0.0")
        wc(rec["price_vs_ma200"]/100,num_fmt="+0.0%;-0.0%",
           bg_=C["pass_cell"] if rec["price_vs_ma200"]>=0 else C["fail_cell"])
        wc(rec["return_6m"]/100,num_fmt="+0.0%;-0.0%",
           bg_=C["pass_cell"] if rec["return_6m"]>=0 else C["fail_cell"])
        for gname in gate_names:
            g=rec["gates"].get(gname,{})
            cell_bg=_gate_cell_bg(g); label=_gate_label(g)
            c=ws.cell(row=i,column=col,value=label)
            c.fill=_fill(cell_bg); c.font=_font(9,bold=True)
            c.alignment=_align("center"); c.border=_BORDER
            score=g.get("score",0); note=g.get("note","")
            c.comment=Comment(f"Score: {score}\n{note}","GateTester V30"); col+=1
        ws.row_dimensions[i].height=15
    wmap={}
    for ci in range(1,len(all_hdrs)+1):
        ltr=get_column_letter(ci)
        if ci<=3: wmap[ltr]=20
        elif ci<=len(base_hdrs): wmap[ltr]=11
        elif ci<=len(base_hdrs)+len(inp_hdrs): wmap[ltr]=10
        else: wmap[ltr]=9
    _cw(ws,wmap)
    ws.auto_filter.ref=f"A3:{get_column_letter(len(all_hdrs))}3"

def _sheet_matrix(wb, records):
    ws=wb.create_sheet("Gate Matrix"); ws.sheet_properties.tabColor="ED7D31"; ws.freeze_panes="F3"
    if not records: return
    gate_names=list(records[0]["gates"].keys())
    hdrs=["Ticker","Industry","Score","/ Max","Passed","Return %"]+gate_names
    _hdr(ws,1,hdrs,bg=C["hdr_dark"]); ws.row_dimensions[1].height=36
    ws.merge_cells(f"A2:{get_column_letter(len(hdrs))}2")
    leg=ws["A2"]
    leg.value="Legend:  PASS=full  partial=graduated  [P]=proxy  N/A=excluded  FAIL=failed"
    leg.font=_font(8,C["hdr_dark"],italic=True); leg.fill=_fill("F2F2F2")
    leg.alignment=_align("left"); ws.row_dimensions[2].height=14
    for i,rec in enumerate(records,3):
        bg=C["gray_light"] if i%2==1 else C["white"]; col=1
        def wc(val,bg_=None,bold_=False,num_fmt=None):
            nonlocal col
            c=ws.cell(row=i,column=col,value=val)
            c.fill=_fill(bg_ or bg); c.font=_font(9,bold=bold_)
            c.alignment=_align("left" if col<=2 else "center"); c.border=_BORDER
            if num_fmt: c.number_format=num_fmt
            col+=1
        wc(rec["ticker"],bold_=True); wc(rec["industry"])
        sc_bg=C["pass_cell"] if rec["weighted_score"]>=rec.get("pass_threshold",5.0) else C["fail_cell"]
        wc(rec["weighted_score"],bg_=sc_bg,num_fmt="0.00")
        wc(rec["max_score"],num_fmt="0.00")
        sp_bg=C["pass_cell"] if rec["strategy_passed"] else C["fail_cell"]
        wc("PASSED" if rec["strategy_passed"] else "FAILED",bg_=sp_bg,bold_=True)
        pct_bg=C["pass_cell"] if rec["price_change"]>=0 else C["fail_cell"]
        wc(rec["price_change"]/100,bg_=pct_bg,num_fmt="+0.00%;-0.00%")
        for gname in gate_names:
            g=rec["gates"].get(gname,{})
            c=ws.cell(row=i,column=col,value=_gate_label(g))
            c.fill=_fill(_gate_cell_bg(g)); c.font=_font(9,bold=True)
            c.alignment=_align("center"); c.border=_BORDER; col+=1
        ws.row_dimensions[i].height=15
    wmap={"A":7,"B":22,"C":9,"D":8,"E":9,"F":10}
    for ci in range(7,len(hdrs)+1): wmap[get_column_letter(ci)]=9
    _cw(ws,wmap)
    ws.auto_filter.ref=f"A1:{get_column_letter(len(hdrs))}1"

def _sheet_raw(wb, records):
    ws=wb.create_sheet("Raw Data"); ws.sheet_properties.tabColor="7030A0"
    hdrs=["Ticker","Sector","Universe","Industry",
          "PS/Growth","GM %","GM Erosion","ROIC %","Inv Days","Inv Trend",
          "Rule 40","Share Growth %","Pricing","Rev Growth %",
          "vs MA200 %","6M Return %","Return %",
          "Score","Max Score","Threshold","Passed",
          "Confidence","Outcome","Direct Gates","Proxy Gates","Rescue Bonus",
          "Min Direct","Direct OK","Quality Kill","Val/Quality Block",
          "Expensive Block","Strong-Arm Block","MA200 Block","Rev Floor Block"]
    _hdr(ws,1,hdrs,bg=C["hdr_dark"]); ws.row_dimensions[1].height=20
    for i,rec in enumerate(records,2):
        bg=C["gray_light"] if i%2==0 else C["white"]
        vals=[rec["ticker"],rec["sector"],rec["universe"],rec["industry"],
              "N/A" if rec["ps_growth"]==999.0 else rec["ps_growth"],
              rec["gm_pct"],rec["gm_erosion"],rec["roic_pct"],
              rec["inv_days"],rec["inv_trend"],rec["rule40"],
              rec["share_growth"],rec["pricing_power"],rec["rev_growth"],
              rec["price_vs_ma200"]/100,rec["return_6m"]/100,rec["price_change"]/100,
              rec["weighted_score"],rec["max_score"],rec.get("pass_threshold",5.0),
              rec["strategy_passed"],rec["confidence"],rec["outcome"],
              rec["gates_direct_passed"],rec["gates_proxy_passed"],
              rec.get("rescue_bonus",0.0),rec.get("min_direct_required",2),
              rec.get("direct_gate_ok",True),rec.get("quality_killed",False),
              rec.get("val_quality_blocked",False),rec.get("expensive_blocked",False),
              rec.get("strong_arm_blocked",False),rec.get("ma200_blocked",False),
              rec.get("tech_rev_blocked",False)]
        for j,v in enumerate(vals,1):
            c=ws.cell(row=i,column=j,value=v)
            c.fill=_fill(bg); c.font=_font(9); c.alignment=_align("center"); c.border=_BORDER
            if j in (15,16,17): c.number_format="+0.00%;-0.00%"
        ws.row_dimensions[i].height=14
    for ci in range(1,len(hdrs)+1): ws.column_dimensions[get_column_letter(ci)].width=14
    ws.auto_filter.ref=f"A1:{get_column_letter(len(hdrs))}1"; ws.freeze_panes="E2"

def write_json(records, out_path, price_col_name=""):
    total=len(records)
    sw=sum(1 for r in records if r["outcome"]=="WIN")
    sl=sum(1 for r in records if r["outcome"]=="LOSS")
    cw=sum(1 for r in records if r["outcome"]=="WIN (contrarian)")
    cl=sum(1 for r in records if r["outcome"]=="LOSS (contrarian)")
    vf=sum(1 for r in records if r["outcome"]=="VETO-FAIL")
    payload={
        "meta":{
            "generated": datetime.now().isoformat(),
            "start_date": price_col_name.split("(")[-1].split("->")[0].strip() if "(" in price_col_name else "",
            "version": "V30", "nasdaq_regime": _ndx_regime,
            "pass_threshold_semi": pass_threshold("semi"),
            "pass_threshold_tech": pass_threshold("tech"),
            "pass_threshold_medtech": pass_threshold("medtech"),
            "pass_threshold_energy": pass_threshold("energy"),
            "tech_rev_floor": TECH_REV_FLOOR,
        },
        "summary":{
            "total":total,"strategy_wins":sw,"strategy_losses":sl,
            "contrarian_wins":cw,"contrarian_losses":cl,"veto_fails":vf,
            "accuracy_pct":round((sw+cw+vf)/total*100,1) if total else 0,
        },
        "stocks":records,
    }
    with open(out_path,"w",encoding="utf-8") as f:
        json.dump(payload,f,indent=2,ensure_ascii=False,default=str)
    print(f"  JSON saved   -> {out_path}")

def fetch_spy_return(start_col_name: str) -> "float | None":
    if not _YF_AVAILABLE: return None
    try:
        m = re.search(r'(\d{4}-\d{2}-\d{2})', start_col_name or "")
        if not m: return None
        start_str = m.group(1); start_dt = datetime.strptime(start_str, "%Y-%m-%d")
        end_dt = datetime.today()
        spy = yf.Ticker("SPY")
        hist = spy.history(start=start_dt.strftime("%Y-%m-%d"), end=end_dt.strftime("%Y-%m-%d"))
        if hist.empty or len(hist) < 2: return None
        p0 = float(hist["Close"].iloc[0]); p1 = float(hist["Close"].iloc[-1])
        return round((p1 / p0 - 1) * 100, 1)
    except Exception: return None

def print_portfolio_metrics(records, price_col_name: str = ""):
    positions = [r for r in records if r["strategy_passed"] and not r["veto"]]
    if not positions: print("  No strategy positions to analyze."); return
    returns = [r["price_change"] for r in positions]
    wins = [r for r in positions if r["price_change"] >= 0]
    losses = [r for r in positions if r["price_change"] < 0]
    n = len(positions); win_rate = len(wins)/n; loss_rate = len(losses)/n
    avg_win = (sum(r["price_change"] for r in wins)/len(wins)) if wins else 0.0
    avg_loss = (sum(r["price_change"] for r in losses)/len(losses)) if losses else 0.0
    expected_value = win_rate * avg_win + loss_rate * avg_loss
    total_gains = sum(r["price_change"] for r in wins) if wins else 0.0
    total_losses = abs(sum(r["price_change"] for r in losses)) if losses else 1e-9
    profit_factor = total_gains / total_losses
    avg_return = sum(returns) / n
    std_return = statistics.stdev(returns) if len(returns) > 1 else 0.0
    sharpe = avg_return / std_return if std_return > 0 else 0.0
    max_dd = abs(min(returns)) if returns else 0.0
    spy_return = fetch_spy_return(price_col_name)
    spy_str = f"{spy_return:+.1f}%" if spy_return is not None else "N/A"
    alpha_str = (f"{avg_return - spy_return:+.1f}%" if spy_return is not None else "N/A")
    conf_multipliers = {"HIGH": 1.5, "MED": 1.0, "LOW": 0.4}
    weighted_returns = []
    for r in positions:
        mult = conf_multipliers.get(r["confidence"], 1.0)
        weighted_returns.append(r["price_change"] * mult)
    w_denom = sum(conf_multipliers.get(r["confidence"], 1.0) for r in positions)
    w_avg_return = sum(weighted_returns) / w_denom if w_denom > 0 else 0.0
    if expected_value > 5.0 and profit_factor > 1.5:
        verdict = "ALPHA GENERATING"; verdict_color = GREEN
    elif expected_value > 0 and profit_factor >= 1.0:
        verdict = "COIN FLIP"; verdict_color = YELLOW
    else:
        verdict = "VALUE DESTROYING"; verdict_color = RED
    print(); print("="*78)
    print("  PORTFOLIO METRICS  (strategy positions only -- strategy_passed=True)")
    print("="*78)
    print(f"  Positions      : {n}  ({len(wins)} wins / {len(losses)} losses)")
    print(f"  Win Rate       : {win_rate*100:.1f}%  |  Loss Rate: {loss_rate*100:.1f}%")
    print(f"  Avg Win        : {avg_win:+.1f}%"); print(f"  Avg Loss       : {avg_loss:+.1f}%")
    print(f"  Expected Value : {expected_value:+.2f}% per position")
    print(f"  Profit Factor  : {profit_factor:.2f}x  (total gains / total losses)")
    print(f"  Sharpe Ratio   : {sharpe:.2f}  (vs 0% risk-free, cross-sectional)")
    print(f"  Worst Position : -{max_dd:.1f}%  (largest single BUY-signal loss)")
    print(f"  SPY Return     : {spy_str}  |  Strategy Avg: {avg_return:+.1f}%  |  Alpha: {alpha_str}")
    print(f"  Conf-Wtd Avg   : {w_avg_return:+.1f}%  (HIGH=1.5x MED=1.0x LOW=0.4x)")
    print(); print(f"  VERDICT: {verdict_color}{BOLD}{verdict}{RESET}"); print("="*78)

def print_sep(char="-",w=78): print(char*w)

def main():
    global _ndx_regime
    _log_file = open(LOG_PATH, "w", encoding="utf-8", errors="replace")
    _log_file.write(f"Gate Tester V30 Log - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    _log_file.write("V30: tech G8 override 0.3->0.5, G7 high-GM tightened gm>68+rev>15, G3 no bear adj\n\n")
    sys.stdout = _Tee(sys.__stdout__, _log_file)
    print("\n"+"="*78)
    print("  GATE TESTER  V30  --  tech G8 override 0.3->0.5, G7 high-GM tightened, G3 no bear adj")
    print("="*78)
    if not MULTI_CSV.exists(): print(f"\n  ERROR: CSV not found -> {MULTI_CSV}"); sys.exit(1)
    df=pd.read_csv(MULTI_CSV)
    _price_col_name = ""
    for col in df.columns:
        if any(k in col for k in ["Price_Change","price_change","Yearly","Price Change"]):
            _price_col_name = col; df.rename(columns={col:"_pct"},inplace=True); break
    if "_pct" not in df.columns:
        df["_pct"]=pd.to_numeric(df.iloc[:,-1],errors="coerce")
    if "MARKET_REGIME" in df.columns:
        regime_vals = df["MARKET_REGIME"].dropna()
        _ndx_regime = str(regime_vals.iloc[0]) if len(regime_vals) > 0 else "BULL_STRONG"
    elif "NDX_Above_MA100" in df.columns:
        ndx_vals = df["NDX_Above_MA100"].dropna()
        if len(ndx_vals) > 0:
            _ndx_regime = "BULL_STRONG" if float(ndx_vals.iloc[0]) == 1.0 else "BEAR_GRIND"
        else: _ndx_regime = "BULL_STRONG"
    else: _ndx_regime = "BULL_STRONG"
    if _ndx_regime in ("BULL_STRONG","BULL_WEAK"): regime_col = f"{GREEN}{_ndx_regime}{RESET}"
    else: regime_col = f"{RED}{_ndx_regime}{RESET}"
    print(f"  NASDAQ regime: {regime_col}  |  "
          f"Thresholds: Semi={pass_threshold('semi'):.1f}  Tech={pass_threshold('tech'):.1f}  "
          f"MedTech={pass_threshold('medtech'):.1f}  Energy={pass_threshold('energy'):.1f}")
    print(f"  Counter-cyclical: {COUNTER_CYCLICAL}")
    print(f"  Weights: G1 x{W_G1} | G8 x{W_G8} (tech: x{W_G8_TECH})")
    print(f"  Direct x{DIRECT_WEIGHT} | Proxy x{PROXY_WEIGHT} | N/A=0")
    print(f"  MinDirect: tech={MIN_DIRECT_GATES_TECH} medtech={MIN_DIRECT_GATES_MEDTECH} other={MIN_DIRECT_GATES_DEFAULT}")
    print(f"  V30: tech G8 override 0.3->0.5, G7 high-GM tightened gm>68+rev>15, G3 no bear adj")
    print("="*78)
    pct_ranks = compute_sector_percentiles(df)
    records=[]
    totals=dict(total=0,sw=0,sl=0,cw=0,cl=0,vf=0,skipped=0)
    universe_stats={u:dict(win=0,loss=0,cwin=0,closs=0,veto=0,n=0)
                    for u in ("energy","tech","medtech","semi")}
    for _,row in df.iterrows():
        ticker = str(row.get("Ticker","UNKNOWN"))
        sector = str(row.get("Sector",""))
        try: price_chg = sf(row["_pct"])
        except Exception: price_chg = 0.0
        sector_pct = pct_ranks.get(ticker, 50.0)
        try: universe, sub, gate_results = run_gates(row, sector_pct)
        except Exception as e:
            totals["skipped"]+=1; print(f"\n  {YELLOW}SKIP{RESET}  {ticker} -- gate error: {e}"); continue
        if gate_results is None:
            totals["skipped"]+=1; print(f"\n  {YELLOW}SKIP{RESET}  {ticker} -- sector '{sector}' not mapped"); continue
        try:
            vetoed, veto_reason = check_veto(row, sector_pct)
            rec=build_record(row,universe,sub,gate_results,price_chg,
                             veto=vetoed,veto_reason=veto_reason,sector_pct_rank=sector_pct)
        except Exception as e:
            totals["skipped"]+=1; print(f"\n  {YELLOW}SKIP{RESET}  {ticker} -- build_record error: {e}"); continue
        records.append(rec)
        w_score = rec["weighted_score"]; w_max = rec["max_score"]
        n_direct = rec["gates_direct_passed"]; n_proxy = rec["gates_proxy_passed"]
        strategy_ok = rec["strategy_passed"]; outcome = rec["outcome"]
        confidence = rec["confidence"]; rescue = rec.get("rescue_bonus",0.0)
        totals["total"]+=1; us=universe_stats[universe]; us["n"]+=1
        if outcome=="WIN":                totals["sw"]+=1; us["win"]+=1
        elif outcome=="LOSS":             totals["sl"]+=1; us["loss"]+=1
        elif outcome=="WIN (contrarian)": totals["cw"]+=1; us["cwin"]+=1
        elif outcome=="VETO-FAIL":        totals["vf"]+=1; us["veto"]+=1
        else:                             totals["cl"]+=1; us["closs"]+=1
        conf_col=GREEN if confidence=="HIGH" else (YELLOW if confidence=="MED" else RED)
        out_col=GREEN if "WIN" in outcome else RED
        thr_used = pass_threshold(universe); min_dg = min_direct_gates(universe)
        print(); print_sep()
        print(f" {BOLD}{CYAN}{ticker}{RESET}  |  {sector}  ->  "
              f"{BOLD}{SUB_LABELS.get(sub,sub)}{RESET}  | sector pct={sector_pct:.0f}%")
        print_sep()
        if vetoed: print(f"  {RED}{BOLD}*** VETOED ***  {veto_reason}{RESET}")
        rescue_note = f"  (includes +{rescue:.1f} momentum rescue)" if rescue > 0 else ""
        direct_note = ""
        if not rec.get("direct_gate_ok",True):
            direct_note = f"  {RED}[BLOCKED: need {min_dg} direct gates, have {n_direct}]{RESET}"
        quality_note = ""
        if rec.get("quality_killed",False):
            quality_note = f"  {RED}[{rec.get('quality_reason','')}]{RESET}"
        val_quality_note = ""
        if rec.get("val_quality_blocked",False):
            val_quality_note = f"  {RED}[{rec.get('val_quality_note','')}]{RESET}"
        expensive_note = ""
        if rec.get("expensive_blocked",False):
            expensive_note = f"  {RED}[{rec.get('expensive_note','')}]{RESET}"
        strong_arm_note = ""
        if rec.get("strong_arm_blocked",False):
            strong_arm_note = f"  {RED}[{rec.get('strong_arm_note','')}]{RESET}"
        ma200_note = ""
        if rec.get("ma200_blocked",False):
            ma200_note = f"  {RED}[{rec.get('ma200_note','')}]{RESET}"
        rev_floor_note = ""
        if rec.get("tech_rev_blocked",False):
            rev_floor_note = f"  {RED}[{rec.get('tech_rev_note','')}]{RESET}"
        print(f"  Score: {BOLD}{w_score:.2f}{RESET}/{thr_used:.1f}  (max={w_max:.1f}){rescue_note}  |  "
              f"Strategy: {'PASSED' if strategy_ok else 'FAILED'}  |  "
              f"Return: {BOLD}{price_chg:+.2f}%{RESET}")
        if direct_note: print(direct_note)
        if quality_note: print(quality_note)
        if val_quality_note: print(val_quality_note)
        if expensive_note: print(expensive_note)
        if strong_arm_note: print(strong_arm_note)
        if ma200_note: print(ma200_note)
        if rev_floor_note: print(rev_floor_note)
        print(f"  Outcome: {out_col}{BOLD}{outcome}{RESET}   |  "
              f"{conf_col}{confidence} confidence{RESET}  |  "
              f"direct={n_direct}  proxy={n_proxy}")
        print()
        for gname,(passed,weight,note,is_proxy,score_override) in gate_results.items():
            actual=score_override if score_override is not None else (weight if passed else 0.0)
            if weight==NA_WEIGHT: icon=f"{YELLOW}N/A{RESET}"; color=YELLOW
            elif passed: icon=f"{GREEN}PASS{RESET}"; color=GREEN
            else: icon=f"{RED}FAIL{RESET}"; color=RED
            proxy_tag=f"{BLUE}[P]{RESET}" if is_proxy else "   "
            mom_tag=f"{MAGENTA}[G8]{RESET}" if gname=="G8 Momentum" else "    "
            clean = re.sub(r'\s*\[(?:P|N/A|proxy|direct)\s*[^\]]*\]', '', note)
            print(f"   {icon} {proxy_tag}{mom_tag} {color}{gname:<26}{RESET}  (+{actual:.2f})  {clean}")
        failed=[g for g,(ok,w,*_) in gate_results.items() if not ok and w!=NA_WEIGHT]
        na_gates=[g for g,(_,w,*_) in gate_results.items() if w==NA_WEIGHT]
        if failed: print(f"\n   {RED}Failed:  {', '.join(failed)}{RESET}")
        if na_gates: print(f"   {YELLOW}N/A:     {', '.join(na_gates)}{RESET}")
    total=totals["total"]; sw,sl=totals["sw"],totals["sl"]
    cw,cl=totals["cw"],totals["cl"]; vf=totals["vf"]
    correct=sw+cw+vf; pct=(correct/total*100) if total else 0
    print(); print("="*78)
    print(f"  SUMMARY -- V30  (NASDAQ regime: {_ndx_regime.upper()}  |  "
          f"threshold: Semi={pass_threshold('semi'):.1f}  Tech={pass_threshold('tech'):.1f}  "
          f"MedTech={pass_threshold('medtech'):.1f}  Energy={pass_threshold('energy'):.1f})")
    print("="*78)
    print(f"  Total tested: {total}   Skipped: {totals['skipped']}  "
          f"Thresholds: Semi={pass_threshold('semi'):.1f} | Tech={pass_threshold('tech'):.1f} | "
          f"MedTech={pass_threshold('medtech'):.1f} | Energy={pass_threshold('energy'):.1f}")
    print()
    print(f"  {GREEN}Strategy WINS    (score>=threshold + positive):{RESET}  {sw}")
    print(f"  {RED}Strategy LOSSES   (score>=threshold + negative):{RESET}  {sl}")
    print()
    print(f"  {GREEN}Contrarian WINS   (score<threshold + negative):{RESET}  {cw}")
    print(f"  {RED}Contrarian LOSSES  (score<threshold + positive):{RESET}  {cl}")
    print(f"  {RED}VETO-FAIL (GM erosion kill-switch):{RESET}              {vf}")
    print()
    print(f"  {BOLD}Overall accuracy: {correct}/{total}  ({pct:.1f}%){RESET}")
    print()
    print(f"  {'-'*56}"); print("  Per-universe:"); print(f"  {'-'*56}")
    for u,s in universe_stats.items():
        if s["n"]==0: continue
        uc=(s["win"]+s["cwin"]+s["veto"]); upc=(uc/s["n"]*100) if s["n"] else 0
        thr_u = pass_threshold(u)
        print(f"  {BOLD}{u.upper():<10}{RESET}  n={s['n']:>3}  thr={thr_u:.1f}  "
              f"W={s['win']}  L={s['loss']}  VETO={s['veto']}  "
              f"C-W={s['cwin']}  C-L={s['closs']}  Accuracy={uc}/{s['n']} ({upc:.0f}%)")
    print("="*78)
    print(f"\n  {BLUE}[P]{RESET}=proxy x{PROXY_WEIGHT}  "
          f"{YELLOW}N/A{RESET}=anti-double-count  "
          f"{MAGENTA}[G8]{RESET}=momentum gate  partial=graduated score")
    print_portfolio_metrics(records, price_col_name=_price_col_name)
    xlsx=OUTPUT_DIR/"gate_report_latest.xlsx"; jfile=OUTPUT_DIR/"gate_report_latest.json"
    print(); print("-"*78)
    write_excel(records,xlsx); write_json(records,jfile,price_col_name=_price_col_name)
    print(f"  Log saved    -> {LOG_PATH}"); print("-"*78); print()

if __name__=="__main__":
    main()